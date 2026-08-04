import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.build_company_source_registry import (
    CATEGORY_REGISTRIES,
    build_workbook,
    verify_workbook,
)
from scripts.linkedin_export_workbook import RegistryEntry, load_linkedin_export


def _write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


class LinkedInExportWorkbookTests(unittest.TestCase):
    def _create_export(self, root: Path) -> None:
        connections = root / "Connections.csv"
        connections.write_text(
            "Notes:\nExported connection data\n\n",
            encoding="utf-8",
        )
        with connections.open("a", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "First Name",
                    "Last Name",
                    "URL",
                    "Email Address",
                    "Company",
                    "Position",
                    "Connected On",
                ]
            )
            writer.writerow(
                [
                    "Example",
                    "Person",
                    "https://www.linkedin.com/in/example-person",
                    "example@example.com",
                    "Wipro Limited",
                    "Data Scientist",
                    "19 Jul 2026",
                ]
            )
            writer.writerow(
                [
                    "",
                    "",
                    "https://www.linkedin.com/in/unnamed-example",
                    "",
                    "Example Private Company",
                    "Engineer",
                    "18 Jul 2026",
                ]
            )

        _write_csv(
            root / "Jobs" / "Job Applications.csv",
            [
                "Application Date",
                "Contact Email",
                "Contact Phone Number",
                "Company Name",
                "Job Title",
                "Job Url",
                "Resume Name",
                "Question And Answers",
            ],
            [
                [
                    "07/19/26, 1:30 PM",
                    "private@example.com",
                    "+910000000000",
                    "Wipro",
                    "Machine Learning Engineer",
                    "https://www.linkedin.com/jobs/view/123",
                    "Resume.docx",
                    "Sensitive answer",
                ]
            ],
        )
        _write_csv(
            root / "Jobs" / "Saved Jobs.csv",
            ["Saved Date", "Job Url", "Job Title", "Company Name"],
            [
                [
                    "7/18/26, 11:15 AM",
                    "https://www.linkedin.com/jobs/view/456",
                    "AI Engineer",
                    "Wipro",
                ]
            ],
        )
        _write_csv(
            root / "SavedJobAlerts.csv",
            ["ALERT_PARAMETERS", "QUERY_CONTEXT", "SAVED_SEARCH_ID"],
            [
                [
                    "{frequency=DAILY,channels=[INAPP_NOTIFICATION, EMAIL],smartExpansionEnabled=true}",
                    (
                        "{keywords=machine learning,searchLocation={geoLocations=[{geo=urn:li:geo:12345,"
                        "radiusInKms=40}]},radiusInKms=40,workplaceTypes={selectedValues=["
                        "urn:li:workplaceType:2,urn:li:workplaceType:3]},sortingType=RELEVANCE}"
                    ),
                    "alert-1",
                ]
            ],
        )
        _write_csv(
            root / "Company Follows.csv",
            ["Organization", "Followed On"],
            [["Wipro", "Sun Jul 19 10:20:30 UTC 2026"]],
        )
        _write_csv(
            root / "Profile.csv",
            [
                "First Name",
                "Last Name",
                "Maiden Name",
                "Address",
                "Birth Date",
                "Headline",
                "Summary",
                "Industry",
                "Zip Code",
                "Geo Location",
                "Twitter Handles",
                "Websites",
                "Instant Messengers",
            ],
            [
                [
                    "Private",
                    "Person",
                    "",
                    "Private address",
                    "1990-01-01",
                    "Machine Learning Engineer",
                    "Professional summary",
                    "Technology",
                    "500002",
                    "500002, Hyderabad, Telangana, India",
                    "",
                    "https://portfolio.example.com",
                    "",
                ]
            ],
        )
        _write_csv(
            root / "Jobs" / "Job Seeker Preferences.csv",
            ["Locations", "Preferred Job Types", "Job Titles", "Open To Recruiters"],
            [["500002", "Full-time", "Machine Learning Engineer", "Yes"]],
        )
        _write_csv(
            root / "Positions.csv",
            ["Company Name", "Title", "Description", "Location", "Started On", "Finished On"],
            [["Wipro", "Engineer", "Built ML systems", "India", "Jun 2021", ""]],
        )
        _write_csv(
            root / "Education.csv",
            ["School Name", "Start Date", "End Date", "Notes", "Degree Name", "Activities"],
            [["Example University", "2018", "2020", "", "MTech", ""]],
        )
        _write_csv(root / "Skills.csv", ["Name"], [["Machine Learning"], ["Python"]])

    def test_private_export_is_normalized_into_job_search_tabs(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self._create_export(root)
            registry_entries = [
                RegistryEntry(
                    category=category,
                    company=company.company,
                    careers_url=company.careers_url,
                    jobs_url=company.jobs_url,
                )
                for category, companies in CATEGORY_REGISTRIES.items()
                for company in companies
            ]
            data = load_linkedin_export(root, registry_entries)

            self.assertEqual(len(data.connections), 2)
            self.assertEqual(data.connections[0]["registry_company"], "Wipro")
            self.assertEqual(data.connections[1]["name"], "Unnamed LinkedIn connection")
            self.assertEqual(data.profile["Geo Location"], "Hyderabad, Telangana, India")
            self.assertNotIn("Locations", data.preferences)
            self.assertNotIn("contact_email", data.applications[0])
            self.assertNotIn("contact_phone", data.applications[0])
            self.assertEqual(data.job_alerts[0]["geo_identifier"], "urn:li:geo:12345")
            self.assertEqual(data.job_alerts[0]["workplace_types"], "Remote, Hybrid")

            output = root / "Company_Source_Registry.xlsx"
            build_workbook(output, linkedin_data=data)
            verify_workbook(output, linkedin_data=data)

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(len(workbook.sheetnames), 14)
            pilot = workbook["Portal Alert Pilot"]
            referral_counts = {
                pilot.cell(row, 1).value: pilot.cell(row, 3).value
                for row in range(5, 10)
            }
            self.assertEqual(referral_counts["Wipro"], 1)
            connections = workbook["LinkedIn Connections"]
            self.assertEqual(connections.max_row, 6)
            self.assertIsNotNone(connections["A5"].hyperlink)
            self.assertIsNotNone(connections["G5"].hyperlink)
            applications = workbook["LinkedIn Applications"]
            application_headers = [
                applications.cell(4, column).value
                for column in range(1, applications.max_column + 1)
            ]
            self.assertNotIn("Contact Email", application_headers)
            self.assertNotIn("Contact Phone", application_headers)
            self.assertNotIn("Question And Answers", application_headers)
            profile_values = {
                profile_cell.value
                for row in workbook["LinkedIn Profile"].iter_rows()
                for profile_cell in row
                if profile_cell.value
            }
            self.assertNotIn("Private address", profile_values)
            self.assertNotIn("1990-01-01", profile_values)


if __name__ == "__main__":
    unittest.main()
