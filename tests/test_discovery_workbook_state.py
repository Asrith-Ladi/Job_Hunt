import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from job_hunt.discovery.models import DiscoveryFilters, DiscoveryJob, SourceCheck
from job_hunt.discovery.state import (
    select_new_or_changed_jobs,
    update_discovery_state,
    update_user_fields,
)
from job_hunt.discovery.workbook import (
    CHECKS_SHEET_NAME,
    DISCOVERY_JOB_COLUMNS,
    JOBS_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    read_discovery_workbook,
    validate_discovery_rows,
    verify_discovery_workbook,
    write_discovery_workbook,
)


def _row(
    description="Build ML systems. 5-8 years of experience.",
    discovered_at="2026-08-01T10:00:00+05:30",
):
    row = DiscoveryJob.create(
        company="Example",
        title="Machine Learning Engineer",
        location="Hyderabad",
        provider="greenhouse",
        source_identifier="example",
        source_type="official_public_api",
        external_job_id="job-1",
        official_url="https://boards.greenhouse.io/example/jobs/1",
        apply_url="https://boards.greenhouse.io/example/jobs/1",
        source_url="https://boards-api.greenhouse.io/v1/boards/example/jobs?content=true",
        description=description,
        posted_at="2026-07-31T08:00:00+00:00",
        date_provenance="first_published",
        discovered_at=discovered_at,
        filters=DiscoveryFilters(),
    ).to_dict()
    row["run_change_status"] = "new"
    return row


def _check():
    return SourceCheck(
        company="Example",
        category="Product Companies",
        provider="greenhouse",
        source_identifier="example",
        strategy="official_public_api",
        source_url="https://boards-api.greenhouse.io/v1/boards/example/jobs?content=true",
        status="success",
        jobs_found=1,
        jobs_exported=1,
        warning="",
        fallback="hosted board",
        checked_at="2026-08-01T10:00:00+05:30",
    ).to_dict()


class DiscoveryWorkbookStateTests(unittest.TestCase):
    def test_workbook_round_trip_has_three_stable_safe_tables(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "company_portals.xlsx"
            row = _row()
            row["notes"] = "=formula text must stay text"
            summary = {
                "run_id": "company_portals-test",
                "mode": "company_portals",
                "sources_checked": 1,
                "jobs_found": 1,
                "jobs_unchanged_from_prior_runs": 0,
                "warnings": 0,
                "keyword_filter": "=unsafe summary formula text",
            }
            write_discovery_workbook(
                output,
                mode="company_portals",
                rows=[row],
                source_checks=[_check()],
                summary=summary,
                run_started_at=datetime(2026, 8, 1, 10, 0),
            )
            verify_discovery_workbook(output, expected_jobs=1, expected_checks=1)

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [JOBS_SHEET_NAME, CHECKS_SHEET_NAME, SUMMARY_SHEET_NAME],
            )
            jobs = workbook[JOBS_SHEET_NAME]
            self.assertEqual(jobs.freeze_panes, "B5")
            self.assertIsNone(jobs.auto_filter.ref)
            for column in ["official_url", "apply_url", "source_url"]:
                cell = jobs.cell(5, DISCOVERY_JOB_COLUMNS.index(column) + 1)
                self.assertIsNotNone(cell.hyperlink)
            notes = jobs.cell(5, DISCOVERY_JOB_COLUMNS.index("notes") + 1)
            self.assertNotEqual(notes.data_type, "f")

            rows, checks, loaded_summary = read_discovery_workbook(output)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["run_change_status"], "new")
            self.assertEqual(len(checks), 1)
            self.assertEqual(loaded_summary["run_id"], "company_portals-test")

            edited = dict(rows[0])
            edited["application_status"] = "reviewing"
            edited["notes"] = "Review tomorrow"
            self.assertEqual(
                validate_discovery_rows([edited], expected_rows=rows)[0]["notes"],
                "Review tomorrow",
            )
            changed_title = dict(edited)
            changed_title["title"] = "Changed protected title"
            with self.assertRaises(ValueError):
                validate_discovery_rows([changed_title], expected_rows=rows)

            workbook = load_workbook(output)
            legacy_jobs = workbook[JOBS_SHEET_NAME]
            legacy_jobs.delete_cols(DISCOVERY_JOB_COLUMNS.index("run_change_status") + 1)
            workbook.save(output)
            legacy_rows, _, _ = read_discovery_workbook(output)
            self.assertEqual(legacy_rows[0]["run_change_status"], "new_or_changed")

    def test_seen_state_preserves_first_seen_and_user_fields(self):
        first = _row(discovered_at="2026-08-01T10:00:00+05:30")
        selected, unchanged = select_new_or_changed_jobs([first], None)
        self.assertEqual(len(selected), 1)
        self.assertEqual(selected[0]["run_change_status"], "new")
        self.assertEqual(unchanged, 0)
        state = update_discovery_state(
            None,
            [first],
            completed_at="2026-08-01T10:01:00+05:30",
        )
        first["application_status"] = "shortlisted"
        first["notes"] = "Strong match"
        state = update_user_fields(state, [first])

        second = _row(discovered_at="2026-08-02T10:00:00+05:30")
        selected, unchanged = select_new_or_changed_jobs([second], state)
        self.assertEqual(selected, [])
        self.assertEqual(unchanged, 1)

        changed = _row(
            description="Build production ML systems. 5-8 years of experience.",
            discovered_at="2026-08-02T10:00:00+05:30",
        )
        selected, unchanged = select_new_or_changed_jobs([changed], state)
        self.assertEqual(unchanged, 0)
        self.assertEqual(selected[0]["run_change_status"], "changed")
        self.assertEqual(selected[0]["first_seen_at"], "2026-08-01T10:00:00+05:30")
        self.assertEqual(selected[0]["last_seen_at"], "2026-08-02T10:00:00+05:30")
        self.assertEqual(selected[0]["application_status"], "shortlisted")
        self.assertEqual(selected[0]["notes"], "Strong match")


if __name__ == "__main__":
    unittest.main()
