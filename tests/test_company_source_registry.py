import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.build_company_source_registry import (
    CATEGORY_REGISTRIES,
    PORTAL_ALERT_HEADERS,
    PORTAL_ALERT_SHEET_NAME,
    PORTAL_ALERT_TABLE_NAME,
    TABLE_NAMES,
    build_workbook,
    verify_workbook,
)


class CompanySourceRegistryTests(unittest.TestCase):
    def test_registry_is_valid_and_categories_are_mutually_exclusive(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "Company_Source_Registry.xlsx"

            build_workbook(output)
            verify_workbook(output)

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Coverage", *CATEGORY_REGISTRIES, PORTAL_ALERT_SHEET_NAME],
            )

            seen: set[str] = set()
            for sheet_name, companies in CATEGORY_REGISTRIES.items():
                sheet = workbook[sheet_name]
                self.assertEqual(sheet.max_row, len(companies) + 4)
                self.assertEqual(sheet.freeze_panes, "A5")
                self.assertIsNone(sheet.auto_filter.ref)
                self.assertIn(TABLE_NAMES[sheet_name], sheet.tables)
                formatting = {
                    str(key.sqref): [
                        formula
                        for rule in rules
                        for formula in (rule.formula or [])
                    ]
                    for key, rules in sheet.conditional_formatting._cf_rules.items()
                }
                row_range = f"A5:O{sheet.max_row}"
                self.assertIn(row_range, formatting)
                self.assertIn(
                    'LEFT($M5,12)="Inaccessible"',
                    formatting[row_range],
                )
                self.assertIn(
                    'LEFT($M5,15)="Manual required"',
                    formatting[row_range],
                )
                for row in range(5, sheet.max_row + 1):
                    company = str(sheet.cell(row, 1).value)
                    normalized = company.casefold()
                    self.assertNotIn(normalized, seen)
                    seen.add(normalized)
                    self.assertIsNotNone(sheet.cell(row, 4).hyperlink)
                    self.assertIsNotNone(sheet.cell(row, 5).hyperlink)

            self.assertEqual(len(seen), 246)
            coverage = workbook["Coverage"]
            self.assertEqual(coverage.freeze_panes, "A6")
            self.assertEqual(coverage.auto_filter.ref, "A5:G10")
            self.assertEqual(coverage["C12"].value, "=SUM(C6:C10)")

            pilot = workbook[PORTAL_ALERT_SHEET_NAME]
            self.assertEqual(pilot.freeze_panes, "D5")
            self.assertIsNone(pilot.auto_filter.ref)
            self.assertIn(PORTAL_ALERT_TABLE_NAME, pilot.tables)
            self.assertEqual(
                [pilot.cell(4, column).value for column in range(1, pilot.max_column + 1)],
                PORTAL_ALERT_HEADERS,
            )
            companies = [pilot.cell(row, 1).value for row in range(5, 10)]
            self.assertEqual(
                companies,
                ["Wipro", "Cognizant", "Infosys", "Accenture", "Google"],
            )
            self.assertEqual(pilot["M5"].value, "Pending browser/account")
            self.assertEqual(pilot["M9"].value, "Pending browser/Google sign-in")
            self.assertTrue(all(pilot.cell(row, 13).value != "Enabled" for row in range(5, 10)))
            self.assertIsNotNone(pilot["D5"].hyperlink)
            self.assertIsNotNone(pilot["I5"].hyperlink)
            self.assertIsNotNone(pilot["S9"].hyperlink)

    def test_expected_category_sizes_and_representative_adapters(self):
        expected_counts = {
            "MNC": 65,
            "Product Companies": 87,
            "Startups": 37,
            "Mid-Sized Companies": 37,
            "Other Companies": 20,
        }
        self.assertEqual(
            {name: len(items) for name, items in CATEGORY_REGISTRIES.items()},
            expected_counts,
        )

        rows = {
            category: {company.company: company for company in companies}
            for category, companies in CATEGORY_REGISTRIES.items()
        }
        self.assertEqual(rows["MNC"]["Wipro"].source_type, "SAP SuccessFactors")
        self.assertIn(
            "boards-api.greenhouse.io",
            rows["Product Companies"]["Stripe"].public_endpoint,
        )
        self.assertEqual(rows["Startups"]["Groww"].source_type, "Greenhouse")
        self.assertIn("greenhouse.io", rows["Startups"]["Groww"].jobs_url)
        self.assertEqual(rows["Product Companies"]["Paytm"].source_type, "Lever")
        self.assertEqual(
            rows["Product Companies"]["IndiaMART"].source_type,
            "SmartRecruiters",
        )
        self.assertEqual(rows["Startups"]["Sprinto"].source_type, "Lever")
        self.assertEqual(
            rows["Mid-Sized Companies"]["Mindtickle"].source_type,
            "SmartRecruiters",
        )
        self.assertEqual(
            rows["Mid-Sized Companies"]["Redis"].source_type,
            "Company-hosted jobs portal",
        )
        self.assertIn(
            "redis.io/company/careers/current-job-openings",
            rows["Mid-Sized Companies"]["Redis"].jobs_url,
        )
        self.assertIn("Cohesity", rows["Mid-Sized Companies"])
        self.assertIn("Fractal Analytics", rows["Other Companies"])

    def test_all_rows_use_https_and_have_one_primary_category(self):
        assignments: dict[str, str] = {}
        for category, companies in CATEGORY_REGISTRIES.items():
            for company in companies:
                key = company.company.casefold().strip()
                self.assertNotIn(key, assignments)
                assignments[key] = category
                self.assertTrue(company.careers_url.startswith("https://"))
                self.assertTrue(company.jobs_url.startswith("https://"))
        self.assertEqual(len(assignments), 246)


if __name__ == "__main__":
    unittest.main()
