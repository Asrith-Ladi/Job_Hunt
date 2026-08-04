import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from job_hunt.gmail_workbook import (
    GMAIL_SHEET_NAME,
    GMAIL_RUN_COLUMNS,
    GMAIL_TABLE_NAME,
    SUMMARY_SHEET_NAME,
    read_gmail_run_workbook,
    validate_editor_rows,
    verify_gmail_run_workbook,
    write_gmail_run_workbook,
)
from job_hunt.integrations.sheets import JOB_COLUMNS


def _row(record_id="job-1"):
    values = {column: "" for column in GMAIL_RUN_COLUMNS}
    values.update(
        {
            "job_record_id": record_id,
            "owner_id": "personal",
            "alert_source": "linkedin",
            "gmail_message_id": "message-1",
            "email_subject": "AI role",
            "email_received_at": "2026-08-01T08:30:00+00:00",
            "company": "Example Company",
            "title": "Machine Learning Engineer",
            "location": "Hyderabad",
            "years_of_experience": "5-8 years",
            "source_url": "https://linkedin.com/jobs/view/123?utm_source=email",
            "official_url": "https://careers.example.com/jobs/123",
            "first_seen_at": "2026-08-01T08:31:00+00:00",
            "last_seen_at": "2026-08-01T08:31:00+00:00",
            "parse_confidence": "high",
            "parse_status": "parsed",
            "company_match": "not_configured",
            "application_status": "not_started",
            "notes": "=unsafe formula text",
            "evidence_message_ids": "message-1",
            "experience_min_years": 5.0,
            "experience_max_years": 8.0,
            "experience_fit": "inside_target",
            "experience_source": "email",
            "referral_count": 2,
            "referral_name": "Alex Example",
            "referral_position": "Talent Acquisition Partner",
            "referral_profile_url": "https://www.linkedin.com/in/alex-example",
            "referral_match_status": "offline_company_match_unverified",
            "referral_eligibility": "Preliminary alert-only fit.",
            "referral_message": (
                "Hi Alex,\n\nPlease refer me.\n\nJob: https://careers.example.com/jobs/123"
            ),
        }
    )
    return values


class GmailRunWorkbookTests(unittest.TestCase):
    def test_run_workbook_is_editable_safe_and_round_trips(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "gmail_alerts_2026-08-01_140000.xlsx"
            summary = {
                "run_id": "run-1",
                "started_at": "2026-08-01T14:00:00+05:30",
                "finished_at": "2026-08-01T14:01:00+05:30",
                "status": "completed",
                "messages_read": 2,
                "messages_supported": 2,
                "jobs_parsed": 3,
                "jobs_after_deduplication": 2,
                "jobs_filtered_out": 0,
                "parsing_warnings": 0,
            }
            rows = [_row("job-1"), _row("job-2")]

            write_gmail_run_workbook(
                output,
                rows,
                summary,
                run_started_at=datetime(2026, 8, 1, 14, 0),
            )
            verify_gmail_run_workbook(output, expected_rows=2)

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(workbook.sheetnames, [GMAIL_SHEET_NAME, SUMMARY_SHEET_NAME])
            sheet = workbook[GMAIL_SHEET_NAME]
            self.assertEqual(sheet.freeze_panes, "B5")
            self.assertIsNone(sheet.auto_filter.ref)
            final_letter = get_column_letter(len(GMAIL_RUN_COLUMNS))
            self.assertEqual(sheet.tables[GMAIL_TABLE_NAME].ref, f"A4:{final_letter}6")
            self.assertIsNotNone(sheet["L5"].hyperlink)
            self.assertIsNotNone(sheet["M5"].hyperlink)
            self.assertNotEqual(sheet["T5"].data_type, "f")
            referral_name = sheet.cell(5, GMAIL_RUN_COLUMNS.index("referral_name") + 1)
            referral_profile = sheet.cell(
                5,
                GMAIL_RUN_COLUMNS.index("referral_profile_url") + 1,
            )
            referral_message = sheet.cell(
                5,
                GMAIL_RUN_COLUMNS.index("referral_message") + 1,
            )
            self.assertEqual(
                referral_name.hyperlink.target,
                "https://www.linkedin.com/in/alex-example",
            )
            self.assertIsNotNone(referral_profile.hyperlink)
            self.assertEqual(
                referral_message.hyperlink.target,
                "https://careers.example.com/jobs/123",
            )

            loaded_rows, loaded_summary = read_gmail_run_workbook(output)
            self.assertEqual(len(loaded_rows), 2)
            self.assertEqual(loaded_rows[0]["job_record_id"], "job-1")
            self.assertEqual(loaded_rows[0]["experience_min_years"], 5)
            self.assertEqual(loaded_summary["run_id"], "run-1")

    def test_editor_rows_cannot_add_remove_or_duplicate_record_ids(self):
        rows = [_row("job-1"), _row("job-2")]
        validated = validate_editor_rows(
            rows,
            expected_record_ids=["job-1", "job-2"],
        )
        self.assertEqual(len(validated), 2)

        with self.assertRaises(ValueError):
            validate_editor_rows(rows[:1], expected_record_ids=["job-1", "job-2"])
        with self.assertRaises(ValueError):
            validate_editor_rows(
                [_row("job-1"), _row("job-1")],
                expected_record_ids=["job-1", "job-2"],
            )

    def test_empty_incremental_run_still_creates_a_valid_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "gmail_alerts_empty.xlsx"
            write_gmail_run_workbook(
                output,
                [],
                {"run_id": "empty", "messages_read": 0},
                run_started_at=datetime(2026, 8, 1, 15, 0),
            )
            verify_gmail_run_workbook(output, expected_rows=0)
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook[GMAIL_SHEET_NAME].tables[GMAIL_TABLE_NAME].ref,
                f"A4:{get_column_letter(len(GMAIL_RUN_COLUMNS))}4",
            )

    def test_legacy_gmail_workbook_loads_with_blank_referral_fields(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "legacy_gmail.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = GMAIL_SHEET_NAME
            for column_number, column in enumerate(JOB_COLUMNS, start=1):
                sheet.cell(4, column_number, column)
                sheet.cell(5, column_number, _row()[column])
            summary = workbook.create_sheet(SUMMARY_SHEET_NAME)
            summary.cell(2, 1, "Metric")
            summary.cell(2, 2, "Value")
            summary.cell(3, 1, "run_id")
            summary.cell(3, 2, "legacy")
            workbook.save(output)

            rows, loaded_summary = read_gmail_run_workbook(output)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["job_record_id"], "job-1")
        self.assertEqual(rows[0]["referral_name"], "")
        self.assertEqual(rows[0]["referral_message"], "")
        self.assertEqual(loaded_summary["run_id"], "legacy")


if __name__ == "__main__":
    unittest.main()
