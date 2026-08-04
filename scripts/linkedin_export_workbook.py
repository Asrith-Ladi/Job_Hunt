"""Import job-search-useful LinkedIn export data into the company registry workbook."""

from __future__ import annotations

import csv
import io
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


LINKEDIN_SHEET_NAMES = [
    "LinkedIn Overview",
    "LinkedIn Profile",
    "LinkedIn Connections",
    "LinkedIn Applications",
    "LinkedIn Saved Jobs",
    "LinkedIn Job Alerts",
    "Followed Companies",
]

LINKEDIN_TABLE_NAMES = {
    "LinkedIn Connections": "LinkedInConnections",
    "LinkedIn Applications": "LinkedInApplications",
    "LinkedIn Saved Jobs": "LinkedInSavedJobs",
    "LinkedIn Job Alerts": "LinkedInJobAlerts",
    "Followed Companies": "LinkedInFollowedCompanies",
}

TITLE_FILL = "0F766E"
NOTE_FILL = "E2E8F0"
PROCESS_FILL = "CCFBF1"
HEADER_FILL = "1E3A5F"
LIGHT_BORDER = "CBD5E1"


@dataclass(frozen=True)
class RegistryEntry:
    category: str
    company: str
    careers_url: str
    jobs_url: str


@dataclass
class LinkedInExportData:
    source_label: str
    connections: list[dict[str, object]]
    applications: list[dict[str, object]]
    saved_jobs: list[dict[str, object]]
    job_alerts: list[dict[str, object]]
    followed_companies: list[dict[str, object]]
    profile: dict[str, str]
    preferences: dict[str, str]
    positions: list[dict[str, str]]
    education: list[dict[str, str]]
    skills: list[str]
    category_stats: list[dict[str, object]]
    top_referrals: list[dict[str, object]]
    registry_companies_reached: int


class _ExportReader:
    def __init__(self, source: Path) -> None:
        self.source = source
        self._archive: zipfile.ZipFile | None = None
        if source.is_file():
            if not zipfile.is_zipfile(source):
                raise ValueError(f"LinkedIn export is not a ZIP file: {source}")
            self._archive = zipfile.ZipFile(source)
        elif not source.is_dir():
            raise FileNotFoundError(f"LinkedIn export was not found: {source}")

    def close(self) -> None:
        if self._archive is not None:
            self._archive.close()

    def read_text(self, relative_path: str) -> str:
        normalized = relative_path.replace("\\", "/")
        if self._archive is not None:
            candidates = {
                name.replace("\\", "/"): name for name in self._archive.namelist()
            }
            archive_name = candidates.get(normalized)
            if archive_name is None:
                raise FileNotFoundError(f"Missing LinkedIn export file: {relative_path}")
            return self._archive.read(archive_name).decode("utf-8-sig")
        path = self.source / Path(relative_path)
        if not path.is_file():
            raise FileNotFoundError(f"Missing LinkedIn export file: {relative_path}")
        return path.read_text(encoding="utf-8-sig")

    def read_optional_text(self, relative_path: str) -> str | None:
        try:
            return self.read_text(relative_path)
        except FileNotFoundError:
            return None


def _read_csv_rows(
    reader: _ExportReader,
    relative_path: str,
    required_headers: set[str],
    *,
    optional: bool = False,
) -> list[dict[str, str]]:
    text = (
        reader.read_optional_text(relative_path)
        if optional
        else reader.read_text(relative_path)
    )
    if text is None:
        return []
    parsed_rows = list(csv.reader(io.StringIO(text, newline="")))
    header_index = next(
        (
            index
            for index, row in enumerate(parsed_rows)
            if required_headers.issubset({value.strip() for value in row})
        ),
        None,
    )
    if header_index is None:
        if optional:
            return []
        raise ValueError(f"Expected columns were not found in {relative_path}")
    headers = [value.strip() for value in parsed_rows[header_index]]
    records: list[dict[str, str]] = []
    for row in parsed_rows[header_index + 1 :]:
        if not any(value.strip() for value in row):
            continue
        padded = [*row, *([""] * max(0, len(headers) - len(row)))]
        records.append(
            {
                header: padded[index].strip() if index < len(padded) else ""
                for index, header in enumerate(headers)
            }
        )
    return records


def _parse_export_date(value: str) -> date | datetime | str | None:
    cleaned = value.strip()
    if not cleaned:
        return None
    formats = (
        "%d %b %Y",
        "%m/%d/%y, %I:%M %p",
        "%d/%m/%y, %I:%M %p",
        "%a %b %d %H:%M:%S %Z %Y",
        "%Y-%m-%d",
    )
    for format_string in formats:
        try:
            parsed = datetime.strptime(cleaned, format_string)
            if "%H" in format_string or "%I" in format_string:
                return parsed
            return parsed.date()
        except ValueError:
            continue
    return cleaned


_TRAILING_COMPANY_TOKENS = {
    "co",
    "company",
    "corp",
    "corporation",
    "inc",
    "incorporated",
    "limited",
    "llc",
    "ltd",
    "plc",
    "private",
    "pvt",
}


def _normalize_company(value: str) -> str:
    ascii_text = (
        unicodedata.normalize("NFKD", value)
        .encode("ascii", "ignore")
        .decode("ascii")
        .casefold()
        .replace("&", " and ")
    )
    tokens = re.findall(r"[a-z0-9]+", ascii_text)
    while tokens and tokens[-1] in _TRAILING_COMPANY_TOKENS:
        tokens.pop()
    return " ".join(tokens)


def _company_aliases(company: str) -> set[str]:
    candidates = {company}
    parenthetical_values = re.findall(r"\(([^)]+)\)", company)
    candidates.update(parenthetical_values)
    candidates.add(re.sub(r"\s*\([^)]+\)", "", company).strip())
    if company.casefold().endswith(" india"):
        candidates.add(company[:-6].strip())
    return {_normalize_company(candidate) for candidate in candidates if candidate}


_MANUAL_ALIASES = {
    "amazon web services": "Amazon",
    "aws": "Amazon",
    "dream11": "Dream Sports (Dream11)",
    "ernst and young": "Ernst & Young (EY)",
    "ey": "Ernst & Young (EY)",
    "hcl technologies": "HCLTech",
    "johnson and johnson": "Johnson & Johnson",
    "j and j": "Johnson & Johnson",
    "larsen and toubro technology services": "L&T Technology Services",
    "ltts": "L&T Technology Services",
    "meta platforms": "Meta",
    "pricewaterhousecoopers": "PwC",
    "procter and gamble": "P&G",
    "tata consultancy services": "Tata Consultancy Services (TCS)",
    "tcs": "Tata Consultancy Services (TCS)",
}


def _build_registry_index(
    entries: Iterable[RegistryEntry],
) -> tuple[dict[str, RegistryEntry], dict[str, int]]:
    entry_list = list(entries)
    by_name = {entry.company: entry for entry in entry_list}
    candidate_aliases: dict[str, list[RegistryEntry]] = defaultdict(list)
    category_counts: dict[str, int] = Counter(entry.category for entry in entry_list)
    for entry in entry_list:
        for alias in _company_aliases(entry.company):
            candidate_aliases[alias].append(entry)
    for alias, canonical_name in _MANUAL_ALIASES.items():
        entry = by_name.get(canonical_name)
        if entry is not None:
            candidate_aliases[_normalize_company(alias)].append(entry)
    index: dict[str, RegistryEntry] = {}
    for alias, matches in candidate_aliases.items():
        unique = {match.company: match for match in matches}
        if len(unique) == 1:
            index[alias] = next(iter(unique.values()))
    return index, category_counts


def _match_registry(
    company: str,
    registry_index: dict[str, RegistryEntry],
) -> tuple[RegistryEntry | None, str]:
    normalized = _normalize_company(company)
    if not normalized:
        return None, "No company supplied"
    match = registry_index.get(normalized)
    if match is not None:
        return match, "Exact/alias snapshot match"
    if normalized.endswith(" india"):
        match = registry_index.get(normalized[:-6].strip())
        if match is not None:
            return match, "Country-suffix alias match"
    return None, "No registry match"


def _extract_map_value(text: str, key: str) -> str:
    matches = list(
        re.finditer(
            rf"(?<![A-Za-z0-9_]){re.escape(key)}\s*[:=]\s*",
            text,
            flags=re.IGNORECASE,
        )
    )
    if not matches:
        return ""
    start = matches[-1].end()
    depth = 0
    quote = ""
    escaped = False
    for index in range(start, len(text)):
        character = text[index]
        if escaped:
            escaped = False
            continue
        if character == "\\":
            escaped = True
            continue
        if quote:
            if character == quote:
                quote = ""
            continue
        if character in {'"', "'"}:
            quote = character
            continue
        if character in "[{(":
            depth += 1
            continue
        if character in "]})":
            if depth == 0:
                return text[start:index].strip().strip('"\'')
            depth -= 1
            continue
        if character == "," and depth == 0:
            return text[start:index].strip().strip('"\'')
    return text[start:].strip().strip('"\'')


def _safe_int(value: str) -> int | str | None:
    cleaned = value.strip()
    if not cleaned:
        return None
    match = re.search(r"-?\d+", cleaned)
    return int(match.group()) if match else cleaned


def _sanitize_geo_location(value: str) -> str:
    return re.sub(r"^\s*\d{5,6}\s*,?\s*", "", value).strip()


def _humanize_export_enum(value: str) -> str:
    cleaned = value.strip().strip("[]")
    if not cleaned:
        return ""
    if cleaned.casefold() in {"true", "false"}:
        return "Yes" if cleaned.casefold() == "true" else "No"
    if re.fullmatch(r"[A-Z0-9_, ]+", cleaned):
        return ", ".join(
            item.strip().replace("_", " ").title()
            for item in cleaned.split(",")
            if item.strip()
        )
    return cleaned


def _extract_geo_identifier(value: str) -> str:
    match = re.search(r"urn:li:geo:(\d+)", value)
    return f"urn:li:geo:{match.group(1)}" if match else ""


def _humanize_workplace_types(value: str) -> str:
    labels = {"1": "On-site", "2": "Remote", "3": "Hybrid"}
    codes = re.findall(r"urn:li:workplaceType:(\d+)", value)
    return ", ".join(labels.get(code, f"Code {code}") for code in dict.fromkeys(codes))


def load_linkedin_export(
    source: Path,
    registry_entries: Iterable[RegistryEntry],
) -> LinkedInExportData:
    reader = _ExportReader(source)
    try:
        registry_index, registry_category_counts = _build_registry_index(
            registry_entries
        )
        raw_connections = _read_csv_rows(
            reader,
            "Connections.csv",
            {"First Name", "Last Name", "URL", "Company"},
        )
        connections: list[dict[str, object]] = []
        for row in raw_connections:
            name = " ".join(
                value for value in (row.get("First Name", ""), row.get("Last Name", "")) if value
            ).strip() or "Unnamed LinkedIn connection"
            company = row.get("Company", "")
            match, method = _match_registry(company, registry_index)
            email = row.get("Email Address", "")
            profile_url = row.get("URL", "")
            if email and profile_url:
                contact_options = "LinkedIn + email"
            elif email:
                contact_options = "Email"
            elif profile_url:
                contact_options = "LinkedIn"
            else:
                contact_options = "No exported contact channel"
            connections.append(
                {
                    "name": name,
                    "company": company,
                    "position": row.get("Position", ""),
                    "registry_company": match.company if match else "",
                    "registry_category": match.category if match else "",
                    "referral_status": (
                        "Target-company connection" if match else "Unmatched company"
                    ),
                    "email": email,
                    "profile_url": profile_url,
                    "connected_on": _parse_export_date(row.get("Connected On", "")),
                    "contact_options": contact_options,
                    "match_method": method,
                    "careers_url": match.careers_url if match else "",
                    "jobs_url": match.jobs_url if match else "",
                }
            )
        category_order = list(registry_category_counts)
        category_rank = {category: index for index, category in enumerate(category_order)}
        connections.sort(
            key=lambda item: (
                0 if item["registry_company"] else 1,
                category_rank.get(str(item["registry_category"]), 99),
                str(item["registry_company"] or item["company"]).casefold(),
                str(item["name"]).casefold(),
            )
        )

        raw_applications = _read_csv_rows(
            reader,
            "Jobs/Job Applications.csv",
            {"Application Date", "Company Name", "Job Title", "Job Url"},
        )
        applications: list[dict[str, object]] = []
        for row in raw_applications:
            match, method = _match_registry(row.get("Company Name", ""), registry_index)
            applications.append(
                {
                    "application_date": _parse_export_date(row.get("Application Date", "")),
                    "company": row.get("Company Name", ""),
                    "job_title": row.get("Job Title", ""),
                    "job_url": row.get("Job Url", ""),
                    "registry_company": match.company if match else "",
                    "registry_category": match.category if match else "",
                    "official_portal": match.jobs_url if match else "",
                    "resume_name": row.get("Resume Name", ""),
                    "match_method": method,
                }
            )
        applications.sort(
            key=lambda item: (
                isinstance(item["application_date"], (date, datetime)),
                item["application_date"] or "",
            ),
            reverse=True,
        )

        raw_saved_jobs = _read_csv_rows(
            reader,
            "Jobs/Saved Jobs.csv",
            {"Saved Date", "Job Url", "Job Title", "Company Name"},
        )
        saved_jobs: list[dict[str, object]] = []
        for row in raw_saved_jobs:
            match, method = _match_registry(row.get("Company Name", ""), registry_index)
            saved_jobs.append(
                {
                    "saved_date": _parse_export_date(row.get("Saved Date", "")),
                    "company": row.get("Company Name", ""),
                    "job_title": row.get("Job Title", ""),
                    "job_url": row.get("Job Url", ""),
                    "registry_company": match.company if match else "",
                    "registry_category": match.category if match else "",
                    "official_portal": match.jobs_url if match else "",
                    "match_method": method,
                }
            )
        saved_jobs.sort(
            key=lambda item: (
                isinstance(item["saved_date"], (date, datetime)),
                item["saved_date"] or "",
            ),
            reverse=True,
        )

        raw_alerts = _read_csv_rows(
            reader,
            "SavedJobAlerts.csv",
            {"ALERT_PARAMETERS", "QUERY_CONTEXT", "SAVED_SEARCH_ID"},
        )
        job_alerts: list[dict[str, object]] = []
        for row in raw_alerts:
            parameters = row.get("ALERT_PARAMETERS", "")
            context = row.get("QUERY_CONTEXT", "")
            job_alerts.append(
                {
                    "alert_id": row.get("SAVED_SEARCH_ID", ""),
                    "frequency": _humanize_export_enum(
                        _extract_map_value(parameters, "frequency")
                    ),
                    "keywords": _extract_map_value(context, "keywords"),
                    "geo_identifier": _extract_geo_identifier(context),
                    "radius_km": _safe_int(_extract_map_value(context, "radiusInKms")),
                    "workplace_types": _humanize_workplace_types(
                        _extract_map_value(context, "workplaceTypes")
                    ),
                    "channels": _humanize_export_enum(
                        _extract_map_value(parameters, "channels")
                    ),
                    "sorting_type": _humanize_export_enum(
                        _extract_map_value(context, "sortingType")
                    ),
                    "smart_expansion": _humanize_export_enum(
                        _extract_map_value(parameters, "smartExpansionEnabled")
                    ),
                }
            )

        raw_follows = _read_csv_rows(
            reader,
            "Company Follows.csv",
            {"Organization", "Followed On"},
        )
        followed_companies: list[dict[str, object]] = []
        for row in raw_follows:
            match, method = _match_registry(row.get("Organization", ""), registry_index)
            followed_companies.append(
                {
                    "organization": row.get("Organization", ""),
                    "followed_on": _parse_export_date(row.get("Followed On", "")),
                    "registry_company": match.company if match else "",
                    "registry_category": match.category if match else "",
                    "careers_url": match.careers_url if match else "",
                    "jobs_url": match.jobs_url if match else "",
                    "registry_status": (
                        "Already in registry" if match else "Registry candidate"
                    ),
                    "match_method": method,
                }
            )
        followed_companies.sort(
            key=lambda item: (
                0 if item["registry_company"] else 1,
                str(item["registry_company"] or item["organization"]).casefold(),
            )
        )

        profile_rows = _read_csv_rows(
            reader,
            "Profile.csv",
            {"Headline", "Summary", "Industry", "Geo Location"},
            optional=True,
        )
        raw_profile = profile_rows[0] if profile_rows else {}
        profile = {
            key: raw_profile.get(key, "")
            for key in ("Headline", "Summary", "Industry", "Geo Location", "Websites")
        }
        profile["Geo Location"] = _sanitize_geo_location(profile["Geo Location"])

        preference_rows = _read_csv_rows(
            reader,
            "Jobs/Job Seeker Preferences.csv",
            {"Locations", "Preferred Job Types", "Job Titles"},
            optional=True,
        )
        raw_preferences = preference_rows[0] if preference_rows else {}
        preference_fields = (
            "Industries",
            "Company Employee Count",
            "Preferred Job Types",
            "Job Titles",
            "Open To Recruiters",
            "Dream Companies",
            "Profile Shared With Job Poster",
            "Job Title For Searching Fast Growing Companies",
            "Introduction Statement",
            "Job Seeker Activity Level",
            "Preferred Start Time Range",
            "Open Candidate Visibility",
            "Job Seeking Urgency Level",
            "Semantic Preferences",
        )
        preferences = {
            key: _humanize_export_enum(raw_preferences.get(key, ""))
            for key in preference_fields
        }

        positions = _read_csv_rows(
            reader,
            "Positions.csv",
            {"Company Name", "Title", "Started On"},
            optional=True,
        )
        education = _read_csv_rows(
            reader,
            "Education.csv",
            {"School Name", "Degree Name"},
            optional=True,
        )
        skill_rows = _read_csv_rows(
            reader,
            "Skills.csv",
            {"Name"},
            optional=True,
        )
        skills = [row.get("Name", "") for row in skill_rows if row.get("Name", "")]

        matched_connections = [
            item for item in connections if item["registry_company"]
        ]
        connection_counts = Counter(
            str(item["registry_company"]) for item in matched_connections
        )
        email_counts = Counter(
            str(item["registry_company"])
            for item in matched_connections
            if item["email"]
        )
        company_entries: dict[str, RegistryEntry] = {}
        for entry in registry_index.values():
            company_entries[entry.company] = entry
        top_referrals = [
            {
                "company": company,
                "category": company_entries[company].category,
                "connections": count,
                "email_contacts": email_counts.get(company, 0),
                "careers_url": company_entries[company].careers_url,
                "jobs_url": company_entries[company].jobs_url,
            }
            for company, count in connection_counts.most_common(15)
        ]
        category_stats: list[dict[str, object]] = []
        for category, registry_count in registry_category_counts.items():
            category_connections = [
                item
                for item in matched_connections
                if item["registry_category"] == category
            ]
            category_stats.append(
                {
                    "category": category,
                    "registry_companies": registry_count,
                    "connections": len(category_connections),
                    "connected_companies": len(
                        {
                            str(item["registry_company"])
                            for item in category_connections
                        }
                    ),
                    "email_contacts": sum(
                        1 for item in category_connections if item["email"]
                    ),
                }
            )

        source_label = source.stem.replace(".zip", "")
        return LinkedInExportData(
            source_label=source_label,
            connections=connections,
            applications=applications,
            saved_jobs=saved_jobs,
            job_alerts=job_alerts,
            followed_companies=followed_companies,
            profile=profile,
            preferences=preferences,
            positions=positions,
            education=education,
            skills=skills,
            category_stats=category_stats,
            top_referrals=top_referrals,
            registry_companies_reached=len(connection_counts),
        )
    finally:
        reader.close()


def _style_title_rows(
    sheet,
    *,
    title: str,
    note: str,
    process: str,
    final_column: int,
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 85
    sheet.sheet_properties.tabColor = "0F766E"
    for row in (1, 2, 3):
        sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=final_column,
        )
    title_cell = sheet.cell(1, 1, title)
    title_cell.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    note_cell = sheet.cell(2, 1, note)
    note_cell.font = Font(name="Aptos", size=10, italic=True, color="334155")
    note_cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 42

    process_cell = sheet.cell(3, 1, process)
    process_cell.font = Font(name="Aptos", size=10, color="115E59")
    process_cell.fill = PatternFill("solid", fgColor=PROCESS_FILL)
    process_cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 28


def _add_table_sheet(
    workbook: Workbook,
    *,
    sheet_name: str,
    title: str,
    note: str,
    process: str,
    columns: list[tuple[str, str]],
    rows: list[dict[str, object]],
    widths: list[int],
    hyperlinks: dict[str, str] | None = None,
    email_fields: set[str] | None = None,
    date_fields: set[str] | None = None,
    status_field: str | None = None,
) -> None:
    sheet = workbook[sheet_name]
    final_column = len(columns)
    _style_title_rows(
        sheet,
        title=title,
        note=note,
        process=process,
        final_column=final_column,
    )
    for column_index, (header, _) in enumerate(columns, start=1):
        cell = sheet.cell(4, column_index, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 38

    hyperlink_fields = hyperlinks or {}
    mail_fields = email_fields or set()
    typed_date_fields = date_fields or set()
    field_columns = {field: index for index, (_, field) in enumerate(columns, start=1)}
    for row_index, row in enumerate(rows, start=5):
        for column_index, (_, field) in enumerate(columns, start=1):
            value = row.get(field, "")
            cell = sheet.cell(row_index, column_index, value)
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if field in typed_date_fields and isinstance(value, (date, datetime)):
                cell.number_format = (
                    "yyyy-mm-dd hh:mm" if isinstance(value, datetime) else "yyyy-mm-dd"
                )
            target_field = hyperlink_fields.get(field)
            target = str(row.get(target_field, "")) if target_field else ""
            if target:
                cell.hyperlink = target
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif field in mail_fields and value:
                cell.hyperlink = f"mailto:{value}"
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 46

    final_row = len(rows) + 4
    table = Table(
        displayName=LINKEDIN_TABLE_NAMES[sheet_name],
        ref=f"A4:{sheet.cell(4, final_column).column_letter}{final_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A5"
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(4, column_index).column_letter].width = width

    if status_field and status_field in field_columns and rows:
        column_letter = sheet.cell(4, field_columns[status_field]).column_letter
        status_range = f"{column_letter}5:{column_letter}{final_row}"
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'LEFT({column_letter}5,6)="Target"'],
                fill=PatternFill("solid", fgColor="DCFCE7"),
            ),
        )
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'LEFT({column_letter}5,7)="Already"'],
                fill=PatternFill("solid", fgColor="DCFCE7"),
            ),
        )
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'LEFT({column_letter}5,8)="Registry"'],
                fill=PatternFill("solid", fgColor="FEF3C7"),
            ),
        )

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:4"
    sheet.oddFooter.center.text = title
    sheet.oddFooter.right.text = "Page &P of &N"


def _add_overview_sheet(workbook: Workbook, data: LinkedInExportData) -> None:
    sheet = workbook["LinkedIn Overview"]
    _style_title_rows(
        sheet,
        title="LinkedIn Job-Search Overview",
        note=(
            f"Snapshot: {data.source_label}. Imported locally for personal job search. "
            "No contact enrichment, scraping, or external lookup was performed."
        ),
        process=(
            "Use registry-matched connections for referrals, application history to avoid "
            "duplicate effort, and saved jobs/alerts/follows as discovery inputs."
        ),
        final_column=8,
    )
    connections_end = len(data.connections) + 4
    applications_end = len(data.applications) + 4
    saved_jobs_end = len(data.saved_jobs) + 4
    alerts_end = len(data.job_alerts) + 4
    follows_end = len(data.followed_companies) + 4
    cards = [
        ("Connections", f"=COUNTA('LinkedIn Connections'!$A$5:$A${connections_end})"),
        (
            "Target-company connections",
            f'=COUNTIF(\'LinkedIn Connections\'!$F$5:$F${connections_end},"Target-company connection")',
        ),
        (
            "Exported connection emails",
            f'=COUNTIF(\'LinkedIn Connections\'!$G$5:$G${connections_end},"<>")',
        ),
        ("Registry companies reached", data.registry_companies_reached),
        (
            "Prior applications",
            f"=COUNTA('LinkedIn Applications'!$A$5:$A${applications_end})",
        ),
        (
            "Saved jobs",
            f"=COUNTA('LinkedIn Saved Jobs'!$A$5:$A${saved_jobs_end})",
        ),
        (
            "Job alerts",
            f"=COUNTA('LinkedIn Job Alerts'!$A$5:$A${alerts_end})",
        ),
        (
            "Followed companies",
            f"=COUNTA('Followed Companies'!$A$5:$A${follows_end})",
        ),
    ]
    for index, (label, value) in enumerate(cards):
        card_row = 5 if index < 4 else 9
        card_column = (index % 4) * 2 + 1
        sheet.merge_cells(
            start_row=card_row,
            start_column=card_column,
            end_row=card_row,
            end_column=card_column + 1,
        )
        sheet.merge_cells(
            start_row=card_row + 1,
            start_column=card_column,
            end_row=card_row + 2,
            end_column=card_column + 1,
        )
        label_cell = sheet.cell(card_row, card_column, label)
        label_cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        label_cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell = sheet.cell(card_row + 1, card_column, value)
        value_cell.font = Font(name="Aptos Display", size=22, bold=True, color="0F766E")
        value_cell.fill = PatternFill("solid", fgColor="F0FDFA")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[card_row].height = 26
        sheet.row_dimensions[card_row + 1].height = 28
        sheet.row_dimensions[card_row + 2].height = 18

    sheet.merge_cells("A13:H13")
    section = sheet["A13"]
    section.value = "Referral Coverage by Registry Category"
    section.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    section.fill = PatternFill("solid", fgColor=TITLE_FILL)
    section.alignment = Alignment(vertical="center")
    category_headers = [
        "Category",
        "Registry Companies",
        "Connections",
        "Connected Companies",
        "Email Contacts",
    ]
    for column, header in enumerate(category_headers, start=1):
        cell = sheet.cell(14, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, item in enumerate(data.category_stats, start=15):
        values = [
            item["category"],
            item["registry_companies"],
            item["connections"],
            item["connected_companies"],
            item["email_contacts"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.fill = PatternFill(
                "solid",
                fgColor="FFFFFF" if row_index % 2 else "F8FAFC",
            )
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row_index].height = 28

    top_start = 22
    sheet.merge_cells(start_row=top_start, start_column=1, end_row=top_start, end_column=8)
    section = sheet.cell(top_start, 1, "Top Registry Companies by Referral Connections")
    section.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    section.fill = PatternFill("solid", fgColor=TITLE_FILL)
    section.alignment = Alignment(vertical="center")
    top_headers = [
        "Company",
        "Category",
        "Connections",
        "Exported Emails",
        "Official Careers Page",
        "Direct Job Portal",
    ]
    for column, header in enumerate(top_headers, start=1):
        cell = sheet.cell(top_start + 1, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_index, item in enumerate(data.top_referrals, start=top_start + 2):
        values = [
            item["company"],
            item["category"],
            item["connections"],
            item["email_contacts"],
            item["careers_url"],
            item["jobs_url"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.fill = PatternFill(
                "solid",
                fgColor="FFFFFF" if row_index % 2 else "F8FAFC",
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in {1, 5, 6} and value:
                target = item["jobs_url"] if column == 1 else value
                cell.hyperlink = str(target)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 40

    privacy_row = top_start + 2 + max(1, len(data.top_referrals)) + 1
    sheet.merge_cells(
        start_row=privacy_row,
        start_column=1,
        end_row=privacy_row + 1,
        end_column=8,
    )
    privacy = sheet.cell(privacy_row, 1)
    privacy.value = (
        "Privacy boundary: private messages, invitations, raw search/ad history, birth date, "
        "home address, the owner's phone/email files, and saved screening answers were not "
        "imported. Connection emails shown here came only from the supplied export."
    )
    privacy.font = Font(name="Aptos", size=10, italic=True, color="7C2D12")
    privacy.fill = PatternFill("solid", fgColor="FFEDD5")
    privacy.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[privacy_row].height = 28
    sheet.row_dimensions[privacy_row + 1].height = 28

    widths = [30, 26, 18, 22, 42, 48, 20, 20]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A14"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def _add_profile_sheet(workbook: Workbook, data: LinkedInExportData) -> None:
    sheet = workbook["LinkedIn Profile"]
    _style_title_rows(
        sheet,
        title="LinkedIn Profile & Job Preferences",
        note=(
            "Only job-search-relevant profile fields are included. Direct self-contact, "
            "birth date, and street-address fields remain excluded."
        ),
        process=(
            "Use this snapshot as supporting evidence for role matching; keep the resume "
            "as the primary source for eligibility claims."
        ),
        final_column=8,
    )
    current_row = 5

    def add_section(title: str) -> None:
        nonlocal current_row
        sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=8,
        )
        cell = sheet.cell(current_row, 1, title)
        cell.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[current_row].height = 26
        current_row += 1

    add_section("Public Professional Profile")
    for field in ("Headline", "Summary", "Industry", "Geo Location", "Websites"):
        field_value = data.profile.get(field, "").strip()
        if not field_value:
            continue
        sheet.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=8,
        )
        label = sheet.cell(current_row, 1, field)
        label.font = Font(name="Aptos", size=10, bold=True, color="334155")
        label.fill = PatternFill("solid", fgColor="F1F5F9")
        label.alignment = Alignment(wrap_text=True, vertical="top")
        value = sheet.cell(current_row, 2, field_value)
        value.font = Font(name="Aptos", size=10, color="1F2937")
        value.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[current_row].height = 70 if field == "Summary" else 34
        current_row += 1

    current_row += 1
    add_section("Job-Seeker Preferences")
    for field, value_text in data.preferences.items():
        if not value_text.strip():
            continue
        sheet.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=8,
        )
        label = sheet.cell(current_row, 1, field)
        label.font = Font(name="Aptos", size=9, bold=True, color="334155")
        label.fill = PatternFill("solid", fgColor="F1F5F9")
        label.alignment = Alignment(wrap_text=True, vertical="top")
        value = sheet.cell(current_row, 2, value_text)
        value.font = Font(name="Aptos", size=9, color="1F2937")
        value.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[current_row].height = 48
        current_row += 1

    current_row += 1
    add_section("Experience")
    experience_headers = [
        "Company Name",
        "Title",
        "Description",
        "Location",
        "Started On",
        "Finished On",
    ]
    for column, header in enumerate(experience_headers, start=1):
        cell = sheet.cell(current_row, column, header)
        cell.font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    current_row += 1
    for position in data.positions:
        for column, field in enumerate(experience_headers, start=1):
            cell = sheet.cell(current_row, column, position.get(field, ""))
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[current_row].height = 62
        current_row += 1

    current_row += 1
    add_section("Education")
    education_headers = [
        "School Name",
        "Degree Name",
        "Start Date",
        "End Date",
        "Activities",
        "Notes",
    ]
    for column, header in enumerate(education_headers, start=1):
        cell = sheet.cell(current_row, column, header)
        cell.font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    current_row += 1
    for item in data.education:
        for column, field in enumerate(education_headers, start=1):
            cell = sheet.cell(current_row, column, item.get(field, ""))
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[current_row].height = 54
        current_row += 1

    current_row += 1
    add_section("Skills")
    for index, skill in enumerate(data.skills, start=1):
        sheet.cell(current_row, 1, index)
        sheet.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=4,
        )
        cell = sheet.cell(current_row, 2, skill)
        cell.font = Font(name="Aptos", size=10, color="1F2937")
        cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[current_row].height = 24
        current_row += 1

    widths = [34, 30, 46, 26, 18, 18, 20, 20]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A6"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def add_linkedin_sheets(workbook: Workbook, data: LinkedInExportData) -> None:
    for sheet_name in LINKEDIN_SHEET_NAMES:
        workbook.create_sheet(sheet_name)
    _add_overview_sheet(workbook, data)
    _add_profile_sheet(workbook, data)

    _add_table_sheet(
        workbook,
        sheet_name="LinkedIn Connections",
        title="LinkedIn Connections for Referral Discovery",
        note=(
            f"{len(data.connections):,} connection records from the supplied export. "
            "Registry matches are conservative exact/alias snapshot matches, not identity guarantees."
        ),
        process=(
            "Filter Referral Status = Target-company connection, then open the person's "
            "profile and the official job portal before requesting a referral."
        ),
        columns=[
            ("Connection Name", "name"),
            ("Current Company", "company"),
            ("Current Position", "position"),
            ("Registry Company", "registry_company"),
            ("Registry Category", "registry_category"),
            ("Referral Status", "referral_status"),
            ("Email Address", "email"),
            ("LinkedIn Profile", "profile_url"),
            ("Connected On", "connected_on"),
            ("Contact Options", "contact_options"),
            ("Match Method", "match_method"),
            ("Official Careers Page", "careers_url"),
            ("Direct Job Portal", "jobs_url"),
        ],
        rows=data.connections,
        widths=[28, 30, 34, 30, 22, 27, 30, 50, 15, 23, 28, 44, 52],
        hyperlinks={
            "name": "profile_url",
            "profile_url": "profile_url",
            "registry_company": "jobs_url",
            "careers_url": "careers_url",
            "jobs_url": "jobs_url",
        },
        email_fields={"email"},
        date_fields={"connected_on"},
        status_field="referral_status",
    )
    workbook["LinkedIn Connections"]["G4"].comment = Comment(
        "Only email addresses already present in the supplied LinkedIn export are shown; no enrichment was performed.",
        "User",
    )

    _add_table_sheet(
        workbook,
        sheet_name="LinkedIn Applications",
        title="LinkedIn Application History",
        note=(
            f"{len(data.applications):,} prior applications. Saved screening questions and "
            "answers and repeated applicant self-contact fields were intentionally excluded."
        ),
        process=(
            "Use this tab to avoid duplicate applications and to reopen the original job or "
            "the company's current official portal."
        ),
        columns=[
            ("Application Date", "application_date"),
            ("Company", "company"),
            ("Job Title", "job_title"),
            ("Original Job URL", "job_url"),
            ("Registry Company", "registry_company"),
            ("Registry Category", "registry_category"),
            ("Current Official Portal", "official_portal"),
            ("Resume Name", "resume_name"),
            ("Match Method", "match_method"),
        ],
        rows=data.applications,
        widths=[19, 30, 38, 52, 30, 22, 52, 28, 28],
        hyperlinks={
            "job_url": "job_url",
            "registry_company": "official_portal",
            "official_portal": "official_portal",
        },
        date_fields={"application_date"},
    )

    _add_table_sheet(
        workbook,
        sheet_name="LinkedIn Saved Jobs",
        title="LinkedIn Saved Jobs",
        note=f"{len(data.saved_jobs):,} saved jobs from the export snapshot.",
        process=(
            "Open the original URL first; if it expired, use the matched current official "
            "portal and search by company/title."
        ),
        columns=[
            ("Saved Date", "saved_date"),
            ("Company", "company"),
            ("Job Title", "job_title"),
            ("Original Job URL", "job_url"),
            ("Registry Company", "registry_company"),
            ("Registry Category", "registry_category"),
            ("Current Official Portal", "official_portal"),
            ("Match Method", "match_method"),
        ],
        rows=data.saved_jobs,
        widths=[19, 30, 38, 52, 30, 22, 52, 28],
        hyperlinks={
            "job_url": "job_url",
            "registry_company": "official_portal",
            "official_portal": "official_portal",
        },
        date_fields={"saved_date"},
    )

    _add_table_sheet(
        workbook,
        sheet_name="LinkedIn Job Alerts",
        title="LinkedIn Saved Job Alerts",
        note=(
            f"{len(data.job_alerts):,} saved alert configurations, normalized from LinkedIn's "
            "exported parameter format."
        ),
        process=(
            "Use the keywords, location, radius, and workplace settings to recreate or tune "
            "alerts without importing raw opaque query-context text."
        ),
        columns=[
            ("Saved Search ID", "alert_id"),
            ("Frequency", "frequency"),
            ("Keywords", "keywords"),
            ("Geo Identifier", "geo_identifier"),
            ("Radius (km)", "radius_km"),
            ("Workplace Types", "workplace_types"),
            ("Channels", "channels"),
            ("Sorting Type", "sorting_type"),
            ("Smart Expansion", "smart_expansion"),
        ],
        rows=data.job_alerts,
        widths=[28, 16, 34, 40, 14, 32, 24, 18, 18],
    )

    _add_table_sheet(
        workbook,
        sheet_name="Followed Companies",
        title="LinkedIn Followed Companies",
        note=(
            f"{len(data.followed_companies):,} followed organizations. Unmatched organizations "
            "are marked as registry candidates rather than silently added."
        ),
        process=(
            "Review Registry Candidate rows as possible future company-source additions; "
            "matched rows link directly to the current official careers/job portals."
        ),
        columns=[
            ("Organization", "organization"),
            ("Followed On", "followed_on"),
            ("Registry Company", "registry_company"),
            ("Registry Category", "registry_category"),
            ("Official Careers Page", "careers_url"),
            ("Direct Job Portal", "jobs_url"),
            ("Registry Status", "registry_status"),
            ("Match Method", "match_method"),
        ],
        rows=data.followed_companies,
        widths=[34, 22, 30, 22, 48, 54, 24, 28],
        hyperlinks={
            "registry_company": "jobs_url",
            "careers_url": "careers_url",
            "jobs_url": "jobs_url",
        },
        date_fields={"followed_on"},
        status_field="registry_status",
    )


def verify_linkedin_sheets(workbook: Workbook, data: LinkedInExportData) -> None:
    for sheet_name in LINKEDIN_SHEET_NAMES:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"LinkedIn worksheet is missing: {sheet_name}")
    expected_rows = {
        "LinkedIn Connections": len(data.connections) + 4,
        "LinkedIn Applications": len(data.applications) + 4,
        "LinkedIn Saved Jobs": len(data.saved_jobs) + 4,
        "LinkedIn Job Alerts": len(data.job_alerts) + 4,
        "Followed Companies": len(data.followed_companies) + 4,
    }
    for sheet_name, final_row in expected_rows.items():
        sheet = workbook[sheet_name]
        if sheet.max_row != final_row:
            raise ValueError(
                f"Unexpected row count on {sheet_name}: {sheet.max_row}; expected {final_row}"
            )
        if sheet.auto_filter.ref is not None:
            raise ValueError(f"Worksheet filter conflicts with table on {sheet_name}")
        table_name = LINKEDIN_TABLE_NAMES[sheet_name]
        if table_name not in sheet.tables:
            raise ValueError(f"LinkedIn table is missing: {table_name}")
        if sheet.freeze_panes != "A5":
            raise ValueError(f"Unexpected freeze pane on {sheet_name}")

    connections = workbook["LinkedIn Connections"]
    headers = [connections.cell(4, column).value for column in range(1, 14)]
    if "Question And Answers" in headers or "Message" in headers:
        raise ValueError("Sensitive export columns were imported unexpectedly")
    for row in range(5, connections.max_row + 1):
        profile_url = connections.cell(row, 8).value
        if profile_url and not connections.cell(row, 1).hyperlink:
            raise ValueError(f"Connection name hyperlink missing at A{row}")
        email = connections.cell(row, 7).value
        if email and not connections.cell(row, 7).hyperlink:
            raise ValueError(f"Connection email hyperlink missing at G{row}")

    overview = workbook["LinkedIn Overview"]
    for coordinate in ("A6", "C6", "E6", "A10", "C10", "E10", "G10"):
        value = overview[coordinate].value
        if not isinstance(value, str) or not value.startswith("="):
            raise ValueError(f"Overview formula is missing at {coordinate}")
    profile = workbook["LinkedIn Profile"]
    profile_headers = {
        str(profile.cell(row, 1).value)
        for row in range(1, profile.max_row + 1)
        if profile.cell(row, 1).value
    }
    for forbidden in ("Address", "Birth Date", "Phone Number", "Email Address"):
        if forbidden in profile_headers:
            raise ValueError(f"Sensitive profile field was imported: {forbidden}")
