"""Build the reusable MNC career-portal registry workbook."""

from __future__ import annotations

import argparse
import concurrent.futures
import re
import textwrap
import urllib.error
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont

try:
    from scripts.linkedin_export_workbook import (
        LINKEDIN_SHEET_NAMES,
        LINKEDIN_TABLE_NAMES,
        LinkedInExportData,
        RegistryEntry,
        add_linkedin_sheets,
        load_linkedin_export,
        verify_linkedin_sheets,
    )
except ModuleNotFoundError:
    from linkedin_export_workbook import (  # type: ignore[no-redef]
        LINKEDIN_SHEET_NAMES,
        LINKEDIN_TABLE_NAMES,
        LinkedInExportData,
        RegistryEntry,
        add_linkedin_sheets,
        load_linkedin_export,
        verify_linkedin_sheets,
    )


DEFAULT_OUTPUT = (
    Path(__file__).resolve().parents[1]
    / "outputs"
    / "mnc_registry_2026-07-31"
    / "Company_Source_Registry.xlsx"
)


@dataclass(frozen=True)
class CompanySource:
    company: str
    sector: str
    priority: str
    careers_url: str
    jobs_url: str
    source_type: str
    source_identifier: str = ""
    public_endpoint: str = ""
    api_key_required: str = "N/A"
    india_jobs: str = "Global search"
    fallback: str = "Official careers page; Gmail alert; manual official job link"
    notes: str = ""


def source(
    company: str,
    sector: str,
    careers_url: str,
    jobs_url: str,
    source_type: str = "Company-specific / auto-detect",
    *,
    priority: str = "High",
    source_identifier: str = "",
    public_endpoint: str = "",
    api_key_required: str = "N/A",
    india_jobs: str = "Global search",
    fallback: str = "Official careers page; Gmail alert; manual official job link",
    notes: str = "",
) -> CompanySource:
    return CompanySource(
        company=company,
        sector=sector,
        priority=priority,
        careers_url=careers_url,
        jobs_url=jobs_url,
        source_type=source_type,
        source_identifier=source_identifier,
        public_endpoint=public_endpoint,
        api_key_required=api_key_required,
        india_jobs=india_jobs,
        fallback=fallback,
        notes=notes,
    )


COMPANIES = [
    source(
        "Tata Consultancy Services (TCS)",
        "IT services & consulting",
        "https://www.tcs.com/careers",
        "https://ibegin.tcsapps.com/candidate/?geography=IN&language=EN",
        "TCS iBegin / company-specific",
        source_identifier="iBegin",
        india_jobs="Yes",
        notes="The official India page links lateral applications into TCS iBegin.",
    ),
    source(
        "Infosys",
        "IT services & consulting",
        "https://www.infosys.com/careers/",
        "https://career.infosys.com/joblist",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Wipro",
        "IT services & consulting",
        "https://careers.wipro.com/",
        "https://careers.wipro.com/search/",
        "SAP SuccessFactors",
        source_identifier="careers.wipro.com",
        public_endpoint="https://careers.wipro.com/services/rss/job/?locale=en_US",
        api_key_required="No",
        india_jobs="Yes",
        notes="Public RSS is useful but undocumented; verify posting-date semantics.",
    ),
    source(
        "HCLTech",
        "IT services & consulting",
        "https://careers.hcltech.com/",
        "https://careers.hcltech.com/",
        india_jobs="Yes",
    ),
    source(
        "Tech Mahindra",
        "IT services & consulting",
        "https://careers.techmahindra.com/",
        "https://careers.techmahindra.com/",
        india_jobs="Yes",
    ),
    source(
        "LTIMindtree",
        "IT services & consulting",
        "https://www.ltimindtree.com/careers/",
        "https://careers.ltimindtree.com/go/",
        "SAP SuccessFactors",
        india_jobs="Yes",
        notes="Public SAP career site; confirm tenant-specific endpoint before automation.",
    ),
    source(
        "Mphasis",
        "IT services & consulting",
        "https://careers.mphasis.com/",
        "https://careers.mphasis.com/home.html",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Persistent Systems",
        "Digital engineering & IT services",
        "https://www.persistent.com/careers/",
        "https://careers.persistent.com/",
        "Company-specific / SAP candidate",
        india_jobs="Yes",
        notes="Confirm tenant-specific adapter before automation.",
    ),
    source(
        "Coforge",
        "IT services & consulting",
        "https://www.coforge.com/careers",
        "https://careers.coforge.com/",
        india_jobs="Yes",
    ),
    source(
        "L&T Technology Services",
        "Engineering R&D services",
        "https://www.ltts.com/careers",
        "https://www.ltts.com/careers",
        india_jobs="Yes",
    ),
    source(
        "Accenture",
        "Consulting & technology services",
        "https://www.accenture.com/in-en/careers",
        "https://www.accenture.com/in-en/careers/jobsearch",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "IBM",
        "Technology & consulting",
        "https://www.ibm.com/careers",
        "https://www.ibm.com/careers/search",
        "Avature / company-specific",
        india_jobs="Yes",
    ),
    source(
        "Capgemini",
        "Consulting & technology services",
        "https://www.capgemini.com/careers/",
        "https://www.capgemini.com/careers/join-capgemini/job-search/",
        "Company-specific / SAP candidate",
        india_jobs="Yes",
    ),
    source(
        "Cognizant",
        "IT services & consulting",
        "https://careers.cognizant.com/global-en/",
        "https://careers.cognizant.com/global-en/jobs/",
        "Phenom",
        india_jobs="Yes",
    ),
    source(
        "Deloitte",
        "Professional services & consulting",
        "https://www.deloitte.com/in/en/careers.html",
        "https://southasiacareers.deloitte.com/",
        "SAP SuccessFactors",
        source_identifier="southasiacareers.deloitte.com",
        india_jobs="Yes",
    ),
    source(
        "EY",
        "Professional services & consulting",
        "https://www.ey.com/en_in/careers",
        "https://careers.ey.com/",
        "SAP SuccessFactors",
        source_identifier="careers.ey.com",
        india_jobs="Yes",
    ),
    source(
        "PwC",
        "Professional services & consulting",
        "https://www.pwc.in/careers.html",
        "https://www.pwc.in/careers.html",
        "Workday / country-specific portals",
        india_jobs="Yes",
        notes="Use the India careers page first; PwC member-firm job portals vary and require manual navigation.",
    ),
    source(
        "KPMG",
        "Professional services & consulting",
        "https://kpmg.com/in/en/careers.html",
        "https://kpmgindia.talentrecruit.com/career-page",
        "TalentRecruit",
        source_identifier="kpmgindia.talentrecruit.com",
        india_jobs="Yes",
        notes="KPMG Global Services uses a separate Oracle Recruiting Cloud portal.",
    ),
    source(
        "CGI",
        "IT & business consulting",
        "https://www.cgi.com/en/careers",
        "https://cgi.njoyn.com/",
        "Njoyn",
        source_identifier="cgi.njoyn.com",
        india_jobs="Yes",
    ),
    source(
        "NTT DATA",
        "IT services & consulting",
        "https://www.nttdata.com/global/en/careers",
        "https://careers-inc.nttdata.com/default/go/View-All-Jobs/3020100/300/",
        "SAP SuccessFactors",
        source_identifier="careers-inc.nttdata.com",
        india_jobs="Yes",
    ),
    source(
        "DXC Technology",
        "IT services & consulting",
        "https://careers.dxc.com/",
        "https://careers.dxc.com/job-search-results/",
        "Company careers site + Workday",
        source_identifier="dxctechnology | wd1",
        india_jobs="Yes",
    ),
    source(
        "EPAM Systems",
        "Digital engineering services",
        "https://www.epam.com/careers",
        "https://www.epam.com/careers/job-listings",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Genpact",
        "Professional services & operations",
        "https://www.genpact.com/careers",
        "https://genpact.wd108.myworkdayjobs.com/External_Careers",
        "Workday",
        source_identifier="genpact | wd108 | External_Careers",
        india_jobs="Yes",
    ),
    source(
        "Publicis Sapient",
        "Digital consulting",
        "https://careers.publicissapient.com/",
        "https://careers.publicissapient.com/job-search",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Thoughtworks",
        "Technology consulting",
        "https://www.thoughtworks.com/careers",
        "https://job-boards.greenhouse.io/thoughtworks",
        "Greenhouse",
        source_identifier="thoughtworks",
        public_endpoint=(
            "https://boards-api.greenhouse.io/v1/boards/thoughtworks/jobs?content=true"
        ),
        api_key_required="No",
        india_jobs="Yes",
        notes="Greenhouse updated_at is not guaranteed to be first publication date.",
    ),
    source(
        "Virtusa",
        "Digital engineering & IT services",
        "https://www.virtusa.com/careers",
        "https://www.virtusa.com/careers/job-search",
        "Company-specific / Workday candidate",
        india_jobs="Yes",
    ),
    source(
        "UST",
        "Digital technology services",
        "https://www.ust.com/en/careers",
        "https://www.ust.com/en/careers",
        "Company-specific / Eightfold candidate",
        india_jobs="Yes",
    ),
    source(
        "Microsoft",
        "Software & cloud",
        "https://careers.microsoft.com/",
        "https://jobs.careers.microsoft.com/global/en/search",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Google",
        "Internet, software & cloud",
        "https://www.google.com/about/careers/applications/",
        "https://www.google.com/about/careers/applications/jobs/results/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Amazon",
        "E-commerce & cloud",
        "https://www.amazon.jobs/",
        "https://www.amazon.jobs/en/search",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Apple",
        "Consumer technology",
        "https://www.apple.com/careers/in/",
        "https://jobs.apple.com/en-in/search",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Meta",
        "Internet & social technology",
        "https://www.metacareers.com/",
        "https://www.metacareers.com/jobs/",
        "Company-specific",
        india_jobs="Global search",
    ),
    source(
        "Oracle",
        "Enterprise software & cloud",
        "https://www.oracle.com/careers/",
        "https://careers.oracle.com/en/sites/jobsearch/",
        "Oracle Recruiting Cloud",
        source_identifier="jobsearch",
        india_jobs="Yes",
    ),
    source(
        "SAP",
        "Enterprise software & cloud",
        "https://www.sap.com/about/careers.html",
        "https://jobs.sap.com/",
        "SAP SuccessFactors",
        source_identifier="jobs.sap.com",
        india_jobs="Yes",
    ),
    source(
        "Salesforce",
        "Enterprise software & cloud",
        "https://careers.salesforce.com/",
        "https://careers.salesforce.com/en/jobs/",
        "Company-specific / Workday candidate",
        india_jobs="Yes",
    ),
    source(
        "Adobe",
        "Creative & enterprise software",
        "https://careers.adobe.com/",
        "https://careers.adobe.com/us/en/search-results",
        "Phenom",
        india_jobs="Yes",
    ),
    source(
        "Cisco",
        "Networking & enterprise technology",
        "https://www.cisco.com/c/en/us/about/careers.html",
        "https://jobs.cisco.com/jobs/SearchJobs/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Intel",
        "Semiconductors",
        "https://www.intel.com/content/www/us/en/jobs/jobs-at-intel.html",
        "https://jobs.intel.com/en/search-jobs",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "NVIDIA",
        "Semiconductors & AI computing",
        "https://www.nvidia.com/en-us/about-nvidia/careers/",
        "https://nvidia.wd5.myworkdayjobs.com/NVIDIAExternalCareerSite",
        "Workday",
        source_identifier="nvidia | wd5 | NVIDIAExternalCareerSite",
        india_jobs="Yes",
        notes="Workday public endpoints are tenant-specific and undocumented.",
    ),
    source(
        "Qualcomm",
        "Semiconductors & wireless technology",
        "https://www.qualcomm.com/company/careers",
        "https://careers.qualcomm.com/careers",
        "Company-specific / Workday candidate",
        india_jobs="Yes",
    ),
    source(
        "Dell Technologies",
        "Enterprise & consumer technology",
        "https://jobs.dell.com/",
        "https://jobs.dell.com/en/search-jobs",
        "Phenom",
        india_jobs="Yes",
    ),
    source(
        "Hewlett Packard Enterprise (HPE)",
        "Enterprise technology",
        "https://careers.hpe.com/",
        "https://careers.hpe.com/us/en/search-results",
        "Phenom",
        india_jobs="Yes",
    ),
    source(
        "ServiceNow",
        "Enterprise cloud software",
        "https://careers.servicenow.com/",
        "https://careers.smartrecruiters.com/ServiceNow",
        "SmartRecruiters",
        source_identifier="ServiceNow",
        public_endpoint="https://api.smartrecruiters.com/v1/companies/ServiceNow/postings",
        api_key_required="No for public postings",
        india_jobs="Yes",
    ),
    source(
        "Atlassian",
        "Collaboration software",
        "https://www.atlassian.com/company/careers",
        "https://www.atlassian.com/company/careers/all-jobs",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Uber",
        "Mobility & delivery technology",
        "https://www.uber.com/global/en/careers/",
        "https://www.uber.com/global/en/careers/list/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "LinkedIn",
        "Professional network & software",
        "https://careers.linkedin.com/",
        "https://careers.linkedin.com/",
        "Company careers / LinkedIn jobs",
        india_jobs="Yes",
        fallback="Gmail alert; official careers page; manual official job link",
        notes="Treat LinkedIn listings as discovery input; do not scrape protected pages.",
    ),
    source(
        "PayPal",
        "Financial technology",
        "https://careers.pypl.com/",
        "https://paypal.eightfold.ai/careers",
        "Eightfold",
        source_identifier="paypal.eightfold.ai",
        india_jobs="Yes",
    ),
    source(
        "Visa",
        "Payments technology",
        "https://corporate.visa.com/en/careers.html",
        "https://jobs.smartrecruiters.com/Visa",
        "SmartRecruiters",
        source_identifier="Visa",
        public_endpoint="https://api.smartrecruiters.com/v1/companies/Visa/postings",
        api_key_required="No for public postings",
        india_jobs="Yes",
    ),
    source(
        "Mastercard",
        "Payments technology",
        "https://careers.mastercard.com/",
        "https://careers.mastercard.com/us/en/search-results",
        "Phenom",
        india_jobs="Yes",
    ),
    source(
        "American Express",
        "Payments & financial services",
        "https://www.americanexpress.com/en-us/careers/",
        "https://careers.americanexpress.com/en/sites/CX_1/jobs",
        "Oracle Recruiting Cloud",
        source_identifier="careers.americanexpress.com | CX_1",
        india_jobs="Yes",
        notes="American Express migrated away from the former Eightfold URL to its current Oracle-hosted public careers experience.",
    ),
    source(
        "JPMorgan Chase",
        "Banking & financial services",
        "https://careers.jpmorgan.com/",
        "https://jpmc.fa.oraclecloud.com/hcmUI/CandidateExperience/en/sites/CX_1001/jobs",
        "Oracle Recruiting Cloud",
        source_identifier="jpmc | CX_1001",
        india_jobs="Yes",
    ),
    source(
        "Goldman Sachs",
        "Banking & financial services",
        "https://www.goldmansachs.com/careers/",
        "https://higher.gs.com/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Morgan Stanley",
        "Banking & financial services",
        "https://www.morganstanley.com/people-opportunities/careers",
        "https://morganstanley.eightfold.ai/careers",
        "Eightfold",
        source_identifier="morganstanley.eightfold.ai",
        india_jobs="Yes",
    ),
    source(
        "Wells Fargo",
        "Banking & financial services",
        "https://www.wellsfargojobs.com/",
        "https://www.wellsfargojobs.com/en/jobs/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Citi",
        "Banking & financial services",
        "https://jobs.citi.com/",
        "https://jobs.citi.com/search-jobs",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Barclays",
        "Banking & financial services",
        "https://home.barclays/careers/",
        "https://search.jobs.barclays/",
        "Company-specific / Avature candidate",
        india_jobs="Yes",
    ),
    source(
        "HSBC",
        "Banking & financial services",
        "https://www.hsbc.com/careers",
        "https://www.hsbc.com/careers/find-a-job",
        "Country-specific portals",
        india_jobs="Yes",
    ),
    source(
        "Deutsche Bank",
        "Banking & financial services",
        "https://careers.db.com/",
        "https://careers.db.com/professionals/search-roles/",
        "Company-specific / Avature candidate",
        india_jobs="Yes",
    ),
    source(
        "Walmart Global Tech",
        "Retail technology",
        "https://tech.walmart.com/content/walmart-global-tech/en_us/careers.html",
        "https://careers.walmart.com/results",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Siemens",
        "Industrial technology",
        "https://www.siemens.com/global/en/company/jobs.html",
        "https://jobs.siemens.com/careers",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Bosch",
        "Mobility & industrial technology",
        "https://www.bosch.in/careers/",
        "https://www.bosch.in/careers/job-offers/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Honeywell",
        "Industrial technology",
        "https://careers.honeywell.com/",
        "https://careers.honeywell.com/en/sites/Honeywell/jobs",
        "Oracle Recruiting Cloud",
        india_jobs="Yes",
    ),
    source(
        "Ericsson",
        "Telecommunications technology",
        "https://www.ericsson.com/en/careers",
        "https://jobs.ericsson.com/",
        "SAP SuccessFactors candidate",
        india_jobs="Yes",
    ),
    source(
        "Nokia",
        "Telecommunications technology",
        "https://www.nokia.com/about-us/careers/",
        "https://jobs.nokia.com/",
        "SAP SuccessFactors candidate",
        india_jobs="Yes",
    ),
    source(
        "Samsung",
        "Consumer electronics & semiconductors",
        "https://www.samsung.com/in/about-us/careers/",
        "https://www.samsung.com/in/about-us/careers/",
        "Country-specific portals",
        india_jobs="Yes",
    ),
    source(
        "Reddit",
        "Internet & social technology",
        "https://redditinc.com/careers",
        "https://job-boards.greenhouse.io/reddit",
        "Greenhouse",
        priority="Medium",
        source_identifier="reddit",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/reddit/jobs?content=true",
        api_key_required="No",
        india_jobs="Limited / varies",
        notes="Official Reddit careers page links jobs to this Greenhouse board.",
    ),
    source(
        "Discord",
        "Communications software",
        "https://discord.com/careers",
        "https://job-boards.greenhouse.io/discord",
        "Greenhouse",
        priority="Reference",
        source_identifier="discord",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/discord/jobs?content=true",
        api_key_required="No",
        india_jobs="Limited / varies",
    ),
    source(
        "Cloudflare",
        "Cloud & cybersecurity",
        "https://www.cloudflare.com/careers/",
        "https://job-boards.greenhouse.io/cloudflare",
        "Greenhouse",
        priority="Medium",
        source_identifier="cloudflare",
        public_endpoint=(
            "https://boards-api.greenhouse.io/v1/boards/cloudflare/jobs?content=true"
        ),
        api_key_required="No",
        india_jobs="Global search",
    ),
    source(
        "Datadog",
        "Cloud monitoring & security",
        "https://careers.datadoghq.com/",
        "https://careers.datadoghq.com/",
        "Company-specific / Greenhouse candidate",
        priority="Medium",
        india_jobs="Global search",
        notes="Confirm the current board token before enabling a Greenhouse adapter.",
    ),
    source(
        "Figure AI",
        "AI & robotics",
        "https://www.figure.ai/careers",
        "https://job-boards.greenhouse.io/figureai",
        "Greenhouse",
        priority="Reference",
        source_identifier="figureai",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/figureai/jobs?content=true",
        api_key_required="No",
        india_jobs="Limited / varies",
    ),
]


_COMPANY_BY_NAME = {company.company: company for company in COMPANIES}


PRODUCT_COMPANIES = [
    _COMPANY_BY_NAME[name]
    for name in [
        "Microsoft",
        "Google",
        "Amazon",
        "Apple",
        "Meta",
        "Oracle",
        "SAP",
        "Salesforce",
        "Adobe",
        "Cisco",
        "Intel",
        "NVIDIA",
        "Qualcomm",
        "Dell Technologies",
        "Hewlett Packard Enterprise (HPE)",
        "ServiceNow",
        "Atlassian",
        "Uber",
        "LinkedIn",
        "PayPal",
        "Visa",
        "Mastercard",
        "American Express",
        "Walmart Global Tech",
        "Samsung",
        "Reddit",
        "Discord",
        "Cloudflare",
        "Datadog",
        "Figure AI",
    ]
] + [
    source(
        "Netflix",
        "Streaming & entertainment technology",
        "https://jobs.netflix.com/",
        "https://jobs.netflix.com/",
        "Company-specific",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    source(
        "Spotify",
        "Audio streaming",
        "https://www.lifeatspotify.com/",
        "https://www.lifeatspotify.com/jobs",
        "Company-specific",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    source(
        "Airbnb",
        "Travel marketplace technology",
        "https://careers.airbnb.com/",
        "https://careers.airbnb.com/positions/",
        "Company-specific",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    source(
        "Stripe",
        "Payments infrastructure",
        "https://stripe.com/jobs",
        "https://stripe.com/jobs/search",
        "Greenhouse-backed company careers",
        source_identifier="stripe",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/stripe/jobs?content=true",
        api_key_required="No",
        india_jobs="Yes",
        notes="Use the official Stripe search URL for applications; the public Greenhouse board is the structured source.",
    ),
    source(
        "Block",
        "Financial technology",
        "https://block.xyz/careers",
        "https://block.xyz/careers/jobs",
        "Company-specific",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    source(
        "Shopify",
        "Commerce software",
        "https://www.shopify.com/careers",
        "https://www.shopify.com/careers",
        "Ashby-backed company careers",
        priority="Medium",
        india_jobs="Global search",
        notes="Job postings are embedded in the official careers page; confirm any structured endpoint before automation.",
    ),
    source(
        "Zoom",
        "Communications software",
        "https://careers.zoom.us/",
        "https://careers.zoom.us/jobs/search",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Twilio",
        "Cloud communications platform",
        "https://www.twilio.com/en-us/company/jobs",
        "https://jobs.twilio.com/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "GitHub",
        "Developer platform",
        "https://www.github.careers/careers-home",
        "https://www.github.careers/experienced-professionals/jobs/categories",
        "Company-specific",
        india_jobs="Global search",
    ),
    source(
        "GitLab",
        "DevSecOps platform",
        "https://about.gitlab.com/jobs/",
        "https://job-boards.greenhouse.io/gitlab",
        "Greenhouse",
        source_identifier="gitlab",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/gitlab/jobs?content=true",
        api_key_required="No",
        india_jobs="Yes",
    ),
    source(
        "MongoDB",
        "Database platform",
        "https://www.mongodb.com/company/careers",
        "https://job-boards.greenhouse.io/mongodb",
        "Greenhouse",
        source_identifier="mongodb",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/mongodb/jobs?content=true",
        api_key_required="No",
        india_jobs="Yes",
    ),
    source(
        "Snowflake",
        "Cloud data platform",
        "https://careers.snowflake.com/",
        "https://careers.snowflake.com/us/en/search-results",
        "Company-specific / Phenom candidate",
        india_jobs="Yes",
    ),
    source(
        "Databricks",
        "Data & AI platform",
        "https://www.databricks.com/company/careers",
        "https://www.databricks.com/company/careers/open-positions",
        "Greenhouse-backed company careers",
        source_identifier="databricks",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/databricks/jobs?content=true",
        api_key_required="No",
        india_jobs="Yes",
    ),
    source(
        "Confluent",
        "Data streaming platform",
        "https://www.confluent.io/careers/",
        "https://careers.confluent.io/",
        "Company-specific / Greenhouse candidate",
        india_jobs="Yes",
        notes="Confirm the current board token before enabling an ATS adapter.",
    ),
    source(
        "Elastic",
        "Search & observability platform",
        "https://www.elastic.co/careers/",
        "https://jobs.elastic.co/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Okta",
        "Identity software",
        "https://www.okta.com/company/careers/",
        "https://www.okta.com/company/careers/",
        "Company-specific / Workday candidate",
        india_jobs="Yes",
    ),
    source(
        "Palo Alto Networks",
        "Cybersecurity",
        "https://jobs.paloaltonetworks.com/en/",
        "https://jobs.paloaltonetworks.com/en/search_jobs",
        "Phenom",
        india_jobs="Yes",
    ),
    source(
        "CrowdStrike",
        "Cybersecurity",
        "https://www.crowdstrike.com/en-us/careers/",
        "https://crowdstrike.wd5.myworkdayjobs.com/crowdstrikecareers",
        "Workday",
        source_identifier="crowdstrike | wd5 | crowdstrikecareers",
        india_jobs="Yes",
        notes="Workday public endpoints are tenant-specific and undocumented.",
    ),
    source(
        "Fortinet",
        "Cybersecurity",
        "https://www.fortinet.com/corporate/careers",
        "https://www.fortinet.com/corporate/careers",
        "Company-specific / Oracle candidate",
        india_jobs="Yes",
    ),
    source(
        "Zscaler",
        "Cloud security",
        "https://www.zscaler.com/careers",
        "https://www.zscaler.com/careers/search",
        "Company-specific / Workday candidate",
        india_jobs="Yes",
    ),
    source(
        "Nutanix",
        "Hybrid multicloud software",
        "https://www.nutanix.com/company/careers",
        "https://careers.nutanix.com/",
        "Company-specific / Phenom candidate",
        india_jobs="Yes",
    ),
    source(
        "Broadcom",
        "Semiconductors & infrastructure software",
        "https://www.broadcom.com/company/careers",
        "https://broadcom.wd1.myworkdayjobs.com/External_Career",
        "Workday",
        source_identifier="broadcom | wd1 | External_Career",
        india_jobs="Yes",
        notes="Workday public endpoints are tenant-specific and undocumented.",
    ),
    source(
        "Intuit",
        "Financial software",
        "https://www.intuit.com/careers/",
        "https://jobs.intuit.com/search-jobs",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Workday",
        "Enterprise cloud software",
        "https://www.workday.com/en-us/company/careers.html",
        "https://workday.wd5.myworkdayjobs.com/Workday",
        "Workday",
        source_identifier="workday | wd5 | Workday",
        india_jobs="Yes",
        notes="Workday public endpoints are tenant-specific and undocumented.",
    ),
    source(
        "HubSpot",
        "CRM & marketing software",
        "https://www.hubspot.com/careers",
        "https://www.hubspot.com/careers/jobs/all",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Freshworks",
        "Business software",
        "https://www.freshworks.com/company/careers/",
        "https://careers.smartrecruiters.com/Freshworks",
        "SmartRecruiters",
        source_identifier="Freshworks",
        public_endpoint="https://api.smartrecruiters.com/v1/companies/Freshworks/postings",
        api_key_required="No for public postings",
        india_jobs="Yes",
    ),
    source(
        "Zoho",
        "Business software",
        "https://www.zoho.com/careers/",
        "https://careers.zohocorp.com/jobs",
        "Zoho Recruit / company-specific",
        source_identifier="zohocorp",
        india_jobs="Yes",
    ),
    source(
        "Razorpay",
        "Payments technology",
        "https://razorpay.com/careers/",
        "https://job-boards.greenhouse.io/razorpaysoftwareprivatelimited",
        "Greenhouse",
        source_identifier="razorpaysoftwareprivatelimited",
        public_endpoint=(
            "https://boards-api.greenhouse.io/v1/boards/"
            "razorpaysoftwareprivatelimited/jobs?content=true"
        ),
        api_key_required="No",
        india_jobs="Yes",
    ),
    source(
        "PhonePe",
        "Payments technology",
        "https://www.phonepe.com/careers/",
        "https://www.phonepe.com/careers/job-openings/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Flipkart",
        "E-commerce technology",
        "https://www.flipkartcareers.com/",
        "https://www.flipkartcareers.com/jobslist",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Swiggy",
        "Consumer internet & delivery",
        "https://careers.swiggy.com/",
        "https://careers.swiggy.com/list.html",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Zomato",
        "Consumer internet & food technology",
        "https://www.zomato.com/careers",
        "https://www.zomato.com/careers",
        "Company-specific",
        india_jobs="Yes",
        notes="The careers page may collect candidate interest instead of exposing a complete structured job feed.",
    ),
    source(
        "Meesho",
        "E-commerce technology",
        "https://www.meesho.io/jobs",
        "https://www.meesho.io/jobs",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "CRED",
        "Consumer financial technology",
        "https://careers.cred.club/",
        "https://careers.cred.club/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "BrowserStack",
        "Software testing platform",
        "https://www.browserstack.com/careers",
        "https://browserstack.hire.trakstar.com/jobs",
        "Trakstar Hire",
        source_identifier="browserstack.hire.trakstar.com",
        india_jobs="Yes",
    ),
    source(
        "Postman",
        "API platform",
        "https://www.postman.com/company/careers/",
        "https://www.postman.com/company/careers/open-positions/",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Chargebee",
        "Subscription billing software",
        "https://www.chargebee.com/careers/",
        "https://jobs.chargebee.com/",
        "SAP SuccessFactors",
        source_identifier="jobs.chargebee.com",
        india_jobs="Yes",
    ),
    source(
        "Druva",
        "Data security software",
        "https://www.druva.com/why-druva/explore/careers",
        "https://www.druva.com/why-druva/explore/careers/jobs",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "InMobi",
        "Advertising technology",
        "https://www.inmobi.com/company/careers",
        "https://www.inmobi.com/company/careers",
        "Company-specific",
        india_jobs="Yes",
    ),
    source(
        "Dream Sports (Dream11)",
        "Sports technology",
        "https://www.dreamsports.group/lifeatdreamsports",
        "https://www.dreamsports.group/lifeatdreamsports",
        "Company-specific",
        india_jobs="Yes",
        notes="Use the official Dream Sports careers page for roles across Dream11 and group products.",
    ),
]


def simple_source(
    company: str,
    sector: str,
    careers_url: str,
    jobs_url: str = "",
    source_type: str = "Company-specific / auto-detect",
    *,
    priority: str = "High",
    india_jobs: str = "Yes",
    source_identifier: str = "",
    public_endpoint: str = "",
    api_key_required: str = "N/A",
    notes: str = "",
) -> CompanySource:
    return source(
        company,
        sector,
        careers_url,
        jobs_url or careers_url,
        source_type,
        priority=priority,
        india_jobs=india_jobs,
        source_identifier=source_identifier,
        public_endpoint=public_endpoint,
        api_key_required=api_key_required,
        notes=notes,
    )


NEW_MNC_COMPANIES = [
    simple_source(
        "ABB",
        "Industrial automation & electrification",
        "https://careers.abb/global/en",
        "https://careers.abb/global/en/search-results",
        "Phenom",
    ),
    simple_source(
        "Schneider Electric",
        "Energy management & industrial automation",
        "https://www.se.com/ww/en/about-us/careers/",
    ),
    simple_source(
        "Philips",
        "Health technology",
        "https://www.careers.philips.com/",
        "https://www.careers.philips.com/professional/global/en/search-results",
        "Phenom",
    ),
    simple_source(
        "GE Aerospace",
        "Aerospace technology",
        "https://careers.geaerospace.com/global/en",
        "https://careers.geaerospace.com/global/en/search-results",
        "Phenom",
    ),
    simple_source(
        "GE HealthCare",
        "Health technology",
        "https://careers.gehealthcare.com/global/en",
        "https://careers.gehealthcare.com/global/en/search-results",
        "Phenom",
    ),
    simple_source(
        "Hitachi",
        "Industrial & digital technology",
        "https://careers.hitachi.com/",
    ),
    simple_source(
        "Toyota",
        "Automotive technology",
        "https://careers.toyotabharat.com/",
        "https://careers.toyotabharat.com/",
        "Company-hosted careers page / manual application",
    ),
    simple_source(
        "Mercedes-Benz Group",
        "Automotive technology",
        "https://group.mercedes-benz.com/careers/",
        "https://jobs.mercedes-benz.com/en/",
    ),
    simple_source(
        "BMW Group",
        "Automotive technology",
        "https://www.bmwgroup.jobs/",
    ),
    simple_source(
        "Volvo Group",
        "Commercial transport technology",
        "https://www.volvogroup.com/en/careers.html",
        "https://jobs.volvogroup.com/",
        "SAP SuccessFactors",
    ),
    simple_source(
        "Caterpillar",
        "Industrial machinery",
        "https://careers.caterpillar.com/",
        "https://careers.caterpillar.com/en/jobs/",
    ),
    simple_source(
        "3M",
        "Industrial & consumer products",
        "https://www.3m.com/3M/en_US/careers-us/",
        india_jobs="Global search",
    ),
    simple_source(
        "Unilever",
        "Consumer products",
        "https://careers.unilever.com/",
        "https://careers.unilever.com/search-jobs",
    ),
    simple_source(
        "Procter & Gamble",
        "Consumer products",
        "https://www.pgcareers.com/global/en",
        "https://www.pgcareers.com/global/en/search-results",
        "Phenom",
    ),
    simple_source(
        "Nestle",
        "Food & consumer products",
        "https://www.nestle.com/jobs",
        "https://www.nestle.com/jobs/search-jobs",
    ),
    simple_source(
        "PepsiCo",
        "Food & beverages",
        "https://www.pepsicojobs.com/main/",
        "https://www.pepsicojobs.com/main/jobs",
    ),
    simple_source(
        "The Coca-Cola Company",
        "Beverages",
        "https://careers.coca-colacompany.com/",
    ),
    simple_source(
        "Shell",
        "Energy",
        "https://www.shell.com/careers.html",
        "https://jobs.shell.com/",
    ),
    simple_source(
        "bp",
        "Energy",
        "https://www.bp.com/en/global/corporate/careers.html",
        "https://careers.bp.com/",
    ),
    simple_source(
        "ExxonMobil",
        "Energy",
        "https://jobs.exxonmobil.com/",
    ),
    simple_source(
        "Airbus",
        "Aerospace",
        "https://www.airbus.com/en/careers",
        "https://ag.wd3.myworkdayjobs.com/Airbus",
        "Workday",
        source_identifier="ag | wd3 | Airbus",
    ),
    simple_source(
        "Boeing",
        "Aerospace",
        "https://jobs.boeing.com/",
        "https://jobs.boeing.com/search-jobs",
    ),
    simple_source(
        "Johnson & Johnson",
        "Healthcare & medical products",
        "https://www.careers.jnj.com/",
        "https://www.careers.jnj.com/en/jobs/",
    ),
    simple_source(
        "Pfizer",
        "Pharmaceuticals",
        "https://www.pfizer.com/about/careers",
        "https://pfizer.wd1.myworkdayjobs.com/PfizerCareers",
        "Workday",
        source_identifier="pfizer | wd1 | PfizerCareers",
    ),
    simple_source(
        "AstraZeneca",
        "Pharmaceuticals",
        "https://careers.astrazeneca.com/",
        "https://careers.astrazeneca.com/search-jobs",
    ),
]


NEW_PRODUCT_COMPANIES = [
    simple_source(
        "AMD",
        "Semiconductors & computing",
        "https://careers.amd.com/careers-home",
        "https://careers.amd.com/careers-home/jobs",
    ),
    simple_source(
        "Arm",
        "Semiconductor IP",
        "https://careers.arm.com/",
        "https://careers.arm.com/search-jobs",
    ),
    simple_source(
        "Micron Technology",
        "Memory & storage semiconductors",
        "https://careers.micron.com/",
        "https://careers.micron.com/careers",
    ),
    simple_source(
        "Texas Instruments",
        "Semiconductors",
        "https://careers.ti.com/",
        "https://careers.ti.com/search-jobs",
    ),
    simple_source(
        "Analog Devices",
        "Semiconductors",
        "https://careers.analog.com/",
    ),
    simple_source(
        "Autodesk",
        "Design & engineering software",
        "https://www.autodesk.com/careers",
        "https://autodesk.wd1.myworkdayjobs.com/Ext",
        "Workday",
        source_identifier="autodesk | wd1 | Ext",
    ),
    simple_source(
        "eBay",
        "E-commerce marketplace",
        "https://jobs.ebayinc.com/",
        "https://jobs.ebayinc.com/us/en/search-results",
    ),
    simple_source(
        "Booking.com",
        "Travel technology",
        "https://careers.booking.com/",
        "https://jobs.booking.com/booking/jobs",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Expedia Group",
        "Travel technology",
        "https://careers.expediagroup.com/",
        "https://careers.expediagroup.com/jobs/",
    ),
    simple_source(
        "DoorDash",
        "Local commerce technology",
        "https://careersatdoordash.com/",
        "https://careersatdoordash.com/job-search/",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    simple_source(
        "Pinterest",
        "Consumer internet",
        "https://www.pinterestcareers.com/",
        "https://www.pinterestcareers.com/jobs/",
        india_jobs="Global search",
    ),
    simple_source(
        "Snap",
        "Consumer internet & AR",
        "https://careers.snap.com/",
        "https://careers.snap.com/jobs",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    simple_source(
        "Coinbase",
        "Cryptocurrency platform",
        "https://www.coinbase.com/careers",
        "https://www.coinbase.com/careers/positions",
        priority="Medium",
        india_jobs="Global search",
    ),
    simple_source(
        "Robinhood",
        "Financial technology",
        "https://careers.robinhood.com/",
        "https://careers.robinhood.com/",
        priority="Medium",
        india_jobs="Limited / varies",
    ),
    simple_source(
        "Grab",
        "Mobility & financial technology",
        "https://www.grab.careers/en/",
        "https://www.grab.careers/en/jobs/",
        india_jobs="Limited / varies",
    ),
    simple_source(
        "Akamai Technologies",
        "Cloud delivery & security",
        "https://www.akamai.com/careers",
        "https://akamaicareers.inflightcloud.com/",
    ),
]


NEW_STARTUP_COMPANIES = [
    simple_source(
        "Zepto",
        "Quick commerce",
        "https://www.zepto.com/",
        "https://www.linkedin.com/company/zeptonow/jobs",
        "Employer-managed LinkedIn jobs fallback",
        notes="No standalone public employer career portal was confirmed; do not scrape protected LinkedIn pages.",
    ),
    simple_source(
        "Groww",
        "Investment technology",
        "https://groww.in/careers",
        "https://job-boards.greenhouse.io/groww",
        "Greenhouse",
        source_identifier="groww",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/groww/jobs?content=true",
        api_key_required="No",
    ),
    simple_source(
        "Zerodha",
        "Investment technology",
        "https://zerodha.com/careers/",
    ),
    simple_source(
        "Urban Company",
        "Services marketplace technology",
        "https://careers.urbancompany.com/",
        "https://careers.urbancompany.com/jobs",
    ),
    simple_source(
        "Ather Energy",
        "Electric mobility technology",
        "https://careers.atherenergy.com/",
        "https://careers.atherenergy.com/jobs",
    ),
    simple_source(
        "Rapido",
        "Mobility technology",
        "https://www.rapido.bike/Careers",
    ),
    simple_source(
        "Delhivery",
        "Logistics technology",
        "https://www.delhivery.com/careers",
    ),
    simple_source(
        "Nykaa",
        "Beauty & commerce technology",
        "https://careers.nykaa.com/",
    ),
    simple_source(
        "Lenskart",
        "Retail & vision technology",
        "https://hiring.lenskart.com/",
    ),
    simple_source(
        "Pine Labs",
        "Payments technology",
        "https://www.pinelabs.com/careers",
    ),
    simple_source(
        "BharatPe",
        "Financial technology",
        "https://bharatpe.com/career/",
    ),
    simple_source(
        "CoinDCX",
        "Cryptocurrency technology",
        "https://careers.coindcx.com/",
        "https://careers.coindcx.com/opportunities",
    ),
    simple_source(
        "CoinSwitch",
        "Financial technology",
        "https://coinswitch.co/",
        "https://in.linkedin.com/company/coinswitch/jobs",
        "Employer-managed LinkedIn jobs fallback",
        notes="No live first-party careers page was confirmed; re-check the official site before enabling an adapter.",
    ),
    simple_source(
        "Innovaccer",
        "Healthcare data platform",
        "https://innovaccer.com/careers",
        "https://innovaccer.com/careers/jobs",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Physics Wallah",
        "Education technology",
        "https://www.pw.live/life",
        notes="Official Life at PW page includes career opportunities; do not use similarly named unofficial recruiting domains.",
    ),
    simple_source(
        "Sarvam AI",
        "Generative AI",
        "https://www.sarvam.ai/careers",
        priority="Medium",
    ),
    simple_source(
        "Observe.AI",
        "Contact-center AI",
        "https://www.observe.ai/careers",
        "https://www.observe.ai/careers#open-positions",
    ),
    simple_source(
        "Yellow.ai",
        "Conversational AI",
        "https://yellow.ai/career/",
    ),
    simple_source(
        "Uniphore",
        "Enterprise AI",
        "https://www.uniphore.com/careers/",
    ),
    simple_source(
        "Porter",
        "Logistics technology",
        "https://porter.in/careers",
    ),
]


NEW_MID_SIZED_COMPANIES = [
    simple_source(
        "Cloudera",
        "Data platform",
        "https://www.cloudera.com/careers.html",
        "https://cloudera.wd5.myworkdayjobs.com/External_Career",
        "Workday",
        source_identifier="cloudera | wd5 | External_Career",
    ),
    simple_source(
        "New Relic",
        "Observability software",
        "https://newrelic.com/about/careers",
        "https://newrelic.wd5.myworkdayjobs.com/External",
        "Workday",
        source_identifier="newrelic | wd5 | External",
    ),
    simple_source(
        "Zendesk",
        "Customer service software",
        "https://jobs.zendesk.com/us/en",
    ),
    simple_source(
        "Rubrik",
        "Data security software",
        "https://www.rubrik.com/company/careers",
        "https://www.rubrik.com/company/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Cohesity",
        "Data security software",
        "https://www.cohesity.com/careers/",
        "https://careers.cohesity.com/open-positions/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Pure Storage",
        "Enterprise storage",
        "https://www.purestorage.com/company/careers.html",
        "https://job-boards.greenhouse.io/purestorage",
        "Greenhouse",
        source_identifier="purestorage",
        public_endpoint="https://boards-api.greenhouse.io/v1/boards/purestorage/jobs?content=true",
        api_key_required="No",
    ),
    simple_source(
        "Informatica",
        "Enterprise data management",
        "https://careers.informatica.com/",
    ),
    simple_source(
        "Pegasystems",
        "Enterprise workflow software",
        "https://www.pega.com/about/careers",
        "https://www.pega.com/about/careers/search",
    ),
    simple_source(
        "UiPath",
        "Automation software",
        "https://careers.uipath.com/careers",
    ),
    simple_source(
        "Automation Anywhere",
        "Automation software",
        "https://www.automationanywhere.com/company/careers",
    ),
    simple_source(
        "Celonis",
        "Process intelligence software",
        "https://www.celonis.com/careers/",
        "https://www.celonis.com/careers/jobs/",
    ),
    simple_source(
        "Alteryx",
        "Analytics automation software",
        "https://www.alteryx.com/careers",
        "https://alteryx.wd5.myworkdayjobs.com/AlteryxCareers",
        "Workday",
        source_identifier="alteryx | wd5 | AlteryxCareers",
    ),
    simple_source(
        "Fivetran",
        "Data integration platform",
        "https://www.fivetran.com/careers",
    ),
    simple_source(
        "dbt Labs",
        "Analytics engineering platform",
        "https://www.getdbt.com/about-us/careers",
        "https://www.getdbt.com/about-us/careers",
        "Company-hosted jobs portal",
        india_jobs="Global search",
        notes="Use the live first-party careers page; the former Greenhouse board API currently returns 404 and must not be treated as stable.",
    ),
    simple_source(
        "Redis",
        "Database platform",
        "https://redis.io/company/careers/current-job-openings/",
        "https://redis.io/company/careers/current-job-openings/",
        "Company-hosted jobs portal",
        india_jobs="Global search",
        notes="Use the live first-party openings page; the former Greenhouse board root now returns 404 and is retained only as historical discovery evidence.",
    ),
    simple_source(
        "Grafana Labs",
        "Observability platform",
        "https://grafana.com/about/careers/",
        "https://grafana.com/about/careers/open-positions/",
        india_jobs="Global search",
    ),
    simple_source(
        "Cockroach Labs",
        "Distributed database platform",
        "https://www.cockroachlabs.com/careers/",
        india_jobs="Global search",
    ),
    simple_source(
        "Canva",
        "Design software",
        "https://www.lifeatcanva.com/",
        "https://www.lifeatcanva.com/en/jobs/",
        india_jobs="Global search",
    ),
    simple_source(
        "Miro",
        "Collaboration software",
        "https://miro.com/careers/",
        "https://miro.com/careers/open-positions/",
        india_jobs="Global search",
    ),
]


OTHER_COMPANIES = [
    simple_source(
        "Fractal Analytics",
        "AI & analytics services",
        "https://fractal.ai/careers/",
    ),
    simple_source(
        "Tiger Analytics",
        "AI & analytics services",
        "https://www.tigeranalytics.com/careers/",
    ),
    simple_source(
        "Tredence",
        "Data science & analytics services",
        "https://www.tredence.com/careers/greatest-of-ai",
    ),
    simple_source(
        "Mu Sigma",
        "Decision sciences",
        "https://www.mu-sigma.com/careers/",
    ),
    simple_source(
        "ZS Associates",
        "Management consulting & analytics",
        "https://www.zs.com/careers",
        "https://jobs.zs.com/",
    ),
    simple_source(
        "EXL",
        "Analytics & operations",
        "https://www.exlservice.com/careers",
    ),
    simple_source(
        "LatentView Analytics",
        "Data analytics",
        "https://www.latentview.com/career/",
    ),
    simple_source(
        "Quantiphi",
        "AI & cloud engineering",
        "https://quantiphi.com/careers/",
    ),
    simple_source(
        "Sigmoid",
        "Data engineering & AI",
        "https://www.sigmoid.com/careers/",
    ),
    simple_source(
        "Course5 Intelligence",
        "Analytics & AI",
        "https://www.course5i.com/careers/",
    ),
    simple_source(
        "Gramener",
        "Data science & visualization",
        "https://gramener.com/careers/",
    ),
    simple_source(
        "MathCo",
        "Data analytics & AI",
        "https://mathco.com/careers/",
    ),
    simple_source(
        "Searce",
        "Cloud & AI consulting",
        "https://www.searce.com/join-us",
    ),
    simple_source(
        "Nagarro",
        "Digital engineering",
        "https://www.nagarro.com/en/careers",
    ),
    simple_source(
        "Globant",
        "Digital engineering",
        "https://career.globant.com/",
    ),
    simple_source(
        "Epsilon",
        "Marketing technology",
        "https://www.epsilon.com/us/careers",
    ),
    simple_source(
        "D. E. Shaw India",
        "Quantitative finance & technology",
        "https://www.deshawindia.com/careers/",
    ),
    simple_source(
        "Arcesium",
        "Financial technology",
        "https://www.arcesium.com/careers",
    ),
    simple_source(
        "Optiver",
        "Market making & technology",
        "https://optiver.com/working-at-optiver/career-opportunities/",
    ),
    simple_source(
        "Tower Research Capital",
        "Quantitative trading & technology",
        "https://www.tower-research.com/open-positions",
    ),
]


INDIA_PRODUCT_EXPANSION = [
    simple_source(
        "Paytm",
        "Payments & financial technology",
        "https://paytm.com/careers",
        "https://jobs.lever.co/paytm",
        "Lever",
        source_identifier="paytm",
        public_endpoint="https://api.lever.co/v0/postings/paytm?mode=json",
        api_key_required="No",
    ),
    simple_source(
        "MakeMyTrip",
        "Travel technology",
        "https://careers.makemytrip.com/",
        "https://careers.makemytrip.com/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "IndiaMART",
        "B2B marketplace technology",
        "https://careers.indiamart.com/",
        "https://careers.smartrecruiters.com/Indiamart1",
        "SmartRecruiters",
        source_identifier="Indiamart1",
        public_endpoint="https://api.smartrecruiters.com/v1/companies/Indiamart1/postings",
        api_key_required="No",
    ),
    simple_source(
        "Jio Platforms",
        "Telecommunications & digital products",
        "https://careers.jio.com/",
        "https://careers.jio.com/",
        "Company-specific",
    ),
    simple_source(
        "MapmyIndia",
        "Maps, geospatial & IoT products",
        "https://www.mapmyindia.com/careers/",
        "https://www.mapmyindia.com/careers/",
        "Company-hosted careers page / manual application",
        priority="Medium",
        notes="The official page currently emphasizes direct resume submission; review openings manually before applying.",
    ),
    simple_source(
        "Juspay",
        "Payments infrastructure",
        "https://juspay.io/careers",
        "https://juspay.io/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Clear",
        "Tax, payments & compliance technology",
        "https://www.clear.in/s/careers",
        "https://www.clear.in/s/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Games24x7",
        "Gaming, AI & data products",
        "https://www.games24x7.com/life",
        "https://www.games24x7.com/life",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "BookMyShow",
        "Entertainment ticketing technology",
        "https://in.bookmyshow.com/careers",
        "https://careers.smartrecruiters.com/BookMyShow",
        "SmartRecruiters",
        source_identifier="BookMyShow",
        public_endpoint="https://api.smartrecruiters.com/v1/companies/BookMyShow/postings",
        api_key_required="No",
    ),
    simple_source(
        "MobiKwik",
        "Payments & consumer financial technology",
        "https://www.mobikwik.com/careers/",
        "https://www.mobikwik.com/careers/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Policybazaar",
        "Insurance & financial technology",
        "https://www.policybazaar.com/careers/",
        "https://policybazaar.hire.trakstar.com/jobs",
        "Trakstar Hire",
        source_identifier="policybazaar",
        notes="The public Trakstar page is company-specific and should remain a manual adapter unless separately qualified.",
    ),
    simple_source(
        "Practo",
        "Digital healthcare products",
        "https://www.practo.com/company/careers",
        "https://practo.app.param.ai/jobs/",
        "Param AI / company-specific",
        source_identifier="practo.app.param.ai",
        notes="The official Practo careers route redirects to its public Param AI jobs experience; treat it as a company-specific adapter.",
    ),
]


INDIA_STARTUP_EXPANSION = [
    simple_source(
        "Krutrim AI Labs",
        "Foundation models & AI infrastructure",
        "https://ai-labs.olakrutrim.com/",
        "https://ai-labs.olakrutrim.com/",
        "Company-hosted careers section / manual application",
        notes="Use the official Join Us section and verify each opening manually.",
    ),
    simple_source(
        "Qure.ai",
        "Healthcare AI",
        "https://jobs.qure.ai/",
        "https://qure.zohorecruit.in/jobs/careers",
        "Zoho Recruit",
        source_identifier="qure.zohorecruit.in",
    ),
    simple_source(
        "Pixxel",
        "Earth-observation & space data",
        "https://www.pixxel.space/careers",
        "https://www.pixxel.space/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Skyroot Aerospace",
        "Space launch technology",
        "https://www.skyroot.in/careers",
        "https://skyroot.zohorecruit.in/jobs/Careers",
        "Zoho Recruit",
        source_identifier="skyroot.zohorecruit.in",
    ),
    simple_source(
        "Agnikul Cosmos",
        "Space launch technology",
        "https://agnikul.in/careers/",
        "https://agnikul.in/careers/",
        "Company-hosted careers page / manual application",
        notes="The official page publishes roles but currently requests applications by email; no automatic submission is permitted.",
    ),
    simple_source(
        "Atlan",
        "Data catalog & AI context platform",
        "https://atlan.com/careers/",
        "https://atlan.com/careers/?p=open-positions",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Neysa",
        "AI cloud & infrastructure",
        "https://neysa.ai/careers/",
        "https://neysa.ai/careers/job-openings/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Ema",
        "Enterprise agentic AI",
        "https://www.ema.ai/careers",
        "https://www.ema.ai/careers/jobs",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "SpotDraft",
        "AI contract lifecycle management",
        "https://www.spotdraft.com/careers",
        "https://spotdraft.freshteam.com/jobs",
        "Freshteam",
        source_identifier="spotdraft",
    ),
    simple_source(
        "Sprinto",
        "Security compliance automation",
        "https://sprinto.com/about-us/",
        "https://jobs.lever.co/Sprinto",
        "Lever",
        source_identifier="Sprinto",
        public_endpoint="https://api.lever.co/v0/postings/Sprinto?mode=json",
        api_key_required="No",
    ),
    simple_source(
        "SuperOps",
        "AI-native IT management",
        "https://superops.com/careers",
        "https://superops.com/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Rocketlane",
        "Customer onboarding software",
        "https://www.rocketlane.com/",
        "https://careers.kula.ai/rocketlane",
        "Kula",
        source_identifier="rocketlane",
    ),
]


INDIA_MID_SIZED_EXPANSION = [
    simple_source(
        "Whatfix",
        "Digital adoption software",
        "https://whatfix.com/careers",
        "https://whatfix101.hire.trakstar.com/jobs",
        "Trakstar Hire",
        source_identifier="whatfix101",
    ),
    simple_source(
        "Icertis",
        "Contract intelligence software",
        "https://www.icertis.com/company/careers/",
        "https://www.icertis.com/company/careers/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Mindtickle",
        "Revenue enablement software",
        "https://www.mindtickle.com/careers/",
        "https://careers.smartrecruiters.com/mindtickle",
        "SmartRecruiters",
        source_identifier="mindtickle",
        public_endpoint="https://api.smartrecruiters.com/v1/companies/mindtickle/postings",
        api_key_required="No",
    ),
    simple_source(
        "CleverTap",
        "Customer engagement software",
        "https://clevertap.com/careers/",
        "https://careers.kula.ai/clevertap",
        "Kula",
        source_identifier="clevertap",
    ),
    simple_source(
        "MoEngage",
        "Customer engagement software",
        "https://www.moengage.com/careers/",
        "https://www.moengage.com/careers/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Gupshup",
        "Conversational AI & messaging",
        "https://www.gupshup.ai/careers",
        "https://www.gupshup.ai/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Amagi",
        "Cloud broadcast & media technology",
        "https://www.amagi.com/careers",
        "https://www.amagi.com/careers",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "Hasura",
        "Data API & AI development platform",
        "https://hasura.io/careers/",
        "https://promptql.io/careers",
        "Company-hosted jobs portal",
        notes="The former Hasura careers URL currently redirects to PromptQL's official careers page.",
    ),
    simple_source(
        "LeadSquared",
        "Sales execution & CRM software",
        "https://www.leadsquared.com/careers/",
        "https://leadsquaredhrms.darwinbox.in/ms/candidatev2/main/careers",
        "Darwinbox",
        source_identifier="leadsquaredhrms.darwinbox.in",
    ),
    simple_source(
        "Darwinbox",
        "Human capital management software",
        "https://darwinbox.com/en-us/careers",
        "https://darwinbox.com/en-us/careers",
        "Company-hosted JavaScript careers page",
    ),
    simple_source(
        "Wingify",
        "Digital experience optimization software",
        "https://wingify.com/careers/",
        "https://wingify.com/careers/",
        "Company-hosted jobs portal",
    ),
    simple_source(
        "GreyOrange",
        "Warehouse robotics software",
        "https://www.greyorange.com/careers/",
        "https://www.greyorange.com/company/careers/",
        "Company-hosted jobs portal",
    ),
]


_ALL_EXISTING_BY_NAME = {
    company.company: company for company in COMPANIES + PRODUCT_COMPANIES
}
_STARTUP_FROM_PRODUCT = {
    "Figure AI",
    "Razorpay",
    "Meesho",
    "CRED",
    "Dream Sports (Dream11)",
}
_MID_SIZED_FROM_PRODUCT = {
    "Discord",
    "BrowserStack",
    "Postman",
    "Chargebee",
    "Druva",
    "InMobi",
}
_PRODUCT_NAMES = {company.company for company in PRODUCT_COMPANIES}

MNC_COMPANIES = [
    company for company in COMPANIES if company.company not in _PRODUCT_NAMES
] + NEW_MNC_COMPANIES
ESTABLISHED_PRODUCT_COMPANIES = [
    company
    for company in PRODUCT_COMPANIES
    if company.company not in _STARTUP_FROM_PRODUCT | _MID_SIZED_FROM_PRODUCT
] + NEW_PRODUCT_COMPANIES + INDIA_PRODUCT_EXPANSION
STARTUP_COMPANIES = [
    _ALL_EXISTING_BY_NAME[name] for name in sorted(_STARTUP_FROM_PRODUCT)
] + NEW_STARTUP_COMPANIES + INDIA_STARTUP_EXPANSION
MID_SIZED_COMPANIES = [
    _ALL_EXISTING_BY_NAME[name] for name in sorted(_MID_SIZED_FROM_PRODUCT)
] + NEW_MID_SIZED_COMPANIES + INDIA_MID_SIZED_EXPANSION

CATEGORY_REGISTRIES = {
    "MNC": MNC_COMPANIES,
    "Product Companies": ESTABLISHED_PRODUCT_COMPANIES,
    "Startups": STARTUP_COMPANIES,
    "Mid-Sized Companies": MID_SIZED_COMPANIES,
    "Other Companies": OTHER_COMPANIES,
}

CATEGORY_DEFINITIONS = {
    "MNC": "Large multinational services, consulting, industrial, banking, healthcare, and consumer employers with India-relevant hiring.",
    "Product Companies": "Established product-led software, cloud, semiconductor, fintech, and consumer-technology employers.",
    "Startups": "Selected high-growth or founder-led companies with strong technology, data, AI, or digital-product roles.",
    "Mid-Sized Companies": "Established specialist product companies generally smaller than the largest global product employers.",
    "Other Companies": "Selected analytics, digital engineering, AI consulting, and quantitative-technology employers relevant to the search.",
}


DOCUMENTED_AUTOMATION_SOURCES = {
    "greenhouse",
    "lever",
    "workable",
    "smartrecruiters",
}
MANUAL_SOURCE_MARKERS = (
    "workday",
    "oracle recruiting",
    "successfactors",
    "icims",
    "taleo",
    "phenom",
    "eightfold",
    "avature",
    "njoyn",
    "darwinbox",
    "zoho recruit",
    "freshteam",
    "trakstar",
    "kula",
    "talentrecruit",
    "ibegin",
    "inflight",
    "param ai",
    "manual application",
    "linkedin jobs",
    "javascript careers page",
)


def requires_manual_approach(company: CompanySource) -> bool:
    source_type = company.source_type.casefold().strip()
    if source_type in DOCUMENTED_AUTOMATION_SOURCES:
        return False
    return any(marker in source_type for marker in MANUAL_SOURCE_MARKERS)


def check_url(url: str) -> tuple[str, str]:
    if not url:
        return "Inaccessible (missing URL)", ""
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/126 Safari/537.36"
            )
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=15) as response:
            code = getattr(response, "status", 200)
            final_url = response.geturl()
        final_path = urlparse(final_url).path.casefold()
        error_paths = ("/error/", "/errors/", "page-not-found", "/404")
        if any(marker in final_path for marker in error_paths):
            return "Inaccessible (redirected to error page)", final_url
        return f"Accessible (HTTP {code})", final_url
    except urllib.error.HTTPError as exc:
        if exc.code in {400, 401, 403, 406, 429, 451}:
            return (
                f"Manual required (HTTP {exc.code}; automated access restricted)",
                url,
            )
        if exc.code in {404, 410}:
            return f"Inaccessible (HTTP {exc.code})", url
        return f"Manual required (HTTP {exc.code}; verify in browser)", url
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        raw_reason = str(getattr(exc, "reason", exc)).replace("\n", " ")
        if "CERTIFICATE_VERIFY_FAILED" in raw_reason:
            reason = "TLS certificate check failed"
        elif "timed out" in raw_reason.lower():
            reason = "connection timed out"
        elif "getaddrinfo failed" in raw_reason.lower():
            return "Inaccessible (DNS check failed)", url
        else:
            reason = raw_reason[:90]
        return f"Manual required ({reason})", url


def validate_urls(urls: list[str]) -> dict[str, tuple[str, str]]:
    results: dict[str, tuple[str, str]] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        future_urls = {executor.submit(check_url, url): url for url in urls}
        for future in concurrent.futures.as_completed(future_urls):
            url = future_urls[future]
            results[url] = future.result()
    return results


def validate_sources(companies: list[CompanySource]) -> dict[str, tuple[str, str]]:
    urls = sorted({item.jobs_url for item in companies if item.jobs_url})
    results = validate_urls(urls)
    for company in companies:
        status, final_url = results.get(
            company.jobs_url,
            ("Inaccessible (not checked)", company.jobs_url),
        )
        if status.startswith("Accessible") and requires_manual_approach(company):
            results[company.jobs_url] = (
                "Manual required (public page; company-specific adapter)",
                final_url,
            )
    return results


def append_redirect_note(notes: str, original_url: str, final_url: str) -> str:
    if not final_url:
        return notes
    original_host = urlparse(original_url).netloc.lower()
    final_host = urlparse(final_url).netloc.lower()
    if not final_host or original_host == final_host:
        return notes
    redirect_note = f"Automated check redirected to {final_host}."
    return f"{notes} {redirect_note}".strip()


def _add_verification_formatting(sheet, final_row: int) -> None:
    """Make dead links red and company-specific/manual sources blue by whole row."""
    row_range = f"A5:O{final_row}"
    status_range = f"M5:M{final_row}"
    sheet.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=['LEFT($M5,12)="Inaccessible"'],
            fill=PatternFill("solid", fgColor="FEE2E2"),
            stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=['LEFT($M5,15)="Manual required"'],
            fill=PatternFill("solid", fgColor="DBEAFE"),
            stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['LEFT($M5,10)="Accessible"'],
            fill=PatternFill("solid", fgColor="DCFCE7"),
        ),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['$M5="Not checked"'],
            fill=PatternFill("solid", fgColor="E5E7EB"),
        ),
    )


def _build_mnc_workbook(
    output_path: Path,
    *,
    validation_results: dict[str, tuple[str, str]] | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MNC"
    sheet.sheet_view.showGridLines = False

    headers = [
        "Company",
        "Sector",
        "Priority",
        "Official Careers Page",
        "Direct Job Portal",
        "ATS / Source Type",
        "Source Identifier",
        "Public Jobs API / Feed",
        "API Key Required",
        "India Jobs",
        "Active",
        "Last Checked",
        "Verification Status",
        "Fallback",
        "Notes",
    ]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(1, 1, "MNC Career Portal Registry")
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0F766E")
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    note = sheet.cell(
        2,
        1,
        (
            f"Starter registry of {len(COMPANIES)} major India-relevant and reference MNCs. "
            "Use official sources first; ATS mappings marked candidate require confirmation. "
            "Greenhouse updated_at is not always the original publication date."
        ),
    )
    note.font = Font(name="Aptos", size=10, italic=True, color="334155")
    note.fill = PatternFill("solid", fgColor="E2E8F0")
    note.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 34

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    summary = sheet.cell(
        3,
        1,
        (
            "Daily path: registry -> direct API/feed or portal -> local date/keyword filtering "
            "-> optional LLM scoring only for shortlisted jobs."
        ),
    )
    summary.font = Font(name="Aptos", size=10, color="115E59")
    summary.fill = PatternFill("solid", fgColor="CCFBF1")
    summary.alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 24

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 32

    checked_on = date.today()
    for row_index, company in enumerate(COMPANIES, start=5):
        status = "Not checked"
        final_url = ""
        if validation_results is not None:
            status, final_url = validation_results.get(
                company.jobs_url,
                ("Manual required (not checked)", company.jobs_url),
            )
        notes = append_redirect_note(company.notes, company.jobs_url, final_url)
        values = [
            company.company,
            company.sector,
            company.priority,
            company.careers_url,
            company.jobs_url,
            company.source_type,
            company.source_identifier,
            company.public_endpoint,
            company.api_key_required,
            company.india_jobs,
            "Yes",
            checked_on if validation_results is not None else None,
            status,
            company.fallback,
            notes,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in (4, 5, 8):
            cell = sheet.cell(row_index, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row_index, 12).number_format = "yyyy-mm-dd"
        sheet.row_dimensions[row_index].height = 72

    final_row = 4 + len(COMPANIES)
    table = Table(displayName="MNCRegistry", ref=f"A4:O{final_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:O{final_row}"

    widths = {
        "A": 30,
        "B": 27,
        "C": 12,
        "D": 43,
        "E": 52,
        "F": 31,
        "G": 34,
        "H": 58,
        "I": 19,
        "J": 18,
        "K": 11,
        "L": 14,
        "M": 42,
        "N": 44,
        "O": 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    thin_gray = Side(style="thin", color="CBD5E1")
    sheet.cell(4, 1).border = Border(left=thin_gray, top=thin_gray, bottom=thin_gray)
    sheet.cell(4, len(headers)).border = Border(right=thin_gray, top=thin_gray, bottom=thin_gray)

    validations = [
        ("C5:C500", '"High,Medium,Reference"'),
        ("J5:J500", '"Yes,Global search,Limited / varies"'),
        ("K5:K500", '"Yes,No"'),
    ]
    for cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a value from the list."
        validation.errorTitle = "Invalid value"
        sheet.add_data_validation(validation)
        validation.add(cell_range)

    _add_verification_formatting(sheet, final_row)

    comments = {
        5: "Detected or researched ATS. Values containing 'candidate' must be confirmed before an adapter is enabled.",
        7: "Board token, company identifier, Workday tenant/site, or other stable source key when known.",
        8: "Only public endpoints are listed. Undocumented endpoints carry higher maintenance risk.",
        12: "Date the portal URL was last checked, not the date a job was published.",
        13: "Automated reachability check. A blocked result does not necessarily mean the public browser page is unavailable.",
    }
    for column, text in comments.items():
        sheet.cell(4, column).comment = Comment(text, "User")

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:4"
    sheet.oddFooter.center.text = "MNC Career Portal Registry"
    sheet.oddFooter.right.text = "Page &P of &N"

    workbook.properties.title = "MNC Career Portal Registry"
    workbook.properties.subject = "Official career pages, job portals, and ATS sources"
    workbook.properties.creator = "Personal Job Hunt"
    workbook.properties.description = (
        "Reusable company-source registry for direct job discovery and reduced repeated web research."
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _add_product_company_sheet(
    output_path: Path,
    *,
    validation_results: dict[str, tuple[str, str]] | None = None,
) -> None:
    if len(PRODUCT_COMPANIES) != len(COMPANIES):
        raise ValueError(
            "The template-copy workflow expects the MNC and product registries to "
            "contain the same number of starter rows."
        )

    workbook = load_workbook(output_path)
    product_sheet = workbook.copy_worksheet(workbook["MNC"])
    product_sheet.title = "Product Companies"
    product_sheet["A1"] = "Product-Based Company Career Portal Registry"
    product_sheet["A2"] = (
        f"Curated starter registry of {len(PRODUCT_COMPANIES)} product-led employers "
        "relevant to software, data, AI, cloud, fintech, and consumer-tech searches. "
        "The list is expandable, not exhaustive; candidate ATS mappings require confirmation."
    )
    product_sheet["A3"] = (
        "Daily path: product-company registry -> direct API/feed or portal -> local "
        "date/keyword filtering -> optional LLM scoring only for shortlisted jobs."
    )

    checked_on = date.today()
    for row_index, company in enumerate(PRODUCT_COMPANIES, start=5):
        status = "Not checked"
        final_url = ""
        if validation_results is not None:
            status, final_url = validation_results.get(
                company.jobs_url,
                ("Manual required (not checked)", company.jobs_url),
            )
        notes = append_redirect_note(company.notes, company.jobs_url, final_url)
        values = [
            company.company,
            company.sector,
            company.priority,
            company.careers_url,
            company.jobs_url,
            company.source_type,
            company.source_identifier,
            company.public_endpoint,
            company.api_key_required,
            company.india_jobs,
            "Yes",
            checked_on if validation_results is not None else None,
            status,
            company.fallback,
            notes,
        ]
        for column, value in enumerate(values, start=1):
            cell = product_sheet.cell(row_index, column)
            cell.value = value
            cell.hyperlink = None
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in (4, 5, 8):
            cell = product_sheet.cell(row_index, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        product_sheet.cell(row_index, 12).number_format = "yyyy-mm-dd"

    for table_name in list(product_sheet.tables):
        del product_sheet.tables[table_name]
    final_row = 4 + len(PRODUCT_COMPANIES)
    table = Table(displayName="ProductCompanyRegistry", ref=f"A4:O{final_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    product_sheet.add_table(table)
    product_sheet.freeze_panes = "A5"
    product_sheet.auto_filter.ref = f"A4:O{final_row}"

    if not product_sheet.data_validations.dataValidation:
        validations = [
            ("C5:C500", '"High,Medium,Reference"'),
            ("J5:J500", '"Yes,Global search,Limited / varies"'),
            ("K5:K500", '"Yes,No"'),
        ]
        for cell_range, formula in validations:
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            validation.error = "Choose a value from the list."
            validation.errorTitle = "Invalid value"
            product_sheet.add_data_validation(validation)
            validation.add(cell_range)

    if not product_sheet.conditional_formatting:
        _add_verification_formatting(product_sheet, final_row)

    product_sheet.oddFooter.center.text = "Product-Based Company Career Portal Registry"
    workbook.properties.title = "Company Career Portal Registry"
    workbook.properties.subject = (
        "Official career pages, direct job portals, and ATS sources for MNC and "
        "product-company discovery"
    )
    workbook.save(output_path)


def _build_legacy_two_tab_workbook(
    output_path: Path,
    *,
    validation_results: dict[str, tuple[str, str]] | None = None,
) -> None:
    _build_mnc_workbook(output_path, validation_results=validation_results)
    _add_product_company_sheet(output_path, validation_results=validation_results)


def _verify_legacy_two_tab_workbook(output_path: Path) -> None:
    workbook = load_workbook(output_path, data_only=False)
    if workbook.sheetnames != ["MNC", "Product Companies"]:
        raise ValueError(f"Unexpected worksheets: {workbook.sheetnames}")
    sheet_specs = [
        ("MNC", COMPANIES, "MNCRegistry"),
        ("Product Companies", PRODUCT_COMPANIES, "ProductCompanyRegistry"),
    ]
    for sheet_name, companies, table_name in sheet_specs:
        sheet = workbook[sheet_name]
        expected_rows = len(companies) + 4
        if sheet.max_row != expected_rows or sheet.max_column != 15:
            raise ValueError(
                f"Unexpected dimensions on {sheet_name}: "
                f"{sheet.max_row}x{sheet.max_column}; expected {expected_rows}x15"
            )
        if sheet["A4"].value != "Company" or sheet["E4"].value != "Direct Job Portal":
            raise ValueError(f"Header verification failed on {sheet_name}")
        if not sheet.tables or table_name not in sheet.tables:
            raise ValueError(f"{table_name} table is missing")
        missing_links = []
        for row in range(5, sheet.max_row + 1):
            for column in (4, 5, 8):
                cell = sheet.cell(row, column)
                if cell.value and not cell.hyperlink:
                    missing_links.append(cell.coordinate)
        if missing_links:
            raise ValueError(
                f"Missing hyperlinks on {sheet_name}: {missing_links[:10]}"
            )
        formula_errors = []
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(
                    token in cell.value
                    for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    formula_errors.append(cell.coordinate)
        if formula_errors:
            raise ValueError(
                f"Formula errors found on {sheet_name}: {formula_errors[:10]}"
            )


REGISTRY_HEADERS = [
    "Company",
    "Sector",
    "Priority",
    "Official Careers Page",
    "Direct Job Portal",
    "ATS / Source Type",
    "Source Identifier",
    "Public Jobs API / Feed",
    "API Key Required",
    "India Jobs",
    "Active",
    "Last Checked",
    "Verification Status",
    "Fallback",
    "Notes",
]

TABLE_NAMES = {
    "MNC": "MNCRegistry",
    "Product Companies": "ProductRegistry",
    "Startups": "StartupRegistry",
    "Mid-Sized Companies": "MidSizedRegistry",
    "Other Companies": "OtherRegistry",
}

PORTAL_ALERT_SHEET_NAME = "Portal Alert Pilot"
PORTAL_ALERT_TABLE_NAME = "PortalAlertPilot"
PORTAL_ALERT_AUDIT_DATE = date(2026, 8, 1)
PORTAL_ALERT_HEADERS = [
    "Company",
    "Registry Category",
    "Referral Connections",
    "Official Job Portal",
    "Alert Supported",
    "Alert Method",
    "Login Required",
    "Google Sign-In",
    "Alert Setup URL",
    "Alert Query",
    "Target Location",
    "Experience Target",
    "Alert Enabled",
    "Last Checked",
    "Confirmed Relevant Jobs",
    "Best Match",
    "Best Match Experience",
    "Best Match Location",
    "Best Match URL",
    "Manual Action Needed",
    "Evidence / Notes",
]

PORTAL_ALERT_PILOT = [
    {
        "company": "Wipro",
        "category": "MNC",
        "jobs_url": "https://careers.wipro.com/search/",
        "alert_supported": "Yes",
        "alert_method": "Wipro careers job alerts / talent community",
        "login_required": "Wipro careers profile required",
        "google_sign_in": "No",
        "alert_url": "https://careers.wipro.com/talentcommunity/subscribe/?locale=en_US",
        "query": "machine learning, AI/ML, GenAI, data science",
        "location": "India; Hyderabad or Bengaluru",
        "experience": "5-8 years",
        "enabled": "Pending browser/account",
        "confirmed_jobs": 0,
        "best_match": "Machine Learning Engineer (nearby range)",
        "best_experience": "3-5 years",
        "best_location": "Bengaluru",
        "best_url": "https://careers.wipro.com/job/Machine-Learning-Engineer/171358-en_US/",
        "manual_action": "Connect browser, create or sign in to a Wipro careers profile, then confirm and enable the alert.",
        "notes": "Two nearby live roles were found, but no exact 5-8-year match was confirmed. A search-indexed AI Engineer role at 5-8 years was expired when opened.",
    },
    {
        "company": "Cognizant",
        "category": "MNC",
        "jobs_url": "https://careers.cognizant.com/global-en/jobs/",
        "alert_supported": "Yes",
        "alert_method": "India talent community; profile-based notifications",
        "login_required": "Talent-community registration form",
        "google_sign_in": "Not confirmed",
        "alert_url": "https://careers.cognizant.com/india-en/talent-community/",
        "query": "machine learning, AI/ML, GenAI, data science",
        "location": "India; Hyderabad or Bengaluru",
        "experience": "5-8 years",
        "enabled": "Pending browser/form",
        "confirmed_jobs": 1,
        "best_match": "Sr Associate - Data Science",
        "best_experience": "6-12 years (overlaps 6-8)",
        "best_location": "Hyderabad / Bengaluru",
        "best_url": "https://careers.cognizant.com/global-en/jobs/00067230902/sr-associate-data-science/",
        "manual_action": "Connect browser, review the India talent-community form, and submit only after confirming its fields.",
        "notes": "The official live role overlaps the target range. A separate 4-7-year AI/ML developer result still needs interactive portal confirmation.",
    },
    {
        "company": "Infosys",
        "category": "MNC",
        "jobs_url": "https://career.infosys.com/joblist",
        "alert_supported": "Not confirmed for India",
        "alert_method": "Interactive job portal; Gmail/app scanner fallback",
        "login_required": "Not confirmed",
        "google_sign_in": "Not confirmed",
        "alert_url": "",
        "query": "machine learning, AI/ML, GenAI, data science",
        "location": "India; Hyderabad or Bengaluru",
        "experience": "5-8 years",
        "enabled": "Not confirmed",
        "confirmed_jobs": None,
        "best_match": "Interactive portal review required",
        "best_experience": "",
        "best_location": "India",
        "best_url": "",
        "manual_action": "Connect browser to confirm current India results and whether the portal offers a native alert after sign-in.",
        "notes": "The official India portal is JavaScript-driven. Public indexing did not provide enough evidence to claim a current match or native India alert.",
    },
    {
        "company": "Accenture",
        "category": "MNC",
        "jobs_url": "https://www.accenture.com/in-en/careers/jobsearch",
        "alert_supported": "No native alert confirmed",
        "alert_method": "Saved job plus daily official-portal scanner fallback",
        "login_required": "No for public search; account may be needed to save/apply",
        "google_sign_in": "Not confirmed",
        "alert_url": "",
        "query": "machine learning, AI/ML, GenAI, data science",
        "location": "India; Hyderabad or Bengaluru",
        "experience": "5-8 years",
        "enabled": "No native alert confirmed",
        "confirmed_jobs": 1,
        "best_match": "AI / ML Engineer",
        "best_experience": "5-10 years; minimum 7.5",
        "best_location": "Hyderabad",
        "best_url": "https://www.accenture.com/in-en/careers/jobdetails?id=ATCI-4822102-S1848464_en&title=AI+%2F+ML+Engineer",
        "manual_action": "Use the daily official-portal scanner or Gmail fallback; optionally save the confirmed role after login.",
        "notes": "One strong official live match was confirmed. The public search exposes Saved Jobs, but no native job-alert control was confirmed.",
    },
    {
        "company": "Google",
        "category": "Product Companies",
        "jobs_url": "https://www.google.com/about/careers/applications/jobs/results/",
        "alert_supported": "Yes",
        "alert_method": "Google Careers job alerts",
        "login_required": "Google account required",
        "google_sign_in": "Yes",
        "alert_url": "https://www.google.com/about/careers/applications/jobs/alerts",
        "query": "machine learning, AI/ML, GenAI, data science",
        "location": "India; Hyderabad or Bengaluru",
        "experience": "5-8 years",
        "enabled": "Pending browser/Google sign-in",
        "confirmed_jobs": 3,
        "best_match": "Senior Software Engineer, AI/ML, AI Garage",
        "best_experience": "5 years minimum; 7 preferred",
        "best_location": "Hyderabad",
        "best_url": "https://www.google.com/about/careers/applications/jobs/results/136390105332884166-senior-software-engineer-aiml-ai-garage",
        "manual_action": "Connect a browser signed in to the chosen Google account, then create and verify the alert.",
        "notes": "Three official current roles fit or closely overlap the target, including Hyderabad AI/ML and Bengaluru DeepMind roles.",
    },
]


def _validate_catalog() -> None:
    if list(CATEGORY_REGISTRIES) != list(CATEGORY_DEFINITIONS):
        raise ValueError("Category order and definitions are inconsistent")
    seen: dict[str, str] = {}
    duplicates: list[str] = []
    for category, companies in CATEGORY_REGISTRIES.items():
        if not companies:
            raise ValueError(f"{category} has no companies")
        for company in companies:
            normalized = company.company.casefold().strip()
            if normalized in seen:
                duplicates.append(
                    f"{company.company} ({seen[normalized]} and {category})"
                )
            seen[normalized] = category
            if not company.careers_url.startswith("https://"):
                raise ValueError(f"Invalid careers URL for {company.company}")
            if not company.jobs_url.startswith("https://"):
                raise ValueError(f"Invalid jobs URL for {company.company}")
    if duplicates:
        raise ValueError(f"Duplicate category assignments: {duplicates}")
    category_by_company = {
        company.company: category
        for category, companies in CATEGORY_REGISTRIES.items()
        for company in companies
    }
    pilot_names: set[str] = set()
    for item in PORTAL_ALERT_PILOT:
        company = str(item["company"])
        if company in pilot_names:
            raise ValueError(f"Duplicate portal-alert pilot company: {company}")
        pilot_names.add(company)
        if category_by_company.get(company) != item["category"]:
            raise ValueError(f"Portal-alert pilot category mismatch: {company}")


def _add_portal_alert_pilot_sheet(
    workbook: Workbook,
    linkedin_data: LinkedInExportData | None,
) -> None:
    sheet = workbook.create_sheet(PORTAL_ALERT_SHEET_NAME)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 75
    last_column = len(PORTAL_ALERT_HEADERS)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title = sheet.cell(1, 1, "Official Portal & Job Alert Pilot")
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0F766E")
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    note = sheet.cell(
        2,
        1,
        (
            "Five-company pilot for AI/ML roles targeting 5-8 years in India, with "
            "Hyderabad and Bengaluru prioritized. Public evidence was checked on "
            "2026-08-01; no alert is marked enabled without an interactive confirmation."
        ),
    )
    note.font = Font(name="Aptos", size=10, italic=True, color="334155")
    note.fill = PatternFill("solid", fgColor="E2E8F0")
    note.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 42

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    summary = sheet.cell(
        3,
        1,
        (
            "Alert support, Google sign-in, and a public job listing are separate facts. "
            "Pending rows require a connected browser and, where applicable, user-only "
            "authentication before the alert can be enabled."
        ),
    )
    summary.font = Font(name="Aptos", size=10, color="115E59")
    summary.fill = PatternFill("solid", fgColor="CCFBF1")
    summary.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 36

    for column, header in enumerate(PORTAL_ALERT_HEADERS, start=1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 48

    referral_counts: dict[str, int] = {}
    if linkedin_data is not None:
        for connection in linkedin_data.connections:
            company = str(connection.get("registry_company") or "")
            if company:
                referral_counts[company] = referral_counts.get(company, 0) + 1

    for row_index, item in enumerate(PORTAL_ALERT_PILOT, start=5):
        company = str(item["company"])
        values = [
            company,
            item["category"],
            referral_counts.get(company, 0),
            item["jobs_url"],
            item["alert_supported"],
            item["alert_method"],
            item["login_required"],
            item["google_sign_in"],
            item["alert_url"],
            item["query"],
            item["location"],
            item["experience"],
            item["enabled"],
            PORTAL_ALERT_AUDIT_DATE,
            item["confirmed_jobs"],
            item["best_match"],
            item["best_experience"],
            item["best_location"],
            item["best_url"],
            item["manual_action"],
            item["notes"],
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in (4, 9, 19):
            cell = sheet.cell(row_index, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row_index, 14).number_format = "yyyy-mm-dd"
        sheet.row_dimensions[row_index].height = 90

    final_row = len(PORTAL_ALERT_PILOT) + 4
    table = Table(
        displayName=PORTAL_ALERT_TABLE_NAME,
        ref=f"A4:{get_column_letter(last_column)}{final_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "D5"

    widths = [
        20, 20, 16, 46, 24, 38, 32, 18, 46, 34, 31, 18, 30, 14, 19,
        42, 27, 25, 50, 50, 58,
    ]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    validations = [
        (f"E5:E{final_row}", '"Yes,No native alert confirmed,Not confirmed for India"'),
        (f"H5:H{final_row}", '"Yes,No,Not confirmed"'),
        (
            f"M5:M{final_row}",
            '"Enabled,Pending browser/account,Pending browser/form,Pending browser/Google sign-in,No native alert confirmed,Not confirmed"',
        ),
    ]
    for cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a value from the list."
        validation.errorTitle = "Invalid value"
        sheet.add_data_validation(validation)
        validation.add(cell_range)

    status_range = f"M5:M{final_row}"
    for formula, color in [
        ('M5="Enabled"', "DCFCE7"),
        ('LEFT(M5,7)="Pending"', "FEF3C7"),
        ('LEFT(M5,9)="No native"', "FEE2E2"),
        ('M5="Not confirmed"', "E5E7EB"),
    ]:
        sheet.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
        )

    comments = {
        3: "Counted from registry-matched LinkedIn connections in the supplied export; blank exports produce zero.",
        5: "Whether the official employer portal exposes a native alert or talent-community notification flow.",
        8: "Whether the alert flow specifically supports Google-account sign-in, not whether a Gmail address can be typed into a form.",
        13: "Only set to Enabled after the portal confirms the alert in an interactive browser session.",
        15: "Confirmed current official roles fitting or overlapping the 5-8-year target; blank means the dynamic portal could not be audited conclusively.",
    }
    for column, comment_text in comments.items():
        sheet.cell(4, column).comment = Comment(comment_text, "User")

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:4"
    sheet.oddFooter.center.text = "Official Portal & Job Alert Pilot"
    sheet.oddFooter.right.text = "Page &P of &N"


def _add_registry_sheet(
    workbook: Workbook,
    sheet_name: str,
    companies: list[CompanySource],
    *,
    validation_results: dict[str, tuple[str, str]] | None,
) -> None:
    sheet = workbook[sheet_name]
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 80

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=15)
    title = sheet.cell(1, 1, f"{sheet_name} Career Portal Registry")
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0F766E")
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
    note = sheet.cell(
        2,
        1,
        (
            f"{CATEGORY_DEFINITIONS[sheet_name]} {len(companies)} companies; "
            "each company appears in one category only. Candidate ATS mappings must "
            "be confirmed before an adapter is enabled."
        ),
    )
    note.font = Font(name="Aptos", size=10, italic=True, color="334155")
    note.fill = PatternFill("solid", fgColor="E2E8F0")
    note.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 42

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=15)
    summary = sheet.cell(
        3,
        1,
        (
            "Daily path: stored API/feed or official portal -> local date/keyword "
            "filtering -> optional LLM scoring only for shortlisted jobs."
        ),
    )
    summary.font = Font(name="Aptos", size=10, color="115E59")
    summary.fill = PatternFill("solid", fgColor="CCFBF1")
    summary.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 26

    for column, header in enumerate(REGISTRY_HEADERS, start=1):
        cell = sheet.cell(4, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 36

    checked_on = date.today()
    for row_index, company in enumerate(companies, start=5):
        status = "Not checked"
        final_url = ""
        if validation_results is not None:
            status, final_url = validation_results.get(
                company.jobs_url,
                ("Manual required (not checked)", company.jobs_url),
            )
        notes = append_redirect_note(company.notes, company.jobs_url, final_url)
        values = [
            company.company,
            company.sector,
            company.priority,
            company.careers_url,
            company.jobs_url,
            company.source_type,
            company.source_identifier,
            company.public_endpoint,
            company.api_key_required,
            company.india_jobs,
            "Yes",
            checked_on if validation_results is not None else None,
            status,
            company.fallback,
            notes,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in (4, 5, 8):
            cell = sheet.cell(row_index, column)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row_index, 12).number_format = "yyyy-mm-dd"
        sheet.row_dimensions[row_index].height = 68

    final_row = len(companies) + 4
    table = Table(
        displayName=TABLE_NAMES[sheet_name],
        ref=f"A4:O{final_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A5"

    widths = {
        "A": 30,
        "B": 27,
        "C": 12,
        "D": 43,
        "E": 52,
        "F": 31,
        "G": 34,
        "H": 58,
        "I": 19,
        "J": 18,
        "K": 11,
        "L": 14,
        "M": 42,
        "N": 44,
        "O": 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    thin_gray = Side(style="thin", color="CBD5E1")
    sheet.cell(4, 1).border = Border(
        left=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )
    sheet.cell(4, 15).border = Border(
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    validations = [
        (f"C5:C{final_row}", '"High,Medium,Reference"'),
        (f"J5:J{final_row}", '"Yes,Global search,Limited / varies"'),
        (f"K5:K{final_row}", '"Yes,No"'),
    ]
    for cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a value from the list."
        validation.errorTitle = "Invalid value"
        sheet.add_data_validation(validation)
        validation.add(cell_range)

    _add_verification_formatting(sheet, final_row)

    comments = {
        5: "Detected or researched ATS. Values containing 'candidate' must be confirmed before an adapter is enabled.",
        7: "Board token, company identifier, Workday tenant/site, or another stable source key when known.",
        8: "Only public endpoints are listed. Undocumented endpoints carry higher maintenance risk.",
        12: "Date the portal URL was last checked, not the date a job was published.",
        13: "Automated reachability check. A restricted result does not mean the public browser page is unavailable.",
    }
    for column, comment_text in comments.items():
        sheet.cell(4, column).comment = Comment(comment_text, "User")

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:4"
    sheet.oddFooter.center.text = f"{sheet_name} Career Portal Registry"
    sheet.oddFooter.right.text = "Page &P of &N"


def _populate_coverage_sheet(workbook: Workbook) -> None:
    sheet = workbook["Coverage"]
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 90
    sheet.merge_cells("A1:G1")
    title = sheet["A1"]
    title.value = "Company Registry Coverage & Classification"
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0F766E")
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells("A2:G2")
    note = sheet["A2"]
    note.value = (
        "Scope: a broad India-relevant job-search universe, not every employer in the "
        "world. Classification is mutually exclusive so a company appears in one list only."
    )
    note.font = Font(name="Aptos", size=10, italic=True, color="334155")
    note.fill = PatternFill("solid", fgColor="E2E8F0")
    note.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 40

    sheet.merge_cells("A3:G3")
    summary = sheet["A3"]
    summary.value = (
        "Coverage check: unique assignments + official careers/job URLs + live portal "
        "reachability. Re-run validation periodically because career systems change."
    )
    summary.font = Font(name="Aptos", size=10, color="115E59")
    summary.fill = PatternFill("solid", fgColor="CCFBF1")
    summary.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 28

    headers = [
        "Category",
        "Definition",
        "Company Count",
        "Accessible",
        "Manual Required",
        "Inaccessible",
        "Duplicate Rule",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(5, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[5].height = 36

    for row_index, (category, companies) in enumerate(
        CATEGORY_REGISTRIES.items(),
        start=6,
    ):
        last_row = len(companies) + 4
        sheet.cell(row_index, 1, category)
        sheet.cell(row_index, 2, CATEGORY_DEFINITIONS[category])
        sheet.cell(
            row_index,
            3,
            f"=COUNTA('{category}'!$A$5:$A${last_row})",
        )
        sheet.cell(
            row_index,
            4,
            f'=COUNTIF(\'{category}\'!$M$5:$M${last_row},"Accessible*")',
        )
        sheet.cell(
            row_index,
            5,
            f'=COUNTIF(\'{category}\'!$M$5:$M${last_row},"Manual required*")',
        )
        sheet.cell(
            row_index,
            6,
            f'=COUNTIF(\'{category}\'!$M$5:$M${last_row},"Inaccessible*")',
        )
        sheet.cell(row_index, 7, "One primary category per company")
        for column in range(1, 8):
            cell = sheet.cell(row_index, column)
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 56

    total_row = 12
    sheet.cell(total_row, 1, "Total unique companies")
    sheet.cell(total_row, 3, "=SUM(C6:C10)")
    sheet.cell(total_row, 4, "=SUM(D6:D10)")
    sheet.cell(total_row, 5, "=SUM(E6:E10)")
    sheet.cell(total_row, 6, "=SUM(F6:F10)")
    for column in range(1, 8):
        cell = sheet.cell(total_row, column)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[total_row].height = 28

    sheet.merge_cells("A14:G14")
    scope = sheet["A14"]
    scope.value = (
        "Coverage statement: all companies included in this workbook were checked for "
        "unique classification and an HTTPS careers/job source. This is not a claim that "
        "every MNC or product company worldwide is listed."
    )
    scope.font = Font(name="Aptos", size=10, italic=True, color="7C2D12")
    scope.fill = PatternFill("solid", fgColor="FFEDD5")
    scope.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[14].height = 44

    widths = {
        "A": 24,
        "B": 78,
        "C": 16,
        "D": 16,
        "E": 18,
        "F": 16,
        "G": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A6"
    sheet.auto_filter.ref = "A5:G10"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.oddFooter.center.text = "Company Registry Coverage"


def build_workbook(
    output_path: Path,
    *,
    validation_results: dict[str, tuple[str, str]] | None = None,
    linkedin_data: LinkedInExportData | None = None,
) -> None:
    _validate_catalog()
    workbook = Workbook()
    workbook.active.title = "Coverage"
    for sheet_name in CATEGORY_REGISTRIES:
        workbook.create_sheet(sheet_name)
    for sheet_name, companies in CATEGORY_REGISTRIES.items():
        _add_registry_sheet(
            workbook,
            sheet_name,
            companies,
            validation_results=validation_results,
        )
    _populate_coverage_sheet(workbook)
    _add_portal_alert_pilot_sheet(workbook, linkedin_data)
    if linkedin_data is not None:
        add_linkedin_sheets(workbook, linkedin_data)
    workbook.active = 0
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.title = "Company Career Portal Registry"
    workbook.properties.subject = (
        "Unique company categories with official career pages, job portals, and ATS sources"
    )
    workbook.properties.creator = "Personal Job Hunt"
    workbook.properties.description = (
        "A reusable source registry for direct job discovery with mutually exclusive "
        "company categories and reduced repeated web research."
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _verify_open_xml_filters(
    output_path: Path,
    *,
    expected_table_count: int,
) -> None:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(output_path) as archive:
        worksheet_entries = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        for worksheet_entry in worksheet_entries[1:]:
            root = ET.fromstring(archive.read(worksheet_entry))
            if root.find(f"{namespace}autoFilter") is not None:
                raise ValueError(
                    f"Conflicting worksheet AutoFilter found in {worksheet_entry}"
                )
        table_entries = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/tables/table") and name.endswith(".xml")
        )
        if len(table_entries) != expected_table_count:
            raise ValueError(
                f"Expected {expected_table_count} table parts, found {len(table_entries)}"
            )
        for table_entry in table_entries:
            root = ET.fromstring(archive.read(table_entry))
            if root.find(f"{namespace}autoFilter") is None:
                raise ValueError(f"Table filter missing from {table_entry}")


def verify_workbook(
    output_path: Path,
    *,
    linkedin_data: LinkedInExportData | None = None,
) -> None:
    _validate_catalog()
    workbook = load_workbook(output_path, data_only=False)
    expected_sheets = ["Coverage", *CATEGORY_REGISTRIES, PORTAL_ALERT_SHEET_NAME]
    if linkedin_data is not None:
        expected_sheets.extend(LINKEDIN_SHEET_NAMES)
    if workbook.sheetnames != expected_sheets:
        raise ValueError(f"Unexpected worksheets: {workbook.sheetnames}")

    seen: set[str] = set()
    for sheet_name, companies in CATEGORY_REGISTRIES.items():
        sheet = workbook[sheet_name]
        expected_rows = len(companies) + 4
        if sheet.max_row != expected_rows or sheet.max_column != 15:
            raise ValueError(
                f"Unexpected dimensions on {sheet_name}: "
                f"{sheet.max_row}x{sheet.max_column}; expected {expected_rows}x15"
            )
        if sheet.auto_filter.ref is not None:
            raise ValueError(f"Worksheet-level filter conflicts with table on {sheet_name}")
        table_name = TABLE_NAMES[sheet_name]
        if table_name not in sheet.tables:
            raise ValueError(f"{table_name} is missing")
        if sheet.tables[table_name].ref != f"A4:O{expected_rows}":
            raise ValueError(f"Unexpected table range on {sheet_name}")
        actual_names = [
            sheet.cell(row, 1).value for row in range(5, expected_rows + 1)
        ]
        expected_names = [company.company for company in companies]
        if actual_names != expected_names:
            raise ValueError(f"Company order mismatch on {sheet_name}")
        for company_name in actual_names:
            normalized = str(company_name).casefold().strip()
            if normalized in seen:
                raise ValueError(f"Duplicate company across tabs: {company_name}")
            seen.add(normalized)
        missing_links: list[str] = []
        for row in range(5, expected_rows + 1):
            for column in (4, 5, 8):
                cell = sheet.cell(row, column)
                if cell.value and not cell.hyperlink:
                    missing_links.append(cell.coordinate)
        if missing_links:
            raise ValueError(
                f"Missing hyperlinks on {sheet_name}: {missing_links[:10]}"
            )

    coverage = workbook["Coverage"]
    for row in range(6, 11):
        for column in range(3, 7):
            value = coverage.cell(row, column).value
            if not isinstance(value, str) or not value.startswith("="):
                raise ValueError(
                    f"Coverage formula missing at {coverage.cell(row, column).coordinate}"
                )
    if len(seen) != sum(len(items) for items in CATEGORY_REGISTRIES.values()):
        raise ValueError("Unique company total does not reconcile")

    pilot = workbook[PORTAL_ALERT_SHEET_NAME]
    expected_pilot_rows = len(PORTAL_ALERT_PILOT) + 4
    expected_pilot_ref = (
        f"A4:{get_column_letter(len(PORTAL_ALERT_HEADERS))}{expected_pilot_rows}"
    )
    if pilot.max_row != expected_pilot_rows or pilot.max_column != len(PORTAL_ALERT_HEADERS):
        raise ValueError("Unexpected Portal Alert Pilot dimensions")
    if pilot.auto_filter.ref is not None:
        raise ValueError("Worksheet-level filter conflicts with Portal Alert Pilot table")
    if PORTAL_ALERT_TABLE_NAME not in pilot.tables:
        raise ValueError("Portal Alert Pilot table is missing")
    if pilot.tables[PORTAL_ALERT_TABLE_NAME].ref != expected_pilot_ref:
        raise ValueError("Unexpected Portal Alert Pilot table range")
    if [pilot.cell(4, column).value for column in range(1, pilot.max_column + 1)] != PORTAL_ALERT_HEADERS:
        raise ValueError("Portal Alert Pilot headers are inconsistent")
    for row in range(5, expected_pilot_rows + 1):
        for column in (4, 9, 19):
            cell = pilot.cell(row, column)
            if cell.value and not cell.hyperlink:
                raise ValueError(f"Missing pilot hyperlink at {cell.coordinate}")
        if pilot.cell(row, 13).value == "Enabled":
            raise ValueError("Pilot alert cannot be marked enabled without UI confirmation")

    formula_errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(
                    token in cell.value
                    for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}")
    if formula_errors:
        raise ValueError(f"Formula errors found: {formula_errors[:10]}")
    if linkedin_data is not None:
        verify_linkedin_sheets(workbook, linkedin_data)
    expected_table_count = len(CATEGORY_REGISTRIES) + 1
    if linkedin_data is not None:
        expected_table_count += len(LINKEDIN_TABLE_NAMES)
    _verify_open_xml_filters(
        output_path,
        expected_table_count=expected_table_count,
    )


def _preview_font(size: int, *, bold: bool = False) -> ImageFont.FreeTypeFont:
    font_name = "arialbd.ttf" if bold else "arial.ttf"
    font_path = Path("C:/Windows/Fonts") / font_name
    try:
        return ImageFont.truetype(str(font_path), size=size)
    except OSError:
        return ImageFont.load_default()


def _draw_wrapped_text(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    text: object,
    *,
    font: ImageFont.ImageFont,
    fill: str,
    padding: int = 6,
) -> None:
    if isinstance(text, (date, datetime)):
        value = text.strftime("%Y-%m-%d")
    else:
        value = "" if text is None else str(text)
    x1, y1, x2, y2 = box
    usable_width = max(1, x2 - x1 - 2 * padding)
    average_character_width = max(5, int(getattr(font, "size", 10) * 0.55))
    wrapped = textwrap.wrap(
        value,
        width=max(4, usable_width // average_character_width),
        break_long_words=True,
        break_on_hyphens=False,
    )
    line_height = max(12, int(getattr(font, "size", 10) * 1.25))
    max_lines = max(1, (y2 - y1 - 2 * padding) // line_height)
    if len(wrapped) > max_lines:
        wrapped = wrapped[:max_lines]
        wrapped[-1] = wrapped[-1][:-1] + "…" if wrapped[-1] else "…"
    draw.multiline_text(
        (x1 + padding, y1 + padding),
        "\n".join(wrapped),
        font=font,
        fill=fill,
        spacing=2,
    )


def _preview_color(color, default: str) -> str:
    if color is not None and color.type == "rgb" and color.rgb:
        value = str(color.rgb)
        if len(value) == 8 and value[:2] == "00":
            return default
        return f"#{value[-6:]}"
    return default


def _preview_formula_value(workbook, formula: str) -> object:
    counta = re.fullmatch(
        r"=COUNTA\('([^']+)'!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)\)",
        formula,
    )
    if counta:
        sheet_name, start_column, start_row, end_column, end_row = counta.groups()
        sheet = workbook[sheet_name]
        return sum(
            1
            for row in sheet[f"{start_column}{start_row}:{end_column}{end_row}"]
            for cell in row
            if cell.value not in (None, "")
        )
    countif = re.fullmatch(
        r'=COUNTIF\(\'([^\']+)\'!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+),"([^"]*)"\)',
        formula,
    )
    if countif:
        sheet_name, start_column, start_row, end_column, end_row, criterion = (
            countif.groups()
        )
        sheet = workbook[sheet_name]
        values = [
            cell.value
            for row in sheet[f"{start_column}{start_row}:{end_column}{end_row}"]
            for cell in row
        ]
        if criterion == "<>":
            return sum(value not in (None, "") for value in values)
        return sum(str(value or "") == criterion for value in values)
    return formula


def render_generic_sheet_preview(
    output_path: Path,
    preview_path: Path,
    *,
    sheet_name: str,
) -> None:
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook[sheet_name]
    table_sheet = (
        sheet_name in LINKEDIN_TABLE_NAMES
        or sheet_name == PORTAL_ALERT_SHEET_NAME
    )
    row_limit = min(sheet.max_row, 16) if table_sheet else sheet.max_row
    column_limit = sheet.max_column
    column_widths: list[int] = []
    for column_number in range(1, column_limit + 1):
        column_letter = get_column_letter(column_number)
        excel_width = sheet.column_dimensions[column_letter].width or 12
        column_widths.append(max(80, min(310, int(excel_width * 7))))
    row_heights: list[int] = []
    for row_number in range(1, row_limit + 1):
        excel_height = sheet.row_dimensions[row_number].height or 22
        row_heights.append(max(22, min(120, int(excel_height * 1.34))))

    x_positions = [1]
    for width in column_widths:
        x_positions.append(x_positions[-1] + width)
    y_positions = [1]
    for height in row_heights:
        y_positions.append(y_positions[-1] + height)
    canvas_width = x_positions[-1] + 1
    canvas_height = y_positions[-1] + 1
    image = Image.new("RGB", (canvas_width, canvas_height), "#FFFFFF")
    draw = ImageDraw.Draw(image)

    merged_by_anchor = {}
    merged_interior = set()
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row > row_limit:
            continue
        anchor = (merged_range.min_row, merged_range.min_col)
        merged_by_anchor[anchor] = merged_range
        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            for column_number in range(
                merged_range.min_col,
                merged_range.max_col + 1,
            ):
                if (row_number, column_number) != anchor:
                    merged_interior.add((row_number, column_number))

    for row_number in range(1, row_limit + 1):
        for column_number in range(1, column_limit + 1):
            if (row_number, column_number) in merged_interior:
                continue
            cell = sheet.cell(row_number, column_number)
            merged_range = merged_by_anchor.get((row_number, column_number))
            end_row = (
                min(merged_range.max_row, row_limit)
                if merged_range is not None
                else row_number
            )
            end_column = (
                merged_range.max_col if merged_range is not None else column_number
            )
            box = (
                x_positions[column_number - 1],
                y_positions[row_number - 1],
                x_positions[end_column],
                y_positions[end_row],
            )
            default_fill = "#FFFFFF"
            if table_sheet and row_number >= 5 and row_number % 2 == 0:
                default_fill = "#F8FAFC"
            fill = _preview_color(cell.fill.fgColor, default_fill)
            draw.rectangle(box, fill=fill, outline="#CBD5E1")
            text_color = _preview_color(cell.font.color, "#1F2937")
            if cell.hyperlink:
                text_color = "#0563C1"
            font_size = int(cell.font.sz or 10)
            font = _preview_font(
                max(10, min(24, int(font_size * 1.15))),
                bold=bool(cell.font.bold),
            )
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                value = _preview_formula_value(workbook, value)
            _draw_wrapped_text(
                draw,
                box,
                value,
                font=font,
                fill=text_color,
                padding=7,
            )

    preview_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(preview_path)


def render_preview(output_path: Path, preview_path: Path, *, sheet_name: str) -> None:
    if sheet_name == "Coverage":
        render_coverage_preview(output_path, preview_path)
        return
    if sheet_name in LINKEDIN_SHEET_NAMES or sheet_name == PORTAL_ALERT_SHEET_NAME:
        render_generic_sheet_preview(
            output_path,
            preview_path,
            sheet_name=sheet_name,
        )
        return

    workbook = load_workbook(output_path, data_only=True)
    sheet = workbook[sheet_name]
    column_widths = [
        220,
        150,
        80,
        230,
        260,
        175,
        160,
        260,
        100,
        110,
        70,
        95,
        220,
        210,
        260,
    ]
    row_heights = [42, 58, 38, 46] + [72] * 11
    canvas_width = sum(column_widths) + 2
    canvas_height = sum(row_heights) + 2
    image = Image.new("RGB", (canvas_width, canvas_height), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    normal_font = _preview_font(12)
    header_font = _preview_font(12, bold=True)
    title_font = _preview_font(22, bold=True)
    note_font = _preview_font(12)

    y = 1
    for row_number, row_height in enumerate(row_heights, start=1):
        if row_number <= 3:
            fills = {1: "#0F766E", 2: "#E2E8F0", 3: "#CCFBF1"}
            colors = {1: "#FFFFFF", 2: "#334155", 3: "#115E59"}
            draw.rectangle(
                (1, y, canvas_width - 1, y + row_height),
                fill=fills[row_number],
                outline="#94A3B8",
            )
            _draw_wrapped_text(
                draw,
                (1, y, canvas_width - 1, y + row_height),
                sheet.cell(row_number, 1).value,
                font=title_font if row_number == 1 else note_font,
                fill=colors[row_number],
                padding=10,
            )
            y += row_height
            continue

        x = 1
        for column_number, column_width in enumerate(column_widths, start=1):
            if row_number == 4:
                fill = "#1E3A5F"
                text_color = "#FFFFFF"
                font = header_font
            else:
                fill = "#FFFFFF" if row_number % 2 else "#F8FAFC"
                text_color = "#1F2937"
                font = normal_font
                status = str(sheet.cell(row_number, 13).value or "")
                if status.startswith("Inaccessible"):
                    fill = "#FEE2E2"
                elif status.startswith("Manual required"):
                    fill = "#DBEAFE"
                elif column_number == 13 and status.startswith("Accessible"):
                    fill = "#DCFCE7"
                if column_number in {4, 5, 8} and sheet.cell(row_number, column_number).value:
                    text_color = "#0563C1"
            box = (x, y, x + column_width, y + row_height)
            draw.rectangle(box, fill=fill, outline="#CBD5E1")
            _draw_wrapped_text(
                draw,
                box,
                sheet.cell(row_number, column_number).value,
                font=font,
                fill=text_color,
            )
            x += column_width
        y += row_height

    preview_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(preview_path)


def render_coverage_preview(output_path: Path, preview_path: Path) -> None:
    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Coverage"]
    column_widths = [210, 660, 140, 140, 160, 140, 250]
    row_heights = [42, 58, 42, 16, 46] + [64] * 5 + [16, 34, 16, 58]
    canvas_width = sum(column_widths) + 2
    canvas_height = sum(row_heights) + 2
    image = Image.new("RGB", (canvas_width, canvas_height), "#FFFFFF")
    draw = ImageDraw.Draw(image)
    normal_font = _preview_font(12)
    header_font = _preview_font(12, bold=True)
    title_font = _preview_font(22, bold=True)
    note_font = _preview_font(12)

    status_counts: dict[str, tuple[int, int, int, int]] = {}
    for category, companies in CATEGORY_REGISTRIES.items():
        category_sheet = workbook[category]
        accessible = manual = inaccessible = 0
        for row in range(5, len(companies) + 5):
            status = str(category_sheet.cell(row, 13).value or "")
            if status.startswith("Accessible"):
                accessible += 1
            elif status.startswith("Manual required"):
                manual += 1
            elif status.startswith("Inaccessible"):
                inaccessible += 1
        status_counts[category] = (
            len(companies),
            accessible,
            manual,
            inaccessible,
        )

    total_counts = tuple(
        sum(counts[index] for counts in status_counts.values()) for index in range(4)
    )

    y = 1
    for row_number, row_height in enumerate(row_heights, start=1):
        if row_number in {1, 2, 3, 14}:
            fills = {1: "#0F766E", 2: "#E2E8F0", 3: "#CCFBF1", 14: "#FFEDD5"}
            colors = {1: "#FFFFFF", 2: "#334155", 3: "#115E59", 14: "#7C2D12"}
            draw.rectangle(
                (1, y, canvas_width - 1, y + row_height),
                fill=fills[row_number],
                outline="#94A3B8",
            )
            _draw_wrapped_text(
                draw,
                (1, y, canvas_width - 1, y + row_height),
                sheet.cell(row_number, 1).value,
                font=title_font if row_number == 1 else note_font,
                fill=colors[row_number],
                padding=10,
            )
            y += row_height
            continue
        if row_number in {4, 11, 13}:
            y += row_height
            continue

        x = 1
        for column_number, column_width in enumerate(column_widths, start=1):
            if row_number == 5:
                fill = "#1E3A5F"
                text_color = "#FFFFFF"
                font = header_font
                value = sheet.cell(row_number, column_number).value
            elif row_number == 12:
                fill = "#0F766E"
                text_color = "#FFFFFF"
                font = header_font
                value = sheet.cell(row_number, column_number).value
                if column_number in range(3, 7):
                    value = total_counts[column_number - 3]
            else:
                fill = "#FFFFFF" if row_number % 2 else "#F8FAFC"
                text_color = "#1F2937"
                font = normal_font
                value = sheet.cell(row_number, column_number).value
                if column_number in range(3, 7):
                    category = str(sheet.cell(row_number, 1).value)
                    value = status_counts[category][column_number - 3]
            box = (x, y, x + column_width, y + row_height)
            draw.rectangle(box, fill=fill, outline="#CBD5E1")
            _draw_wrapped_text(
                draw,
                box,
                value,
                font=font,
                fill=text_color,
            )
            x += column_width
        y += row_height

    preview_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(preview_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Check each public job portal once and store a reachability status.",
    )
    parser.add_argument(
        "--preview",
        type=Path,
        help="Render a compact PNG preview of the workbook's MNC sheet.",
    )
    parser.add_argument(
        "--product-preview",
        type=Path,
        help="Render a compact PNG preview of the Product Companies sheet.",
    )
    parser.add_argument(
        "--preview-dir",
        type=Path,
        help="Render a PNG preview of every worksheet into this directory.",
    )
    parser.add_argument(
        "--linkedin-export",
        type=Path,
        help=(
            "Optional LinkedIn data-export folder or ZIP. Only job-search-useful files "
            "are imported into privacy-conscious workbook tabs."
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    all_companies = [
        company
        for companies in CATEGORY_REGISTRIES.values()
        for company in companies
    ]
    linkedin_data = None
    if args.linkedin_export:
        registry_entries = [
            RegistryEntry(
                category=category,
                company=company.company,
                careers_url=company.careers_url,
                jobs_url=company.jobs_url,
            )
            for category, companies in CATEGORY_REGISTRIES.items()
            for company in companies
        ]
        linkedin_data = load_linkedin_export(
            args.linkedin_export,
            registry_entries,
        )
    validation_results = validate_sources(all_companies) if args.validate else None
    endpoint_results = None
    if args.validate:
        endpoints = sorted(
            {
                company.public_endpoint
                for company in all_companies
                if company.public_endpoint
            }
        )
        endpoint_results = validate_urls(endpoints)
    build_workbook(
        args.output,
        validation_results=validation_results,
        linkedin_data=linkedin_data,
    )
    verify_workbook(args.output, linkedin_data=linkedin_data)
    if args.preview:
        render_preview(args.output, args.preview, sheet_name="MNC")
    if args.product_preview:
        render_preview(
            args.output,
            args.product_preview,
            sheet_name="Product Companies",
        )
    if args.preview_dir:
        preview_sheets = ["Coverage", *CATEGORY_REGISTRIES, PORTAL_ALERT_SHEET_NAME]
        if linkedin_data is not None:
            preview_sheets.extend(LINKEDIN_SHEET_NAMES)
        for sheet_name in preview_sheets:
            safe_name = (
                sheet_name.casefold()
                .replace(" ", "_")
                .replace("-", "_")
            )
            render_preview(
                args.output,
                args.preview_dir / f"{safe_name}.png",
                sheet_name=sheet_name,
            )
    print(f"Created {args.output}")
    for category, companies in CATEGORY_REGISTRIES.items():
        print(f"{category}: {len(companies)}")
    print(f"Total unique companies: {len(all_companies)}")
    if linkedin_data is not None:
        print(f"LinkedIn connections: {len(linkedin_data.connections)}")
        print(
            "LinkedIn target-company connections: "
            f"{sum(bool(item['registry_company']) for item in linkedin_data.connections)}"
        )
        print(
            "LinkedIn exported connection emails: "
            f"{sum(bool(item['email']) for item in linkedin_data.connections)}"
        )
        print(f"LinkedIn applications: {len(linkedin_data.applications)}")
        print(f"LinkedIn saved jobs: {len(linkedin_data.saved_jobs)}")
        print(f"LinkedIn job alerts: {len(linkedin_data.job_alerts)}")
        print(f"LinkedIn followed companies: {len(linkedin_data.followed_companies)}")
    if validation_results is not None:
        counts: dict[str, int] = {}
        for status, _ in validation_results.values():
            category = status.split(" (")[0]
            counts[category] = counts.get(category, 0) + 1
        for category, count in sorted(counts.items()):
            print(f"{category}: {count}")
    if endpoint_results is not None:
        endpoint_counts: dict[str, int] = {}
        for status, _ in endpoint_results.values():
            category = status.split(" (")[0]
            endpoint_counts[category] = endpoint_counts.get(category, 0) + 1
        print(f"Public endpoints checked: {len(endpoint_results)}")
        for category, count in sorted(endpoint_counts.items()):
            print(f"Endpoint {category}: {count}")
        for endpoint, (status, _) in sorted(endpoint_results.items()):
            if not status.startswith("Accessible"):
                print(f"Endpoint review: {status} | {endpoint}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
