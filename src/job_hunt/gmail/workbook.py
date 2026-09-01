"""Create and update the editable workbook produced by one Gmail-alert run."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from job_hunt.network.referrals import REFERRAL_COLUMNS
from job_hunt.integrations.sheets import JOB_COLUMNS


GMAIL_SHEET_NAME = "Gmail Alerts"
SUMMARY_SHEET_NAME = "Run Summary"
GMAIL_TABLE_NAME = "GmailAlertsTable"
GMAIL_RUN_COLUMNS = [*JOB_COLUMNS, *REFERRAL_COLUMNS]

EDITABLE_GMAIL_COLUMNS = {
    "company",
    "title",
    "location",
    "years_of_experience",
    "official_url",
    "application_status",
    "notes",
    "experience_min_years",
    "experience_max_years",
    "experience_fit",
    "experience_source",
}

APPLICATION_STATUSES = [
    "not_started",
    "saved",
    "reviewing",
    "shortlisted",
    "applied",
    "interviewing",
    "offer",
    "rejected",
    "withdrawn",
    "expired",
]

EXPERIENCE_FIT_STATUSES = [
    "unknown",
    "inside_target",
    "overlaps_target",
    "outside_target",
]

DATE_COLUMNS = {
    "email_received_at",
    "alert_posted_at",
    "first_seen_at",
    "last_seen_at",
}

URL_COLUMNS = {"source_url", "official_url", "referral_profile_url"}

WIDTHS = {
    "job_record_id": 27,
    "owner_id": 12,
    "alert_source": 13,
    "gmail_message_id": 22,
    "email_subject": 38,
    "email_received_at": 20,
    "company": 26,
    "title": 42,
    "location": 25,
    "years_of_experience": 22,
    "alert_posted_at": 18,
    "source_url": 48,
    "official_url": 48,
    "first_seen_at": 20,
    "last_seen_at": 20,
    "parse_confidence": 17,
    "parse_status": 25,
    "company_match": 18,
    "application_status": 20,
    "notes": 40,
    "evidence_message_ids": 30,
    "experience_min_years": 18,
    "experience_max_years": 18,
    "experience_fit": 20,
    "experience_source": 22,
    "referral_count": 15,
    "referral_name": 24,
    "referral_position": 30,
    "referral_profile_url": 44,
    "referral_match_status": 30,
    "referral_eligibility": 58,
    "referral_message": 68,
}


def _plain_value(value: Any) -> Any:
    """Return a JSON/Excel-friendly scalar without importing pandas."""

    if value is None:
        return ""
    if value.__class__.__module__.startswith(("pandas", "numpy")) and str(value) in {
        "<NA>",
        "NaT",
        "nan",
    }:
        return ""
    if hasattr(value, "item") and callable(value.item):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
    if isinstance(value, float) and value != value:
        return ""
    return value


def normalize_editor_rows(value: Any) -> list[dict[str, Any]]:
    """Normalize editable UI rows into stable job dictionaries."""

    if hasattr(value, "to_dict"):
        try:
            value = value.to_dict(orient="records")
        except TypeError:
            value = value.to_dict()
    if isinstance(value, Mapping):
        value = list(value.values())
    if not isinstance(value, (list, tuple)):
        raise ValueError("The edited Gmail table could not be read.")

    rows: list[dict[str, Any]] = []
    for item in value:
        if not isinstance(item, Mapping):
            raise ValueError("Every edited Gmail row must be a record.")
        rows.append({column: _plain_value(item.get(column, "")) for column in GMAIL_RUN_COLUMNS})
    return rows


def validate_editor_rows(
    rows: Iterable[Mapping[str, Any]],
    *,
    expected_record_ids: Iterable[str] | None = None,
) -> list[dict[str, Any]]:
    """Reject row loss, row injection, or changes to canonical record identifiers."""

    normalized = normalize_editor_rows(list(rows))
    record_ids = [str(row.get("job_record_id") or "").strip() for row in normalized]
    if any(not record_id for record_id in record_ids):
        raise ValueError("Every Gmail job must retain its job record ID.")
    if len(record_ids) != len(set(record_ids)):
        raise ValueError("The edited Gmail table contains duplicate job record IDs.")

    if expected_record_ids is not None:
        expected = {str(value) for value in expected_record_ids}
        actual = set(record_ids)
        if expected != actual:
            raise ValueError(
                "Rows cannot be added or removed from this run. Change application status "
                "or notes instead."
            )
    return normalized


def _parse_excel_date(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            return raw


def _is_public_url(value: Any) -> bool:
    parsed = urlsplit(str(value or "").strip())
    return parsed.scheme.casefold() in {"http", "https"} and bool(parsed.hostname)


def _set_cell_value(cell, column: str, value: Any) -> None:
    if column in DATE_COLUMNS:
        value = _parse_excel_date(value)
    value = _plain_value(value)
    cell.value = value
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        cell.data_type = "s"
    if column in DATE_COLUMNS and isinstance(value, (date, datetime)):
        cell.number_format = "yyyy-mm-dd hh:mm"
    elif column in {"experience_min_years", "experience_max_years"}:
        cell.number_format = "0.0"


def _summary_items(summary: Mapping[str, Any]) -> list[tuple[str, Any]]:
    preferred_order = [
        "run_id",
        "started_at",
        "finished_at",
        "status",
        "messages_read",
        "messages_supported",
        "jobs_parsed",
        "jobs_after_deduplication",
        "jobs_filtered_out",
        "parsing_warnings",
    ]
    items = [(key, summary.get(key, "")) for key in preferred_order]
    items.extend((str(key), value) for key, value in summary.items() if key not in preferred_order)
    return items


def write_gmail_run_workbook(
    output_path: Path,
    rows: Iterable[Mapping[str, Any]],
    summary: Mapping[str, Any],
    *,
    run_started_at: datetime,
) -> Path:
    """Atomically write one dated, editable Gmail-alert run workbook."""

    output_path = Path(output_path)
    normalized_rows = validate_editor_rows(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = GMAIL_SHEET_NAME
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 75
    final_column = len(GMAIL_RUN_COLUMNS)
    final_letter = get_column_letter(final_column)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_column)
    title = sheet.cell(1, 1, f"Gmail Alerts — {run_started_at.date().isoformat()}")
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0F766E")
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=final_column)
    note = sheet.cell(
        2,
        1,
        (
            "One immutable Gmail-ingestion snapshot. User corrections made in Streamlit "
            "rewrite this same file; raw email bodies and OAuth credentials are excluded."
        ),
    )
    note.font = Font(name="Aptos", size=10, italic=True, color="334155")
    note.fill = PatternFill("solid", fgColor="E2E8F0")
    note.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 38

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=final_column)
    duplicate_count = max(
        0,
        int(summary.get("jobs_parsed") or 0) - int(summary.get("jobs_after_deduplication") or 0),
    )
    status = sheet.cell(
        3,
        1,
        (
            f"Messages: {int(summary.get('messages_read') or 0):,}  |  "
            f"Parsed jobs: {int(summary.get('jobs_parsed') or 0):,}  |  "
            f"Unique jobs: {len(normalized_rows):,}  |  "
            f"Duplicates merged: {duplicate_count:,}  |  "
            f"Warnings: {int(summary.get('parsing_warnings') or 0):,}"
        ),
    )
    status.font = Font(name="Aptos", size=10, color="115E59")
    status.fill = PatternFill("solid", fgColor="CCFBF1")
    status.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 28

    for column_number, column in enumerate(GMAIL_RUN_COLUMNS, start=1):
        cell = sheet.cell(4, column_number, column)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 44

    for row_number, row in enumerate(normalized_rows, start=5):
        for column_number, column in enumerate(GMAIL_RUN_COLUMNS, start=1):
            cell = sheet.cell(row_number, column_number)
            _set_cell_value(cell, column, row.get(column, ""))
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in URL_COLUMNS and _is_public_url(cell.value):
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif column == "referral_name":
                profile_url = row.get("referral_profile_url", "")
                if cell.value and _is_public_url(profile_url):
                    cell.hyperlink = str(profile_url)
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif column == "referral_message":
                job_url = row.get("official_url") or row.get("source_url") or ""
                if cell.value and _is_public_url(job_url):
                    cell.hyperlink = str(job_url)
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_number].height = 64

    final_row = max(4, len(normalized_rows) + 4)
    table = Table(displayName=GMAIL_TABLE_NAME, ref=f"A4:{final_letter}{final_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "B5"

    for column_number, column in enumerate(GMAIL_RUN_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(column_number)].width = WIDTHS.get(column, 18)

    application_column = get_column_letter(GMAIL_RUN_COLUMNS.index("application_status") + 1)
    application_validation = DataValidation(
        type="list",
        formula1='"{0}"'.format(",".join(APPLICATION_STATUSES)),
        allow_blank=False,
    )
    sheet.add_data_validation(application_validation)
    if final_row >= 5:
        application_validation.add(f"{application_column}5:{application_column}{final_row}")

    fit_column = get_column_letter(GMAIL_RUN_COLUMNS.index("experience_fit") + 1)
    fit_validation = DataValidation(
        type="list",
        formula1='"{0}"'.format(",".join(EXPERIENCE_FIT_STATUSES)),
        allow_blank=False,
    )
    sheet.add_data_validation(fit_validation)
    if final_row >= 5:
        fit_validation.add(f"{fit_column}5:{fit_column}{final_row}")

    if final_row >= 5:
        status_range = f"{application_column}5:{application_column}{final_row}"
        for formula, color in [
            (f'{application_column}5="applied"', "DCFCE7"),
            (f'{application_column}5="shortlisted"', "FEF3C7"),
            (f'{application_column}5="rejected"', "FEE2E2"),
            (f'{application_column}5="expired"', "E5E7EB"),
        ]:
            sheet.conditional_formatting.add(
                status_range,
                FormulaRule(
                    formula=[formula],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )

    header_comments = {
        "job_record_id": "Stable identifier. Streamlit edits cannot change, add, or remove run rows.",
        "source_url": "Alert-provided job URL after privacy-safe tracking normalization.",
        "official_url": "Editable only when a public official employer URL is known.",
        "application_status": "User-maintained workflow state preserved when this run file is saved again.",
        "notes": "User-maintained review notes; do not place credentials or private message content here.",
        "referral_name": (
            "Top offline same-company candidate. Verify current employment before messaging."
        ),
        "referral_match_status": (
            "An offline snapshot match is a lead, not proof of current employment or referral willingness."
        ),
        "referral_eligibility": (
            "Preliminary resume evidence based on the alert only; official JD skills are not checked."
        ),
        "referral_message": (
            "Copy-ready LinkedIn request. Clicking the Excel cell opens the selected job URL."
        ),
    }
    for column, comment_text in header_comments.items():
        column_number = GMAIL_RUN_COLUMNS.index(column) + 1
        sheet.cell(4, column_number).comment = Comment(comment_text, "User")

    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:4"
    sheet.oddFooter.center.text = "Personal Job Hunt — Gmail Alerts"
    sheet.oddFooter.right.text = "Page &P of &N"

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells("A1:B1")
    summary_title = summary_sheet["A1"]
    summary_title.value = "Gmail Run Summary"
    summary_title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    summary_title.fill = PatternFill("solid", fgColor="0F766E")
    summary_title.alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 32
    summary_sheet.append(["Metric", "Value"])
    for cell in summary_sheet[2]:
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for key, value in _summary_items(summary):
        summary_sheet.append([key, _plain_value(value)])
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 68
    for row in summary_sheet.iter_rows(min_row=3):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    summary_sheet.freeze_panes = "A3"

    workbook.active = 0
    workbook.properties.title = "Personal Job Hunt Gmail Alert Run"
    workbook.properties.subject = "Normalized and deduplicated Gmail job alerts"
    workbook.properties.creator = "Personal Job Hunt"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    workbook.save(temporary_path)
    temporary_path.replace(output_path)
    return output_path


def read_gmail_run_workbook(
    workbook_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Read a previously generated run file for Streamlit resume/edit support."""

    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames[:2] != [GMAIL_SHEET_NAME, SUMMARY_SHEET_NAME]:
        raise ValueError("This is not a supported Gmail run workbook.")
    sheet = workbook[GMAIL_SHEET_NAME]
    headers = [sheet.cell(4, column).value for column in range(1, sheet.max_column + 1)]
    if headers == GMAIL_RUN_COLUMNS:
        source_columns = GMAIL_RUN_COLUMNS
    elif headers == JOB_COLUMNS:
        source_columns = JOB_COLUMNS
    else:
        raise ValueError("The Gmail run workbook headers do not match the stable schema.")

    rows: list[dict[str, Any]] = []
    for row_number in range(5, sheet.max_row + 1):
        values = {column: "" for column in GMAIL_RUN_COLUMNS}
        for column_number, column in enumerate(source_columns, start=1):
            value = sheet.cell(row_number, column_number).value
            if isinstance(value, (date, datetime)):
                value = value.isoformat()
            values[column] = "" if value is None else value
        if values.get("job_record_id"):
            rows.append(values)

    summary_sheet = workbook[SUMMARY_SHEET_NAME]
    summary = {
        str(summary_sheet.cell(row, 1).value): summary_sheet.cell(row, 2).value
        for row in range(3, summary_sheet.max_row + 1)
        if summary_sheet.cell(row, 1).value
    }
    return rows, summary


def verify_gmail_run_workbook(workbook_path: Path, *, expected_rows: int) -> None:
    """Verify structure, table ownership, identifiers, links, and formula safety."""

    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames != [GMAIL_SHEET_NAME, SUMMARY_SHEET_NAME]:
        raise ValueError(f"Unexpected Gmail workbook sheets: {workbook.sheetnames}")
    sheet = workbook[GMAIL_SHEET_NAME]
    expected_final_row = max(4, expected_rows + 4)
    expected_ref = f"A4:{get_column_letter(len(GMAIL_RUN_COLUMNS))}{expected_final_row}"
    if sheet.max_column != len(GMAIL_RUN_COLUMNS) or sheet.max_row != expected_final_row:
        raise ValueError("Unexpected Gmail workbook dimensions.")
    if sheet.auto_filter.ref is not None:
        raise ValueError("The Gmail table must own its filter.")
    if GMAIL_TABLE_NAME not in sheet.tables:
        raise ValueError("The Gmail alerts table is missing.")
    if sheet.tables[GMAIL_TABLE_NAME].ref != expected_ref:
        raise ValueError("Unexpected Gmail alerts table range.")
    if sheet.freeze_panes != "B5":
        raise ValueError("The Gmail sheet freeze pane is inconsistent.")

    identifiers = [
        sheet.cell(row, 1).value for row in range(5, sheet.max_row + 1) if sheet.cell(row, 1).value
    ]
    if len(identifiers) != len(set(identifiers)):
        raise ValueError("Duplicate Gmail job record IDs were written.")
    for row in range(5, sheet.max_row + 1):
        for column in URL_COLUMNS:
            cell = sheet.cell(row, GMAIL_RUN_COLUMNS.index(column) + 1)
            if cell.value and _is_public_url(cell.value) and not cell.hyperlink:
                raise ValueError(f"Missing Gmail job hyperlink at {cell.coordinate}")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError(f"Unexpected formula at {worksheet.title}!{cell.coordinate}")
