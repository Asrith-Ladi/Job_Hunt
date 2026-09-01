"""Read the public company-source tables from the canonical Excel registry."""

from __future__ import annotations

import hashlib
import shutil
import threading
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from job_hunt.discovery.detection import DetectionResult, detect_source
from job_hunt.discovery.models import SourceConfig, clean_text
from job_hunt.integrations.drive_storage import (
    EXCEL_MIME_TYPE,
    build_drive_service,
    download_drive_file,
    drive_file_url,
    ensure_job_hunt_folders,
    find_child_file,
    upload_or_update_file,
)
from job_hunt.runtime.paths import AppPaths, REGISTRY_FILE_NAME
from job_hunt.runtime.state import load_local_state, save_local_state


CATEGORY_SHEETS = (
    "MNC",
    "Product Companies",
    "Startups",
    "Mid-Sized Companies",
    "Other Companies",
)
MINIMUM_COMPANY_COUNT = 210


@dataclass(frozen=True)
class CompanyRegistryEntry:
    company_id: str
    company: str
    category: str
    sector: str
    priority: str
    careers_url: str
    portal_url: str
    source_type_label: str
    source_identifier: str
    public_feed_url: str
    api_key_required: str
    india_jobs: str
    active: str
    last_checked: str
    verification_status: str
    fallback: str
    notes: str
    detection: DetectionResult

    @property
    def adapter_ready(self) -> bool:
        return self.detection.adapter_ready and bool(self.detection.identifier)

    def to_dict(self) -> dict[str, Any]:
        value = asdict(self)
        value["adapter_ready"] = self.adapter_ready
        return value

    def to_source_config(self) -> SourceConfig:
        return SourceConfig(
            company=self.company,
            provider=self.detection.provider,
            identifier=self.detection.identifier,
            category=self.category,
            careers_url=self.careers_url,
            portal_url=self.portal_url,
            public_feed_url=self.public_feed_url,
            region=self.detection.region,
            company_id=self.company_id,
            fallback=self.fallback or self.detection.fallback,
            source_type_label=self.source_type_label,
        )


def _company_id(company: str, category: str) -> str:
    identity = f"{clean_text(company).casefold()}|{clean_text(category).casefold()}"
    return hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16]


def _text(row: dict[str, Any], key: str) -> str:
    return clean_text(row.get(key))


def load_company_registry(path: Path) -> list[CompanyRegistryEntry]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError("The canonical company registry workbook was not found.")
    workbook = load_workbook(path, data_only=True, read_only=False)
    entries: list[CompanyRegistryEntry] = []
    seen: set[str] = set()
    for sheet_name in CATEGORY_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Registry sheet is missing: {sheet_name}")
        sheet = workbook[sheet_name]
        table_names = list(sheet.tables.keys())
        if len(table_names) != 1:
            raise ValueError(f"Registry sheet must contain one table: {sheet_name}")
        table = sheet.tables[table_names[0]]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        for row_number in range(min_row + 1, max_row + 1):
            row = dict(
                zip(
                    headers,
                    [
                        sheet.cell(row_number, column).value
                        for column in range(min_col, max_col + 1)
                    ],
                )
            )
            company = _text(row, "Company")
            if not company:
                continue
            normalized = company.casefold()
            if normalized in seen:
                raise ValueError(f"Company appears in more than one registry category: {company}")
            seen.add(normalized)
            careers_url = _text(row, "Official Careers Page")
            portal_url = _text(row, "Direct Job Portal")
            feed_url = _text(row, "Public Jobs API / Feed")
            source_type = _text(row, "ATS / Source Type")
            identifier = _text(row, "Source Identifier")
            detection = detect_source(
                source_type_label=source_type,
                identifier=identifier,
                urls=(portal_url, careers_url, feed_url),
            )
            entries.append(
                CompanyRegistryEntry(
                    company_id=_company_id(company, sheet_name),
                    company=company,
                    category=sheet_name,
                    sector=_text(row, "Sector"),
                    priority=_text(row, "Priority"),
                    careers_url=careers_url,
                    portal_url=portal_url,
                    source_type_label=source_type,
                    source_identifier=identifier,
                    public_feed_url=feed_url,
                    api_key_required=_text(row, "API Key Required"),
                    india_jobs=_text(row, "India Jobs"),
                    active=_text(row, "Active"),
                    last_checked=_text(row, "Last Checked"),
                    verification_status=_text(row, "Verification Status"),
                    fallback=_text(row, "Fallback"),
                    notes=_text(row, "Notes"),
                    detection=detection,
                )
            )
    if len(entries) < MINIMUM_COMPANY_COUNT:
        workbook.close()
        raise ValueError(
            f"Expected at least {MINIMUM_COMPANY_COUNT} unique registry companies, "
            f"found {len(entries)}."
        )
    workbook.close()
    return entries


@dataclass(frozen=True)
class RegistrySnapshot:
    entries: list[CompanyRegistryEntry]
    sync_status: str
    source: str
    warning: str = ""
    drive_url: str = ""
    drive_modified_time: str = ""
    synced_at: str = ""

    def metadata(self) -> dict[str, Any]:
        return {
            "sync_status": self.sync_status,
            "source": self.source,
            "warning": self.warning,
            "drive_url": self.drive_url,
            "drive_modified_time": self.drive_modified_time,
            "synced_at": self.synced_at,
        }


class CompanyRegistryRepository:
    """Load the Drive-authoritative registry through a validated local cache."""

    def __init__(
        self,
        paths: AppPaths,
        google_connection,
        *,
        registry_loader: Callable[[Path], list[CompanyRegistryEntry]] = load_company_registry,
        drive_service_factory: Callable[[Any], Any] = build_drive_service,
    ) -> None:
        self.paths = paths
        self.google_connection = google_connection
        self.registry_loader = registry_loader
        self.drive_service_factory = drive_service_factory
        self._lock = threading.Lock()

    @staticmethod
    def _digest(path: Path) -> str:
        digest = hashlib.md5(usedforsecurity=False)
        with Path(path).open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _ensure_local_cache(self) -> None:
        cache_path = self.paths.registry_path
        if cache_path.is_file():
            return
        seed_path = self.paths.registry_seed_path
        if not seed_path.is_file():
            return
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        candidate = cache_path.with_name(f".{cache_path.stem}.seed{cache_path.suffix}")
        shutil.copyfile(seed_path, candidate)
        try:
            self.registry_loader(candidate)
            candidate.replace(cache_path)
        finally:
            candidate.unlink(missing_ok=True)

    def _fallback(self, warning: str) -> RegistrySnapshot:
        self._ensure_local_cache()
        entries = self.registry_loader(self.paths.registry_path)
        return RegistrySnapshot(
            entries=entries,
            sync_status="local_fallback",
            source="local_cache",
            warning=warning,
            synced_at=datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        )

    def load(self) -> RegistrySnapshot:
        with self._lock:
            return self._load_locked()

    def _load_locked(self) -> RegistrySnapshot:
        self._ensure_local_cache()
        try:
            credentials = self.google_connection.require_credentials()
        except (AttributeError, RuntimeError):
            return self._fallback(
                "Google Drive is unavailable; showing the last validated registry cache."
            )

        try:
            drive_service = self.drive_service_factory(credentials)
            folders = ensure_job_hunt_folders(drive_service)
            source_folder_id = str(folders["source"]["id"])
            remote = find_child_file(
                drive_service,
                REGISTRY_FILE_NAME,
                parent_id=source_folder_id,
                mime_type=EXCEL_MIME_TYPE,
            )
        except Exception:
            return self._fallback(
                "Drive registry refresh failed; showing the last validated registry cache."
            )

        state = load_local_state(self.paths.app_state_path)
        source_ids = dict(state.get("drive_source_file_ids") or {})
        revisions = dict(state.get("drive_source_revisions") or {})
        synced_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

        if remote is None:
            if not self.paths.registry_path.is_file():
                raise FileNotFoundError(
                    "The Drive registry and its local bootstrap seed are unavailable."
                )
            entries = self.registry_loader(self.paths.registry_path)
            created = upload_or_update_file(
                drive_service,
                self.paths.registry_path,
                parent_id=source_folder_id,
            )
            remote = created
            status = "drive_seeded"
        else:
            remote_id = str(remote.get("id") or "")
            remote_md5 = str(remote.get("md5Checksum") or "")
            remote_revision = {
                "file_id": remote_id,
                "modified_time": str(remote.get("modifiedTime") or ""),
                "md5_checksum": remote_md5,
            }
            stored_revision = revisions.get(REGISTRY_FILE_NAME)
            local_matches = self.paths.registry_path.is_file() and (
                (remote_md5 and self._digest(self.paths.registry_path) == remote_md5)
                or (not remote_md5 and stored_revision == remote_revision)
            )
            if local_matches:
                entries = self.registry_loader(self.paths.registry_path)
                status = "drive_current"
            else:
                candidate = self.paths.registry_path.with_name(
                    f".{self.paths.registry_path.stem}.drive{self.paths.registry_path.suffix}"
                )
                try:
                    download_drive_file(drive_service, remote_id, candidate)
                    entries = self.registry_loader(candidate)
                    self.paths.registry_path.parent.mkdir(parents=True, exist_ok=True)
                    candidate.replace(self.paths.registry_path)
                    status = "drive_refreshed"
                except Exception:
                    candidate.unlink(missing_ok=True)
                    return self._fallback(
                        "The Drive registry failed download or validation; showing the last "
                        "validated registry cache."
                    )

        remote_id = str(remote.get("id") or "")
        source_ids[REGISTRY_FILE_NAME] = remote_id
        revisions[REGISTRY_FILE_NAME] = {
            "file_id": remote_id,
            "modified_time": str(remote.get("modifiedTime") or ""),
            "md5_checksum": str(remote.get("md5Checksum") or ""),
        }
        state["drive_source_file_ids"] = source_ids
        state["drive_source_revisions"] = revisions
        save_local_state(self.paths.app_state_path, state)
        return RegistrySnapshot(
            entries=entries,
            sync_status=status,
            source="google_drive",
            drive_url=str(remote.get("webViewLink") or drive_file_url(remote_id)),
            drive_modified_time=str(remote.get("modifiedTime") or ""),
            synced_at=synced_at,
        )
