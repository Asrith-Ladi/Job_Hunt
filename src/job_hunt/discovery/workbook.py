"""Verified dated Excel artifacts for portal and ATS discovery runs."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


JOBS_SHEET_NAME = "Jobs"
CHECKS_SHEET_NAME = "Source Checks"
SUMMARY_SHEET_NAME = "Run Summary"
JOBS_TABLE_NAME = "DiscoveryJobsTable"
CHECKS_TABLE_NAME = "DiscoverySourceChecksTable"

DISCOVERY_JOB_COLUMNS = [
    "job_record_id",
    "company",
    "title",
    "location",
    "provider",
    "source_identifier",
    "source_type",
    "external_job_id",
    "official_url",
    "apply_url",
    "source_url",
    "description",
    "department",
    "employment_type",
    "workplace_type",
    "experience_text",
    "experience_min_years",
    "experience_max_years",
    "experience_fit",
    "posted_at",
    "updated_at",
    "date_provenance",
    "discovered_at",
    "first_seen_at",
    "last_seen_at",
    "source_confidence",
    "source_status",
    "run_change_status",
    "application_status",
    "notes",
]
LEGACY_DISCOVERY_JOB_COLUMNS = [
    column for column in DISCOVERY_JOB_COLUMNS if column != "run_change_status"
]

SOURCE_CHECK_COLUMNS = [
    "company",
    "category",
    "provider",
    "source_identifier",
    "strategy",
    "source_url",
    "status",
    "jobs_found",
    "jobs_exported",
    "warning",
    "fallback",
    "checked_at",
]

EDITABLE_DISCOVERY_COLUMNS = {"application_status", "notes"}
APPLICATION_STATUSES = [
    "not_started",
    "saved",
    "reviewing",
    "shortlisted",
    "applied",
    "interviewing",
    "offer",
    "rejected",
    "withdrawn",
    "expired",
]
DATE_COLUMNS = {"posted_at", "updated_at", "discovered_at", "first_seen_at", "last_seen_at"}
URL_COLUMNS = {"official_url", "apply_url", "source_url"}

JOB_WIDTHS = {
    "job_record_id": 27,
    "company": 25,
    "title": 42,
    "location": 26,
    "provider": 18,
    "source_identifier": 24,
    "source_type": 24,
    "external_job_id": 24,
    "official_url": 48,
    "apply_url": 42,
    "source_url": 48,
    "description": 72,
    "department": 26,
    "employment_type": 20,
    "workplace_type": 18,
    "experience_text": 18,
    "experience_min_years": 18,
    "experience_max_years": 18,
    "experience_fit": 20,
    "posted_at": 20,
    "updated_at": 20,
    "date_provenance": 25,
    "discovered_at": 20,
    "first_seen_at": 20,
    "last_seen_at": 20,
    "source_confidence": 18,
    "source_status": 20,
    "run_change_status": 20,
    "application_status": 20,
    "notes": 42,
}

CHECK_WIDTHS = {
    "company": 28,
    "category": 24,
    "provider": 18,
    "source_identifier": 24,
    "strategy": 22,
    "source_url": 50,
    "status": 22,
    "jobs_found": 14,
    "jobs_exported": 14,
    "warning": 52,
    "fallback": 45,
    "checked_at": 20,
}


def _plain_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    if hasattr(value, "item") and callable(value.item):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
    return value


def _is_public_url(value: Any) -> bool:
    try:
        parsed = urlsplit(str(value or "").strip())
    except ValueError:
        return False
    return parsed.scheme.casefold() == "https" and bool(parsed.hostname)


def _parse_date(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        try:
            return date.fromisoformat(raw)
        except ValueError:
            return raw


def _set_cell(cell, column: str, value: Any) -> None:
    value = _parse_date(value) if column in DATE_COLUMNS or column == "checked_at" else value
    value = _plain_value(value)
    cell.value = value
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        cell.data_type = "s"
    if isinstance(value, (date, datetime)):
        cell.number_format = "yyyy-mm-dd hh:mm"
    elif column in {"experience_min_years", "experience_max_years"}:
        cell.number_format = "0.0"
    elif column in {"jobs_found", "jobs_exported"}:
        cell.number_format = "#,##0"


def _normalize_records(value: Any, columns: list[str]) -> list[dict[str, Any]]:
    if hasattr(value, "to_dict"):
        try:
            value = value.to_dict(orient="records")
        except TypeError:
            value = value.to_dict()
    if isinstance(value, Mapping):
        value = list(value.values())
    if not isinstance(value, (list, tuple)):
        raise ValueError("The discovery table could not be read.")
    output: list[dict[str, Any]] = []
    for item in value:
        if not isinstance(item, Mapping):
            raise ValueError("Every discovery row must be a record.")
        output.append({column: _plain_value(item.get(column, "")) for column in columns})
    return output


def validate_discovery_rows(
    rows: Iterable[Mapping[str, Any]],
    *,
    expected_rows: Iterable[Mapping[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    normalized = _normalize_records(list(rows), DISCOVERY_JOB_COLUMNS)
    identifiers = [str(row.get("job_record_id") or "").strip() for row in normalized]
    if any(not value for value in identifiers):
        raise ValueError("Every discovered job must retain its record ID.")
    if len(identifiers) != len(set(identifiers)):
        raise ValueError("The edited discovery table contains duplicate record IDs.")
    if expected_rows is not None:
        expected = {str(row.get("job_record_id") or ""): dict(row) for row in expected_rows}
        if set(identifiers) != set(expected):
            raise ValueError("Rows cannot be added to or removed from a discovery run.")
        for row in normalized:
            baseline = expected[str(row["job_record_id"])]
            for column in DISCOVERY_JOB_COLUMNS:
                if column in EDITABLE_DISCOVERY_COLUMNS:
                    continue
                if str(_plain_value(row.get(column, ""))) != str(
                    _plain_value(baseline.get(column, ""))
                ):
                    raise ValueError(f"The protected field '{column}' cannot be changed.")
    for row in normalized:
        if row["application_status"] not in APPLICATION_STATUSES:
            raise ValueError("An unsupported application status was supplied.")
    return normalized


def normalize_source_checks(value: Any) -> list[dict[str, Any]]:
    return _normalize_records(value, SOURCE_CHECK_COLUMNS)


def _title_rows(sheet, title: str, note: str, status: str, final_column: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_column)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0F766E")
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=final_column)
    cell = sheet.cell(2, 1, note)
    cell.font = Font(name="Aptos", size=10, italic=True, color="334155")
    cell.fill = PatternFill("solid", fgColor="E2E8F0")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[2].height = 38

    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=final_column)
    cell = sheet.cell(3, 1, status)
    cell.font = Font(name="Aptos", size=10, color="115E59")
    cell.fill = PatternFill("solid", fgColor="CCFBF1")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 28


def _write_table_sheet(
    sheet,
    *,
    columns: list[str],
    rows: list[dict[str, Any]],
    widths: Mapping[str, int],
    table_name: str,
) -> None:
    for column_number, column in enumerate(columns, start=1):
        cell = sheet.cell(4, column_number, column)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 44
    for row_number, row in enumerate(rows, start=5):
        for column_number, column in enumerate(columns, start=1):
            cell = sheet.cell(row_number, column_number)
            _set_cell(cell, column, row.get(column, ""))
            cell.font = Font(name="Aptos", size=9, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in URL_COLUMNS and _is_public_url(cell.value):
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_number].height = 72
    final_row = max(4, len(rows) + 4)
    final_letter = get_column_letter(len(columns))
    table = Table(displayName=table_name, ref=f"A4:{final_letter}{final_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "B5"
    for number, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(number)].width = widths.get(column, 18)


def write_discovery_workbook(
    output_path: Path,
    *,
    mode: str,
    rows: Iterable[Mapping[str, Any]],
    source_checks: Iterable[Mapping[str, Any]],
    summary: Mapping[str, Any],
    run_started_at: datetime,
) -> Path:
    normalized_rows = validate_discovery_rows(rows)
    normalized_checks = normalize_source_checks(list(source_checks))
    workbook = Workbook()
    jobs = workbook.active
    jobs.title = JOBS_SHEET_NAME
    jobs.sheet_view.showGridLines = False
    jobs.sheet_view.zoomScale = 75
    label = "Company Portals" if mode == "company_portals" else "ATS Sources"
    _title_rows(
        jobs,
        f"{label} — {run_started_at.date().isoformat()}",
        (
            "Current official/public matches. Run change status distinguishes new, changed, and "
            "previously seen jobs; only application status and notes are editable."
        ),
        (
            f"Sources checked: {int(summary.get('sources_checked') or 0):,}  |  "
            f"Current matches: {len(normalized_rows):,}  |  "
            f"New/changed: {int(summary.get('jobs_new_or_changed_this_run') or 0):,}  |  "
            f"Previously seen: {int(summary.get('jobs_unchanged_from_prior_runs') or 0):,}  |  "
            f"Warnings: {int(summary.get('warnings') or 0):,}"
        ),
        len(DISCOVERY_JOB_COLUMNS),
    )
    _write_table_sheet(
        jobs,
        columns=DISCOVERY_JOB_COLUMNS,
        rows=normalized_rows,
        widths=JOB_WIDTHS,
        table_name=JOBS_TABLE_NAME,
    )

    status_column = get_column_letter(DISCOVERY_JOB_COLUMNS.index("application_status") + 1)
    validation = DataValidation(
        type="list",
        formula1='"{0}"'.format(",".join(APPLICATION_STATUSES)),
        allow_blank=False,
    )
    jobs.add_data_validation(validation)
    final_job_row = len(normalized_rows) + 4
    if final_job_row >= 5:
        validation.add(f"{status_column}5:{status_column}{final_job_row}")
        status_range = f"{status_column}5:{status_column}{final_job_row}"
        for formula, color in [
            (f'{status_column}5="applied"', "DCFCE7"),
            (f'{status_column}5="shortlisted"', "FEF3C7"),
            (f'{status_column}5="rejected"', "FEE2E2"),
            (f'{status_column}5="expired"', "E5E7EB"),
        ]:
            jobs.conditional_formatting.add(
                status_range,
                FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
            )
        change_column = get_column_letter(
            DISCOVERY_JOB_COLUMNS.index("run_change_status") + 1
        )
        change_range = f"{change_column}5:{change_column}{final_job_row}"
        for formula, color in [
            (f'{change_column}5="new"', "DCFCE7"),
            (f'{change_column}5="changed"', "FEF3C7"),
            (f'{change_column}5="previously_seen"', "E5E7EB"),
        ]:
            jobs.conditional_formatting.add(
                change_range,
                FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
            )
    jobs.cell(4, DISCOVERY_JOB_COLUMNS.index("job_record_id") + 1).comment = Comment(
        "Stable protected identifier. Browser edits cannot change, add, or remove run rows.",
        "User",
    )
    jobs.cell(4, DISCOVERY_JOB_COLUMNS.index("date_provenance") + 1).comment = Comment(
        "Explains whether posted_at is provider-supplied or unknown. Sitemap lastmod is not publication time.",
        "User",
    )
    jobs.cell(4, DISCOVERY_JOB_COLUMNS.index("run_change_status") + 1).comment = Comment(
        "New and changed jobs are distinguished from currently active jobs already seen in an earlier run.",
        "User",
    )

    checks = workbook.create_sheet(CHECKS_SHEET_NAME)
    checks.sheet_view.showGridLines = False
    _title_rows(
        checks,
        "Source Checks",
        "One auditable result per selected company/source. Access blocks stop safely and retain a fallback.",
        f"Attempts: {len(normalized_checks):,}",
        len(SOURCE_CHECK_COLUMNS),
    )
    _write_table_sheet(
        checks,
        columns=SOURCE_CHECK_COLUMNS,
        rows=normalized_checks,
        widths=CHECK_WIDTHS,
        table_name=CHECKS_TABLE_NAME,
    )

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells("A1:B1")
    summary_sheet["A1"] = f"{label} Run Summary"
    summary_sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
    summary_sheet["A1"].alignment = Alignment(vertical="center")
    summary_sheet.row_dimensions[1].height = 32
    summary_sheet.append(["Metric", "Value"])
    for cell in summary_sheet[2]:
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for row_number, (key, value) in enumerate(summary.items(), start=3):
        _set_cell(summary_sheet.cell(row_number, 1), "metric", str(key))
        _set_cell(summary_sheet.cell(row_number, 2), "summary_value", value)
    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 72
    for row in summary_sheet.iter_rows(min_row=3):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color="1F2937")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    summary_sheet.freeze_panes = "A3"

    workbook.active = 0
    workbook.properties.title = f"Personal Job Hunt {label} Run"
    workbook.properties.creator = "Personal Job Hunt"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    for sheet in workbook.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(f"{output_path.stem}.tmp{output_path.suffix}")
    workbook.save(temporary)
    temporary.replace(output_path)
    return output_path


def read_discovery_workbook(
    workbook_path: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames != [JOBS_SHEET_NAME, CHECKS_SHEET_NAME, SUMMARY_SHEET_NAME]:
        raise ValueError("This is not a supported discovery run workbook.")

    def read_table(sheet_name: str, columns: list[str]) -> list[dict[str, Any]]:
        sheet = workbook[sheet_name]
        headers = [sheet.cell(4, number).value for number in range(1, len(columns) + 1)]
        active_columns = columns
        legacy_jobs = False
        if headers != columns and sheet_name == JOBS_SHEET_NAME:
            legacy_headers = [
                sheet.cell(4, number).value
                for number in range(1, len(LEGACY_DISCOVERY_JOB_COLUMNS) + 1)
            ]
            if legacy_headers == LEGACY_DISCOVERY_JOB_COLUMNS:
                active_columns = LEGACY_DISCOVERY_JOB_COLUMNS
                legacy_jobs = True
        if headers != columns and not legacy_jobs:
            raise ValueError(f"The {sheet_name} headers do not match the stable schema.")
        values: list[dict[str, Any]] = []
        for row_number in range(5, sheet.max_row + 1):
            row: dict[str, Any] = {}
            for number, column in enumerate(active_columns, start=1):
                value = sheet.cell(row_number, number).value
                if isinstance(value, (date, datetime)):
                    value = value.isoformat()
                row[column] = "" if value is None else value
            if any(row.values()):
                if legacy_jobs:
                    row["run_change_status"] = "new_or_changed"
                values.append(row)
        return values

    rows = [
        row for row in read_table(JOBS_SHEET_NAME, DISCOVERY_JOB_COLUMNS) if row["job_record_id"]
    ]
    checks = read_table(CHECKS_SHEET_NAME, SOURCE_CHECK_COLUMNS)
    summary_sheet = workbook[SUMMARY_SHEET_NAME]
    summary = {
        str(summary_sheet.cell(row, 1).value): summary_sheet.cell(row, 2).value
        for row in range(3, summary_sheet.max_row + 1)
        if summary_sheet.cell(row, 1).value
    }
    return rows, checks, summary


def verify_discovery_workbook(
    workbook_path: Path,
    *,
    expected_jobs: int,
    expected_checks: int,
) -> None:
    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames != [JOBS_SHEET_NAME, CHECKS_SHEET_NAME, SUMMARY_SHEET_NAME]:
        raise ValueError(f"Unexpected discovery workbook sheets: {workbook.sheetnames}")
    expectations = [
        (JOBS_SHEET_NAME, DISCOVERY_JOB_COLUMNS, JOBS_TABLE_NAME, expected_jobs),
        (CHECKS_SHEET_NAME, SOURCE_CHECK_COLUMNS, CHECKS_TABLE_NAME, expected_checks),
    ]
    for sheet_name, columns, table_name, row_count in expectations:
        sheet = workbook[sheet_name]
        final_row = max(4, row_count + 4)
        expected_ref = f"A4:{get_column_letter(len(columns))}{final_row}"
        if sheet.freeze_panes != "B5" or sheet.auto_filter.ref is not None:
            raise ValueError(f"The {sheet_name} navigation/filter structure is invalid.")
        if table_name not in sheet.tables or sheet.tables[table_name].ref != expected_ref:
            raise ValueError(f"The {sheet_name} table range is invalid.")
    jobs = workbook[JOBS_SHEET_NAME]
    identifiers = [
        jobs.cell(row, 1).value for row in range(5, jobs.max_row + 1) if jobs.cell(row, 1).value
    ]
    if len(identifiers) != len(set(identifiers)):
        raise ValueError("Duplicate discovery record IDs were written.")
    for row in range(5, jobs.max_row + 1):
        for column in URL_COLUMNS:
            cell = jobs.cell(row, DISCOVERY_JOB_COLUMNS.index(column) + 1)
            if cell.value and _is_public_url(cell.value) and not cell.hyperlink:
                raise ValueError(f"Missing hyperlink at {JOBS_SHEET_NAME}!{cell.coordinate}")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError(f"Unexpected formula at {sheet.title}!{cell.coordinate}")
