"""Offline referral suggestions used by Gmail jobs and network review.

The enrichment deliberately uses only the non-contact fields in the saved
LinkedIn export snapshot.  A company match is a lead to verify, not proof that
the person still works there or will provide a referral.
"""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping
from urllib.parse import urlsplit

from openpyxl import load_workbook

from job_hunt.jobs.enrichment import (
    Connection,
    canonical_company,
    cold_referral_message,
    connection_relevance,
    normalize_text,
    personal_resume_profile,
)


CONNECTIONS_SHEET_NAME = "LinkedIn Connections"
REFERRAL_COLUMNS = [
    "referral_count",
    "referral_name",
    "referral_position",
    "referral_profile_url",
    "referral_match_status",
    "referral_eligibility",
    "referral_message",
]
REFERRAL_CANDIDATES_FIELD = "referral_candidates"

_REQUIRED_HEADERS = {
    "Connection Name",
    "Current Company",
    "Current Position",
    "Registry Company",
    "LinkedIn Profile",
    "Connected On",
}


@dataclass(frozen=True)
class RegistryConnectionRecord:
    """One contact-free row from the canonical LinkedIn connection snapshot."""

    source_row_number: int
    full_name: str
    first_name: str
    last_name: str
    current_company: str
    current_position: str
    registry_company: str
    registry_category: str
    referral_status: str
    email_address: str
    linkedin_url: str
    connected_on: str
    match_method: str
    official_careers_page: str
    direct_job_portal: str

    @property
    def company(self) -> str:
        return self.registry_company or self.current_company

    def as_connection(self) -> Connection:
        return Connection(
            first_name=self.first_name,
            last_name=self.last_name,
            linkedin_url=self.linkedin_url,
            company=self.company,
            position=self.current_position,
            connected_on=self.connected_on,
        )


def _valid_linkedin_profile(value: object) -> bool:
    parsed = urlsplit(str(value or "").strip())
    hostname = (parsed.hostname or "").casefold()
    return (
        parsed.scheme.casefold() in {"http", "https"}
        and (hostname == "linkedin.com" or hostname.endswith(".linkedin.com"))
        and parsed.path.casefold().startswith("/in/")
    )


def _name_parts(value: object) -> tuple[str, str]:
    parts = str(value or "").strip().split()
    if not parts:
        return "", ""
    return parts[0], " ".join(parts[1:])


def _as_text(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value or "").strip()


def load_registry_connection_records(
    registry_path: Path,
    *,
    include_email: bool = False,
) -> list[RegistryConnectionRecord]:
    """Load every named snapshot row, with email only for explicit private-UI use."""

    registry_path = Path(registry_path)
    if not registry_path.is_file():
        raise FileNotFoundError("The company registry workbook is unavailable.")

    workbook = load_workbook(
        registry_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if CONNECTIONS_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("The LinkedIn Connections registry tab is unavailable.")
        sheet = workbook[CONNECTIONS_SHEET_NAME]

        # Some Excel/Drive rewrites omit the worksheet ``dimension`` cache.  In
        # openpyxl read-only mode that leaves max_row/max_column as None even
        # though the sheet contains valid rows.  Calculate the bounds from the
        # XML stream before using them; this is read-only and does not rewrite
        # the authoritative registry workbook.
        if sheet.max_row is None or sheet.max_column is None:
            sheet.calculate_dimension(force=True)
        max_row = int(sheet.max_row or 0)
        max_column = int(sheet.max_column or 0)
        if not max_row or not max_column:
            raise ValueError("The LinkedIn Connections registry tab is empty.")

        header_row = 0
        header_indexes: dict[str, int] = {}
        for row_number in range(1, min(max_row, 10) + 1):
            values = [
                sheet.cell(row_number, column).value for column in range(1, max_column + 1)
            ]
            candidate = {
                str(value).strip(): index for index, value in enumerate(values) if value is not None
            }
            if _REQUIRED_HEADERS.issubset(candidate):
                header_row = row_number
                header_indexes = candidate
                break
        if not header_row:
            raise ValueError("The LinkedIn Connections registry headers are unsupported.")

        records: list[RegistryConnectionRecord] = []
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=header_row + 1),
            start=header_row + 1,
        ):

            def allowed_value(header: str) -> object:
                index = header_indexes.get(header)
                return cells[index].value if index is not None else None

            full_name = _as_text(allowed_value("Connection Name"))
            first_name, last_name = _name_parts(full_name)
            if not first_name:
                continue
            profile_url = _as_text(allowed_value("LinkedIn Profile"))
            records.append(
                RegistryConnectionRecord(
                    source_row_number=row_number,
                    full_name=full_name,
                    first_name=first_name,
                    last_name=last_name,
                    current_company=_as_text(allowed_value("Current Company")),
                    current_position=_as_text(allowed_value("Current Position")),
                    registry_company=_as_text(allowed_value("Registry Company")),
                    registry_category=_as_text(allowed_value("Registry Category")),
                    referral_status=_as_text(allowed_value("Referral Status")),
                    email_address=(
                        _as_text(allowed_value("Email Address")) if include_email else ""
                    ),
                    linkedin_url=profile_url if _valid_linkedin_profile(profile_url) else "",
                    connected_on=_as_text(allowed_value("Connected On")),
                    match_method=_as_text(allowed_value("Match Method")),
                    official_careers_page=_as_text(allowed_value("Official Careers Page")),
                    direct_job_portal=_as_text(allowed_value("Direct Job Portal")),
                )
            )
        return records
    finally:
        workbook.close()


def load_registry_connection_profiles(registry_path: Path) -> list[Connection]:
    """Load saved LinkedIn profiles suitable for same-company referral matching."""

    return [
        record.as_connection()
        for record in load_registry_connection_records(registry_path)
        if record.linkedin_url
    ]


def load_registry_connections(registry_path: Path) -> list[Connection]:
    """Load company-tagged profiles suitable for same-company referral matching."""

    return [
        connection
        for connection in load_registry_connection_profiles(registry_path)
        if connection.company
    ]


def _connection_index(connections: Iterable[Connection]) -> dict[str, list[Connection]]:
    indexed: dict[str, list[Connection]] = defaultdict(list)
    for connection in connections:
        key = normalize_text(canonical_company(connection.company))
        if key:
            indexed[key].append(connection)
    for key, matches in indexed.items():
        matches.sort(
            key=lambda item: (
                -connection_relevance(item),
                normalize_text(item.position),
                normalize_text(item.full_name),
            )
        )
        seen_profiles: set[str] = set()
        unique_matches: list[Connection] = []
        for item in matches:
            profile_key = normalize_text(item.linkedin_url)
            if profile_key in seen_profiles:
                continue
            seen_profiles.add(profile_key)
            unique_matches.append(item)
        indexed[key] = unique_matches
    return dict(indexed)


def _role_strengths(title: object) -> list[str]:
    normalized = normalize_text(title)
    if re.search(
        r"\b(gen ai|generative ai|machine learning|ml|artificial intelligence|ai)\b", normalized
    ):
        return ["production AI/ML and GenAI/RAG systems", "Python delivery on AWS/cloud"]
    if re.search(r"\b(data engineer|analytics engineer|etl|data platform)\b", normalized):
        return ["Python/SQL data engineering", "AWS/cloud production delivery"]
    if re.search(r"\b(data scientist|statistics|nlp|deep learning)\b", normalized):
        return ["production machine learning and statistics", "Python and SQL"]
    if re.search(r"\b(software|backend|api|platform|developer)\b", normalized):
        return ["Python/REST API engineering", "distributed systems on AWS/cloud"]
    return ["production Python engineering", "machine learning on AWS/cloud"]


def _number(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _experience_note(row: Mapping[str, Any]) -> str:
    profile = personal_resume_profile()
    minimum = _number(row.get("experience_min_years"))
    maximum = _number(row.get("experience_max_years"))
    stated = str(row.get("years_of_experience") or "").strip()
    if minimum is None and maximum is None:
        return "The alert does not state a numeric experience requirement, so the role level still needs verification."
    inside_minimum = minimum is None or profile.years_experience >= minimum
    inside_maximum = maximum is None or profile.years_experience <= maximum
    range_text = stated or "the stated experience range"
    if inside_minimum and inside_maximum:
        return f"My documented {profile.years_experience:g} years is within {range_text}."
    return (
        f"The alert lists {range_text}; my documented experience is "
        f"{profile.years_experience:g} years, so I would verify the level first."
    )


def _eligibility_summary(row: Mapping[str, Any], strengths: list[str]) -> str:
    profile = personal_resume_profile()
    return (
        f"Preliminary alert-only fit: {profile.years_experience:g} years documented; "
        f"relevant evidence includes {strengths[0]} and {strengths[1]}. "
        f"{_experience_note(row)} Official JD requirements have not been checked."
    )


def _blank_referral_fields(status: str) -> dict[str, Any]:
    return {
        "referral_count": 0,
        "referral_name": "",
        "referral_position": "",
        "referral_profile_url": "",
        "referral_match_status": status,
        "referral_eligibility": "",
        "referral_message": "",
        REFERRAL_CANDIDATES_FIELD: [],
    }


def enrich_gmail_referrals(
    rows: Iterable[Mapping[str, Any]],
    registry_path: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Add deterministic offline referral suggestions to Gmail-alert rows."""

    records = [dict(row) for row in rows]
    try:
        connections = load_registry_connections(registry_path)
        indexed = _connection_index(connections)
        availability = "available"
    except (FileNotFoundError, OSError, ValueError, KeyError):
        connections = []
        indexed = {}
        availability = "connections_unavailable"

    jobs_with_candidate = 0
    enriched: list[dict[str, Any]] = []
    for record in records:
        company = str(record.get("company") or "").strip()
        if not company:
            record.update(_blank_referral_fields("company_unavailable"))
            enriched.append(record)
            continue
        if availability != "available":
            record.update(_blank_referral_fields(availability))
            enriched.append(record)
            continue

        key = normalize_text(canonical_company(company))
        matches = indexed.get(key, [])
        if not matches:
            record.update(_blank_referral_fields("no_offline_company_match"))
            enriched.append(record)
            continue

        strengths = _role_strengths(record.get("title"))
        experience_note = _experience_note(record)
        job_url = str(record.get("official_url") or record.get("source_url") or "").strip()
        referral_candidates = [
            {
                "name": connection.full_name,
                "position": connection.position,
                "profile_url": connection.linkedin_url,
                "message": cold_referral_message(
                    connection,
                    company,
                    str(record.get("title") or "this job").strip(),
                    job_url,
                    strengths,
                    experience_note=experience_note,
                ),
            }
            for connection in matches
        ]
        connection = matches[0]
        record.update(
            {
                "referral_count": len(referral_candidates),
                "referral_name": connection.full_name,
                "referral_position": connection.position,
                "referral_profile_url": connection.linkedin_url,
                "referral_match_status": "offline_company_match_unverified",
                "referral_eligibility": _eligibility_summary(record, strengths),
                "referral_message": referral_candidates[0]["message"],
                REFERRAL_CANDIDATES_FIELD: referral_candidates,
            }
        )
        jobs_with_candidate += 1
        enriched.append(record)

    return enriched, {
        "referral_enrichment_status": availability,
        "offline_connections_loaded": len(connections),
        "jobs_with_referral_candidate": jobs_with_candidate,
    }
